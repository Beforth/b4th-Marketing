from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from marketing_app.hrms_rbac import hrms_login_required
from marketing_app.permissions import (
    require_permission, require_any_permission, check_permission,
    MARKETING_PERMISSIONS
)
from marketing_app.user_utils import get_django_user
from marketing_app.user_helpers import get_user_info_dict, set_user_info_on_model
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import (
    Q,
    Count,
    Sum,
    Max,
    Value,
    ExpressionWrapper,
    F,
    FloatField,
    IntegerField,
)
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import calendar
from .models import Campaign, Lead, EmailTemplate, CampaignMetric, LeadActivity, Customer, CustomerLocation, Region, Visit, VisitParticipant, Expense, Exhibition, Quotation, PurchaseOrder, PaymentFollowUp, WorkOrder, Manufacturing, Dispatch, URS, GADrawing, TechnicalDiscussion, Negotiation, QuotationRevision, QCTracking, ProductionPlan, PackingDetails, DispatchChecklist, BudgetCategory, AnnualExhibitionBudget, BudgetAllocation, BudgetApproval, InquiryLog, FollowUpStatus, ProjectToday, OrderExpectedNextMonth, MISPurchaseOrder, NewData, NewDataDetails, ODPlan, ODPlanVisitReport, ODPlanRemarks, PODetails, POStatus, WorkOrderFormat, WeeklySummary, CallingDetails, HotOrders, PendingPayment2024, PendingPayment2025, OrderLoss, DSR
from django.contrib.auth import get_user_model
import sys

User = get_user_model()

@hrms_login_required
@require_permission(MARKETING_PERMISSIONS['campaign.view'])
def marketing_dashboard(request):
    """Marketing Dashboard View"""
    from marketing_app.permission_filters import (
        filter_campaigns_by_permission, filter_leads_by_permission,
        filter_customers_by_permission, can_view_reports
    )
    
    # Get current date and calculate date ranges
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # Filter data based on permissions
    customers_qs = Customer.objects.all()
    customers_qs = filter_customers_by_permission(request, customers_qs)
    total_customers = customers_qs.count()
    active_leads = Lead.objects.filter(status__in=['new', 'contacted', 'qualified']).count()
    pending_quotations = Quotation.objects.filter(status='sent').count()
    production_orders = Manufacturing.objects.filter(status__in=['started', 'in_progress', 'qc_started']).count()
    
    # Recent activities
    recent_activities = []
    
    # Recent POs
    recent_pos = PurchaseOrder.objects.select_related('customer').order_by('-created_at')[:3]
    for po in recent_pos:
        recent_activities.append({
            'type': 'po_received',
            'title': f'PO Received from {po.customer.name}',
            'description': f'Purchase Order #{po.po_number} for ₹{po.total_amount}',
            'status': 'Completed',
            'time': po.created_at,
            'icon': 'check-circle',
            'color': 'green'
        })
    
    # Recent visits
    recent_visits = Visit.objects.select_related('customer').order_by('-created_at')[:3]
    for visit in recent_visits:
        recent_activities.append({
            'type': 'visit_completed',
            'title': f'Visit Completed - {visit.customer.name}',
            'description': f'{visit.get_visit_type_display()} completed',
            'status': 'In Progress',
            'time': visit.created_at,
            'icon': 'map-pin',
            'color': 'blue'
        })
    
    # Recent production
    recent_manufacturing = Manufacturing.objects.select_related('work_order__purchase_order__customer').order_by('-created_at')[:3]
    for mfg in recent_manufacturing:
        recent_activities.append({
            'type': 'production_started',
            'title': f'Production Started - Batch #{mfg.batch_number}',
            'description': f'Manufacturing started for {mfg.work_order.purchase_order.customer.name}',
            'status': 'Started',
            'time': mfg.created_at,
            'icon': 'factory',
            'color': 'orange'
        })
    
    # Sort activities by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    
    # Sales Pipeline Data (This Month)
    first_day_of_month = today.replace(day=1)
    new_leads_this_month = Lead.objects.filter(created_at__date__gte=first_day_of_month).count()
    new_leads_last_month = Lead.objects.filter(
        created_at__date__gte=first_day_of_month - timedelta(days=30),
        created_at__date__lt=first_day_of_month
    ).count()
    leads_growth = ((new_leads_this_month - new_leads_last_month) / new_leads_last_month * 100) if new_leads_last_month > 0 else 0
    
    quotations_sent_this_month = Quotation.objects.filter(
        sent_date__date__gte=first_day_of_month,
        sent_date__isnull=False
    ).count()
    quotations_sent_last_month = Quotation.objects.filter(
        sent_date__date__gte=first_day_of_month - timedelta(days=30),
        sent_date__date__lt=first_day_of_month,
        sent_date__isnull=False
    ).count()
    quotations_growth = ((quotations_sent_this_month - quotations_sent_last_month) / quotations_sent_last_month * 100) if quotations_sent_last_month > 0 else 0
    
    pos_this_month = PurchaseOrder.objects.filter(created_at__date__gte=first_day_of_month).count()
    pos_last_month = PurchaseOrder.objects.filter(
        created_at__date__gte=first_day_of_month - timedelta(days=30),
        created_at__date__lt=first_day_of_month
    ).count()
    pos_growth = ((pos_this_month - pos_last_month) / pos_last_month * 100) if pos_last_month > 0 else 0
    
    deliveries_this_month = Dispatch.objects.filter(
        dispatch_date__date__gte=first_day_of_month,
        dispatch_date__isnull=False
    ).count()
    deliveries_last_month = Dispatch.objects.filter(
        dispatch_date__date__gte=first_day_of_month - timedelta(days=30),
        dispatch_date__date__lt=first_day_of_month,
        dispatch_date__isnull=False
    ).count()
    deliveries_growth = ((deliveries_this_month - deliveries_last_month) / deliveries_last_month * 100) if deliveries_last_month > 0 else 0
    
    # Regional Performance Data
    regional_performance = []
    regions = Region.objects.all()
    for region in regions:
        region_customers = customers_qs.filter(region=region)
        region_pos = PurchaseOrder.objects.filter(customer__region=region)
        
        # Calculate total sales for this month
        region_sales_this_month = region_pos.filter(
            created_at__date__gte=first_day_of_month
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Get monthly target (if set)
        monthly_target = region.monthly_target or Decimal('0')
        target_percentage = (region_sales_this_month / monthly_target * 100) if monthly_target > 0 else 0
        
        regional_performance.append({
            'region': region.name,
            'sales': region_sales_this_month,
            'target': monthly_target,
            'percentage': round(target_percentage, 1),
        })
    
    # Alerts & Notifications
    expiring_quotations_today = Quotation.objects.filter(
        valid_until=today,
        status__in=['sent', 'shared']
    ).count()
    
    # Follow-ups due today (visits scheduled for today)
    follow_ups_due_today = Visit.objects.filter(
        scheduled_date__date=today,
        status__in=['scheduled', 'in_progress']
    ).count()
    
    # QC inspections pending
    qc_pending = QCTracking.objects.filter(status__in=['pending', 'in_progress']).count()
    
    # Deliveries scheduled today
    deliveries_today = Dispatch.objects.filter(
        expected_delivery_date=today,
        status__in=['pending', 'dispatched', 'in_transit']
    ).count()
    
    context = {
        'total_customers': total_customers,
        'active_leads': active_leads,
        'pending_quotations': pending_quotations,
        'production_orders': production_orders,
        'recent_activities': recent_activities[:4],
        # Sales Pipeline
        'new_leads_this_month': new_leads_this_month,
        'leads_growth': round(leads_growth, 1),
        'quotations_sent_this_month': quotations_sent_this_month,
        'quotations_growth': round(quotations_growth, 1),
        'pos_this_month': pos_this_month,
        'pos_growth': round(pos_growth, 1),
        'deliveries_this_month': deliveries_this_month,
        'deliveries_growth': round(deliveries_growth, 1),
        # Regional Performance
        'regional_performance': regional_performance,
        # Alerts
        'expiring_quotations_today': expiring_quotations_today,
        'follow_ups_due_today': follow_ups_due_today,
        'qc_pending': qc_pending,
        'deliveries_today': deliveries_today,
        # Date
        'today': today,
    }
    
    return render(request, 'marketing/dashboard.html', context)

@login_required
def campaign_form(request):
    """Campaign Form Interface"""
    if request.method == 'POST':
        # Handle form submission (placeholder)
        messages.success(request, 'Campaign updated successfully!')
        return redirect('marketing:campaign_list')
    
    context = {}
    return render(request, 'marketing/campaign_form.html', context)


@hrms_login_required
@require_permission(MARKETING_PERMISSIONS['campaign.view'])
def campaign_list(request):
    """Campaign List View"""
    campaigns = Campaign.objects.select_related('created_by').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        campaigns = campaigns.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(campaign_type__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        campaigns = campaigns.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(campaigns, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Campaign.STATUS_CHOICES,
    }
    
    return render(request, 'marketing/campaign_list.html', context)

@login_required
def campaign_create(request):
    """Create New Campaign"""
    if request.method == 'POST':
        # Handle form submission
        name = request.POST.get('name')
        description = request.POST.get('description')
        campaign_type = request.POST.get('campaign_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        budget = request.POST.get('budget', 0)
        target_audience = request.POST.get('target_audience')
        goals = request.POST.get('goals')
        
        try:
            # Get user info from HRMS session
            user_info = get_user_info_dict(request)

            
            campaign = Campaign.objects.create(
                name=name,
                description=description,
                campaign_type=campaign_type,
                start_date=start_date,
                end_date=end_date,
                budget=budget,
                target_audience=target_audience,
                goals=goals,
                # Store HRMS user info instead of Django User
                created_by_user_id=user_info['user_id'],
                created_by_username=user_info['username'],
                created_by_email=user_info['email'],
                created_by_full_name=user_info['full_name'],
            )
            messages.success(request, f'Campaign "{campaign.name}" created successfully!')
            return redirect('marketing:campaign_detail', campaign_id=campaign.id)
        except Exception as e:
            messages.error(request, f'Error creating campaign: {str(e)}')
    
    context = {
        'campaign_types': Campaign.CAMPAIGN_TYPES,
    }
    
    return render(request, 'marketing/campaign_form.html', context)

@login_required
def campaign_detail(request, campaign_id):
    """Campaign Detail View"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    
    # Get campaign metrics
    metrics = CampaignMetric.objects.filter(campaign=campaign).order_by('date')
    
    # Get leads from this campaign
    leads = Lead.objects.filter(campaign=campaign).order_by('-created_at')
    
    context = {
        'campaign': campaign,
        'metrics': metrics,
        'leads': leads,
    }
    
    return render(request, 'marketing/campaign_detail.html', context)

@login_required
def lead_list(request):
    """Lead List View"""
    leads = Lead.objects.select_related('campaign').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        leads = leads.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(company__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        leads = leads.filter(status=status_filter)
    
    # Filter by source
    source_filter = request.GET.get('source', '')
    if source_filter:
        leads = leads.filter(source=source_filter)
    
    # Filter by assigned user (using HRMS user_id or Django User ID)
    assigned_filter = request.GET.get('assigned_to', '')
    if assigned_filter:
        # Try filtering by HRMS user_id first, then fall back to Django User ID
        leads = leads.filter(
            Q(assigned_to_user_id=assigned_filter) | 
            Q(assigned_to_id=assigned_filter)
        )
    
    # Filter by campaign
    campaign_filter = request.GET.get('campaign', '')
    if campaign_filter:
        leads = leads.filter(campaign_id=campaign_filter)
    
    # Pagination
    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'source_filter': source_filter,
        'assigned_filter': assigned_filter,
        'campaign_filter': campaign_filter,
        'status_choices': Lead.STATUS_CHOICES,
        'source_choices': Lead.SOURCE_CHOICES,
        'users': User.objects.filter(is_active=True),
        'campaigns': Campaign.objects.filter(status='active'),
    }
    
    return render(request, 'marketing/lead_list.html', context)

@login_required
def lead_create(request):
    """Create New Lead"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        # Check if email already exists
        if email and Lead.objects.filter(email=email).exists():
            messages.error(request, f'A lead with email "{email}" already exists. Please use a different email address.')
            # Return to form with existing data
            context = {
                'source_choices': Lead.SOURCE_CHOICES,
                'campaigns': Campaign.objects.filter(status='active'),
                'users': User.objects.filter(is_active=True),
                'form_data': request.POST,  # Pass form data back to template
            }
            return render(request, 'marketing/lead_form.html', context)
        
        try:
            lead = Lead.objects.create(
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                email=email,
                phone=request.POST.get('phone'),
                company=request.POST.get('company'),
                position=request.POST.get('position'),
                source=request.POST.get('source'),
                notes=request.POST.get('notes'),
                assigned_to_id=request.POST.get('assigned_to') if request.POST.get('assigned_to') else None,
                campaign_id=request.POST.get('campaign') if request.POST.get('campaign') else None,
            )
            messages.success(request, f'Lead "{lead.full_name}" created successfully!')
            return redirect('marketing:lead_detail', lead.id)
        except Exception as e:
            messages.error(request, f'Error creating lead: {str(e)}')
    
    context = {
        'source_choices': Lead.SOURCE_CHOICES,
        'campaigns': Campaign.objects.filter(status='active'),
        'users': User.objects.filter(is_active=True),
    }
    
    return render(request, 'marketing/lead_form.html', context)

@login_required
def lead_detail(request, lead_id):
    """Lead Detail View"""
    lead = get_object_or_404(Lead.objects.select_related('assigned_to', 'campaign'), id=lead_id)
    
    # Get lead activities
    activities = LeadActivity.objects.filter(lead=lead).select_related('performed_by').order_by('-performed_at')
    
    # Get related visits
    visits = Visit.objects.filter(customer__email=lead.email).order_by('-scheduled_date')
    
    # Get related quotations
    quotations = Quotation.objects.filter(customer__email=lead.email).order_by('-created_at')
    
    context = {
        'lead': lead,
        'activities': activities,
        'visits': visits,
        'quotations': quotations,
    }
    
    return render(request, 'marketing/lead_detail.html', context)

@login_required
def email_templates(request):
    """Email Templates List"""
    user_info = get_user_info_dict(request)
    templates = EmailTemplate.objects.filter(created_by_username=user_info['username']).order_by('-created_at')
    
    context = {
        'templates': templates,
    }
    
    return render(request, 'marketing/email_templates.html', context)

@login_required
def campaign_analytics(request):
    """Campaign Analytics View"""
    # Get analytics data
    campaigns = Campaign.objects.all().prefetch_related('leads')
    
    # Calculate metrics
    total_campaigns = campaigns.count()
    active_campaigns = campaigns.filter(status='active').count()
    total_budget = campaigns.aggregate(total=Sum('budget'))['total'] or 0
    
    # Calculate total leads from all campaigns
    total_leads = Lead.objects.filter(campaign__in=campaigns).count()
    
    # Calculate conversion rate
    converted_leads = Lead.objects.filter(campaign__in=campaigns, status='converted').count()
    conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
    
    context = {
        'total_campaigns': total_campaigns,
        'active_campaigns': active_campaigns,
        'total_budget': total_budget,
        'total_leads': total_leads,
        'conversion_rate': conversion_rate,
        'campaigns': campaigns,
    }
    
    return render(request, 'marketing/campaign_analytics.html', context)

@login_required
def lead_import(request):
    """Lead Import View"""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('lead_file')
        if uploaded_file:
            try:
                # Process CSV/Excel file
                # This is a placeholder - implement actual import logic
                messages.success(request, 'Lead import completed successfully!')
                return redirect('marketing:lead_list')
            except Exception as e:
                messages.error(request, f'Error importing leads: {str(e)}')
    
    return render(request, 'marketing/lead_import.html')

@login_required
def lead_scoring(request):
    """Lead Scoring View"""
    # Get leads with scores
    leads = Lead.objects.select_related('assigned_to').order_by('-score', '-created_at')
    
    # Calculate scoring statistics
    total_leads = leads.count()
    high_score_leads = leads.filter(score__gte=80).count()
    medium_score_leads = leads.filter(score__gte=50, score__lt=80).count()
    low_score_leads = leads.filter(score__lt=50).count()
    
    context = {
        'leads': leads,
        'total_leads': total_leads,
        'high_score_leads': high_score_leads,
        'medium_score_leads': medium_score_leads,
        'low_score_leads': low_score_leads,
    }
    
    return render(request, 'marketing/lead_scoring.html', context)

@login_required
def landing_pages(request):
    """Landing Pages View"""
    return render(request, 'marketing/landing_pages.html')

@login_required
def social_posts(request):
    """Social Posts View"""
    return render(request, 'marketing/social_posts.html')

@login_required
def campaign_performance(request):
    """Campaign Performance View"""
    return render(request, 'marketing/campaign_performance.html')

@login_required
def roi_analysis(request):
    """ROI Analysis View"""
    return render(request, 'marketing/roi_analysis.html')

@login_required
def conversion_funnel(request):
    """Conversion Funnel View"""
    return render(request, 'marketing/conversion_funnel.html')

@login_required
def workflows(request):
    """Workflows View"""
    context = {}
    return render(request, 'marketing/workflows.html', context)

@login_required
def triggers(request):
    """Triggers View"""
    context = {}
    return render(request, 'marketing/triggers.html', context)

@login_required
def sequences(request):
    """Sequences View"""
    context = {}
    return render(request, 'marketing/sequences.html', context)

@login_required
def production_planning(request):
    """Production Planning Interface"""
    if request.method == 'POST':
        work_order_id = request.POST.get('work_order')
        department = request.POST.get('department')
        priority = request.POST.get('priority')
        assigned_to_id = request.POST.get('assigned_to')
        planned_start_date = request.POST.get('planned_start_date')
        planned_end_date = request.POST.get('planned_end_date')
        resource_requirements = request.POST.get('resource_requirements')
        special_instructions = request.POST.get('special_instructions')
        
        # Generate plan number
        date = timezone.now()
        plan_number = f"PP{date.strftime('%Y%m%d')}{str(date.microsecond)[-3:]}"
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        production_plan = ProductionPlan.objects.create(
            work_order_id=work_order_id,
            plan_number=plan_number,
            department=department,
            priority=priority,
            assigned_to_id=assigned_to_id if assigned_to_id else None,
            planned_start_date=planned_start_date,
            planned_end_date=planned_end_date,
            resource_requirements=resource_requirements,
            special_instructions=special_instructions,
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
            status='draft'
        )
        
        messages.success(request, f'Production plan {plan_number} created successfully!')
        return redirect('marketing:production_planning')
    
    # Get production plans
    production_plans = ProductionPlan.objects.select_related('work_order__purchase_order__customer', 'assigned_to').order_by('-created_at')
    
    # Calculate statistics
    total_plans = production_plans.count()
    in_progress_plans = production_plans.filter(status='in_progress').count()
    completed_plans = production_plans.filter(status='completed').count()
    overdue_plans = production_plans.filter(
        planned_end_date__lt=timezone.now().date(),
        status__in=['draft', 'approved', 'in_progress']
    ).count()
    
    # Get work orders for form
    work_orders = WorkOrder.objects.select_related('purchase_order__customer').filter(status='approved')
    
    # Get users for assignment
    users = User.objects.filter(is_active=True)
    
    context = {
        'production_plans': production_plans,
        'work_orders': work_orders,
        'users': users,
        'total_plans': total_plans,
        'in_progress_plans': in_progress_plans,
        'completed_plans': completed_plans,
        'overdue_plans': overdue_plans,
    }
    return render(request, 'marketing/production_planning.html', context)

@login_required
def live_gps_tracking(request):
    """Live GPS Tracking Dashboard"""
    # Get current active visits with GPS data
    active_visits = Visit.objects.filter(
        scheduled_date__date=timezone.now().date(),
        status='in_progress'
    ).select_related('customer', 'location', 'assigned_to')
    
    # Get today's completed visits
    completed_visits = Visit.objects.filter(
        scheduled_date__date=timezone.now().date(),
        status='completed'
    ).select_related('customer', 'location', 'assigned_to')
    
    # Get upcoming visits
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date__gt=timezone.now().date()
    ).select_related('customer', 'location', 'assigned_to').order_by('scheduled_date')[:10]
    
    # Get all unique users from visits (using RBAC username)
    # Get unique usernames from active and completed visits
    active_usernames = set(active_visits.values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    completed_usernames = set(completed_visits.values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    all_usernames = active_usernames | completed_usernames
    
    user_status = {}
    
    for username in all_usernames:
        # Check if user has any active visits today
        active_visit = active_visits.filter(assigned_to_username=username).first()
        if active_visit:
            user_status[username] = {
                'status': 'active',
                'visit': active_visit,
                'location': f"{active_visit.location.city if active_visit.location else 'Unknown'}, {active_visit.location.state if active_visit.location else 'Unknown'}"
            }
        else:
            # Check if user has completed visits today
            completed_visit = completed_visits.filter(assigned_to_username=username).first()
            if completed_visit:
                user_status[username] = {
                    'status': 'completed',
                    'visit': completed_visit,
                    'location': f"{completed_visit.location.city if completed_visit.location else 'Unknown'}, {completed_visit.location.state if completed_visit.location else 'Unknown'}"
                }
            else:
                user_status[username] = {
                    'status': 'available',
                    'visit': None,
                    'location': 'Office'
                }
    
    # Calculate statistics
    total_active = active_visits.count()
    total_completed = completed_visits.count()
    total_upcoming = upcoming_visits.count()
    total_users = users.count()
    active_users = len([u for u in user_status.values() if u['status'] == 'active'])
    
    context = {
        'active_visits': active_visits,
        'completed_visits': completed_visits,
        'upcoming_visits': upcoming_visits,
        'users': users,
        'user_status': user_status,
        'total_active': total_active,
        'total_completed': total_completed,
        'total_upcoming': total_upcoming,
        'total_users': total_users,
        'active_users': active_users,
    }
    return render(request, 'marketing/live_gps_tracking.html', context)

@login_required
def progress_tracking_dashboard(request):
    """Progress Tracking Dashboard"""
    # Get date range from request
    start_date = request.GET.get('start_date', (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))
    
    # Convert to date objects
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Sales Progress
    total_quotations = Quotation.objects.filter(created_at__date__range=[start_date, end_date]).count()
    accepted_quotations = Quotation.objects.filter(
        created_at__date__range=[start_date, end_date],
        status='accepted'
    ).count()
    conversion_rate = (accepted_quotations / total_quotations * 100) if total_quotations > 0 else 0
    
    # Production Progress
    total_manufacturing = Manufacturing.objects.filter(created_at__date__range=[start_date, end_date]).count()
    completed_manufacturing = Manufacturing.objects.filter(
        created_at__date__range=[start_date, end_date],
        status='ready_dispatch'
    ).count()
    production_completion_rate = (completed_manufacturing / total_manufacturing * 100) if total_manufacturing > 0 else 0
    
    # Visit Progress
    total_visits = Visit.objects.filter(scheduled_date__date__range=[start_date, end_date]).count()
    completed_visits = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date],
        status='completed'
    ).count()
    visit_completion_rate = (completed_visits / total_visits * 100) if total_visits > 0 else 0
    
    # Lead Progress
    total_leads = Lead.objects.filter(created_at__date__range=[start_date, end_date]).count()
    converted_leads = Lead.objects.filter(
        created_at__date__range=[start_date, end_date],
        status='converted'
    ).count()
    lead_conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
    
    # Regional Performance
    regional_performance = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).values('customer__region__name').annotate(
        total_visits=Count('id'),
        completed_visits=Count('id', filter=Q(status='completed'))
    )
    
    # User Performance
    user_performance = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).values('assigned_to__username').annotate(
        total_visits=Count('id'),
        completed_visits=Count('id', filter=Q(status='completed'))
    )
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_quotations': total_quotations,
        'accepted_quotations': accepted_quotations,
        'conversion_rate': round(conversion_rate, 2),
        'total_manufacturing': total_manufacturing,
        'completed_manufacturing': completed_manufacturing,
        'production_completion_rate': round(production_completion_rate, 2),
        'total_visits': total_visits,
        'completed_visits': completed_visits,
        'visit_completion_rate': round(visit_completion_rate, 2),
        'total_leads': total_leads,
        'converted_leads': converted_leads,
        'lead_conversion_rate': round(lead_conversion_rate, 2),
        'regional_performance': regional_performance,
        'user_performance': user_performance,
    }
    return render(request, 'marketing/progress_tracking_dashboard.html', context)

@login_required
def performance_analytics_detailed(request):
    """Detailed Performance Analytics"""
    # Get date range from request
    start_date = request.GET.get('start_date', (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))
    
    # Convert to date objects
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Sales Performance
    sales_data = Quotation.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).values('created_at__date').annotate(
        total_quotations=Count('id'),
        accepted_quotations=Count('id', filter=Q(status='accepted')),
        total_value=Sum('total_amount'),
        accepted_value=Sum('total_amount', filter=Q(status='accepted'))
    ).order_by('created_at__date')
    
    # Production Performance
    production_data = Manufacturing.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).values('created_at__date').annotate(
        total_batches=Count('id'),
        completed_batches=Count('id', filter=Q(status='ready_dispatch'))
    ).order_by('created_at__date')
    
    # Visit Performance
    visit_data = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).values('scheduled_date__date').annotate(
        total_visits=Count('id'),
        completed_visits=Count('id', filter=Q(status='completed'))
    ).order_by('scheduled_date__date')
    
    # Lead Performance
    lead_data = Lead.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).values('created_at__date').annotate(
        total_leads=Count('id'),
        converted_leads=Count('id', filter=Q(status='converted'))
    ).order_by('created_at__date')
    
    # Top Performing Users
    top_users = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).values('assigned_to__username').annotate(
        total_visits=Count('id'),
        completed_visits=Count('id', filter=Q(status='completed'))
    ).annotate(
        completion_rate=ExpressionWrapper(
            F('completed_visits') * 100.0 / F('total_visits'),
            output_field=FloatField()
        )
    ).order_by('-completion_rate')[:10]
    
    # Top Customers
    top_customers = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).values('customer__name').annotate(
        total_visits=Count('id'),
        completed_visits=Count('id', filter=Q(status='completed'))
    ).order_by('-total_visits')[:10]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'sales_data': sales_data,
        'production_data': production_data,
        'visit_data': visit_data,
        'lead_data': lead_data,
        'top_users': top_users,
        'top_customers': top_customers,
    }
    return render(request, 'marketing/performance_analytics_detailed.html', context)

@login_required
def export_data_advanced(request):
    """Advanced Export Functionality"""
    if request.method == 'POST':
        export_type = request.POST.get('export_type')
        date_range = request.POST.get('date_range')
        format_type = request.POST.get('format_type', 'excel')
        
        # Parse date range
        if date_range == 'custom':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
        elif date_range == 'last_7_days':
            start_date = (timezone.now().date() - timedelta(days=7)).strftime('%Y-%m-%d')
            end_date = timezone.now().date().strftime('%Y-%m-%d')
        elif date_range == 'last_30_days':
            start_date = (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
            end_date = timezone.now().date().strftime('%Y-%m-%d')
        elif date_range == 'this_month':
            start_date = timezone.now().date().replace(day=1).strftime('%Y-%m-%d')
            end_date = timezone.now().date().strftime('%Y-%m-%d')
        else:  # all_time
            start_date = '2020-01-01'
            end_date = timezone.now().date().strftime('%Y-%m-%d')
        
        # Generate export data based on type
        if export_type == 'sales':
            data = Quotation.objects.filter(
                created_at__date__range=[start_date, end_date]
            ).select_related('customer', 'urs')
        elif export_type == 'production':
            data = Manufacturing.objects.filter(
                created_at__date__range=[start_date, end_date]
            ).select_related('work_order__purchase_order__customer')
        elif export_type == 'visits':
            data = Visit.objects.filter(
                scheduled_date__date__range=[start_date, end_date]
            ).select_related('customer', 'location', 'assigned_to')
        elif export_type == 'leads':
            data = Lead.objects.filter(
                created_at__date__range=[start_date, end_date]
            )
        elif export_type == 'customers':
            data = Customer.objects.all()
        else:  # comprehensive
            data = {
                'quotations': Quotation.objects.filter(created_at__date__range=[start_date, end_date]),
                'manufacturing': Manufacturing.objects.filter(created_at__date__range=[start_date, end_date]),
                'visits': Visit.objects.filter(scheduled_date__date__range=[start_date, end_date]),
                'leads': Lead.objects.filter(created_at__date__range=[start_date, end_date]),
            }
        
        messages.success(request, f'{export_type.title()} data exported successfully!')
        return redirect('marketing:export_data_advanced')
    
    context = {}
    return render(request, 'marketing/export_data_advanced.html', context)

@login_required
def exhibition_planning_interface(request):
    """Exhibition Planning Interface"""
    if request.method == 'POST':
        exhibition_name = request.POST.get('exhibition_name')
        event_date = request.POST.get('event_date')
        venue = request.POST.get('venue')
        city = request.POST.get('city')
        organizer = request.POST.get('organizer')
        booth_size = request.POST.get('booth_size')
        expected_visitors = request.POST.get('expected_visitors')
        budget = request.POST.get('budget')
        objectives = request.POST.get('objectives')
        target_audience = request.POST.get('target_audience')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        exhibition = Exhibition.objects.create(
            name=exhibition_name,
            event_date=event_date,
            venue=venue,
            city=city,
            organizer=organizer,
            booth_size=booth_size,
            expected_visitors=expected_visitors,
            budget=budget,
            objectives=objectives,
            target_audience=target_audience,
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
            status='planning'
        )
        
        messages.success(request, f'Exhibition "{exhibition_name}" planned successfully!')
        return redirect('marketing:exhibition_planning_interface')
    
    # Get exhibitions
    exhibitions = Exhibition.objects.all().order_by('-event_date')
    
    # Calculate statistics
    total_exhibitions = exhibitions.count()
    upcoming_exhibitions = exhibitions.filter(event_date__gte=timezone.now().date()).count()
    completed_exhibitions = exhibitions.filter(event_date__lt=timezone.now().date()).count()
    active_exhibitions = exhibitions.filter(status='active').count()
    
    # Get upcoming exhibitions for timeline
    upcoming_timeline = exhibitions.filter(
        event_date__gte=timezone.now().date()
    ).order_by('event_date')[:5]
    
    context = {
        'exhibitions': exhibitions,
        'total_exhibitions': total_exhibitions,
        'upcoming_exhibitions': upcoming_exhibitions,
        'completed_exhibitions': completed_exhibitions,
        'active_exhibitions': active_exhibitions,
        'upcoming_timeline': upcoming_timeline,
    }
    return render(request, 'marketing/exhibition_planning_interface.html', context)

@login_required
def vendor_management(request):
    """Vendor Management Interface"""
    if request.method == 'POST':
        vendor_name = request.POST.get('vendor_name')
        vendor_type = request.POST.get('vendor_type')
        contact_person = request.POST.get('contact_person')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        services = request.POST.get('services')
        rating = request.POST.get('rating')
        
        # Create vendor (assuming we have a Vendor model or use Exhibition model)
        # For now, we'll store vendor info in a structured way
        vendor_data = {
            'name': vendor_name,
            'type': vendor_type,
            'contact_person': contact_person,
            'email': email,
            'phone': phone,
            'address': address,
            'services': services,
            'rating': rating,
            'created_by': request.user.username,
            'created_at': timezone.now().isoformat(),
        }
        
        messages.success(request, f'Vendor "{vendor_name}" added successfully!')
        return redirect('marketing:vendor_management')
    
    # Get vendors from Customer model (filtering by vendor-related types or using a tag/field)
    # For now, we'll use customers that might be vendors, or return empty list
    # In a real implementation, you'd have a Vendor model or a customer_type='vendor'
    vendors = []
    # Option: Use Customer model if you have vendor customers
    # vendors_customers = Customer.objects.filter(customer_type='vendor') if 'vendor' in [c[0] for c in Customer.CUSTOMER_TYPES] else Customer.objects.none()
    
    # Calculate statistics
    total_vendors = len(vendors)
    active_vendors = len([v for v in vendors if v.get('status') == 'active'])
    avg_rating = sum(v.get('rating', 0) for v in vendors) / len(vendors) if vendors else 0
    
    context = {
        'vendors': vendors,
        'total_vendors': total_vendors,
        'active_vendors': active_vendors,
        'avg_rating': round(avg_rating, 1),
    }
    return render(request, 'marketing/vendor_management.html', context)

@login_required
def visitor_database_management(request):
    """Visitor Database Management"""
    if request.method == 'POST':
        visitor_name = request.POST.get('visitor_name')
        company = request.POST.get('company')
        designation = request.POST.get('designation')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        exhibition_id = request.POST.get('exhibition')
        interest_level = request.POST.get('interest_level')
        notes = request.POST.get('notes')
        
        # Create visitor record (placeholder for now)
        visitor_data = {
            'name': visitor_name,
            'company': company,
            'designation': designation,
            'email': email,
            'phone': phone,
            'exhibition': exhibition_id,
            'interest_level': interest_level,
            'notes': notes,
            'created_by': request.user.username,
            'created_at': timezone.now().isoformat(),
        }
        
        messages.success(request, f'Visitor "{visitor_name}" added to database!')
        return redirect('marketing:visitor_database_management')
    
    # Get exhibitions for dropdown
    exhibitions = Exhibition.objects.all().order_by('-event_date')
    
    # Get visitors from Lead model (leads generated from exhibitions/events)
    exhibition_filter = request.GET.get('exhibition', '')
    visitors_qs = Lead.objects.filter(source='event').order_by('-created_at')
    
    if exhibition_filter:
        # Filter by exhibition date range if exhibition is selected
        try:
            exhibition = Exhibition.objects.get(id=exhibition_filter)
            visitors_qs = visitors_qs.filter(
                created_at__date__gte=exhibition.start_date,
                created_at__date__lte=exhibition.end_date
            )
        except Exhibition.DoesNotExist:
            pass
    
    visitors = []
    for lead in visitors_qs[:100]:  # Limit to 100 for performance
        # Determine interest level based on lead score
        if lead.score >= 70:
            interest_level = 'High'
        elif lead.score >= 40:
            interest_level = 'Medium'
        else:
            interest_level = 'Low'
        
        # Get associated exhibition if any
        exhibition_name = 'N/A'
        if lead.campaign:
            # Try to find exhibition by campaign or date
            exhibition = Exhibition.objects.filter(
                start_date__lte=lead.created_at.date(),
                end_date__gte=lead.created_at.date()
            ).first()
            if exhibition:
                exhibition_name = exhibition.name
        
        visitors.append({
            'name': lead.full_name,
            'company': lead.company or 'N/A',
            'designation': lead.position or 'N/A',
            'email': lead.email,
            'phone': lead.phone or 'N/A',
            'exhibition': exhibition_name,
            'interest_level': interest_level,
            'visit_date': lead.created_at.strftime('%Y-%m-%d'),
            'follow_up_status': lead.status.title()
        })
    
    # Calculate statistics
    total_visitors = len(visitors)
    high_interest = len([v for v in visitors if v['interest_level'] == 'High'])
    pending_followup = len([v for v in visitors if v['follow_up_status'] in ['New', 'Contacted', 'Qualified']])
    
    context = {
        'visitors': visitors,
        'exhibitions': exhibitions,
        'total_visitors': total_visitors,
        'high_interest': high_interest,
        'pending_followup': pending_followup,
    }
    return render(request, 'marketing/visitor_database_management.html', context)

@login_required
def exhibition_expense_tracking(request):
    """Exhibition Expense Tracking"""
    if request.method == 'POST':
        exhibition_id = request.POST.get('exhibition')
        expense_type = request.POST.get('expense_type')
        amount = request.POST.get('amount')
        description = request.POST.get('description')
        vendor = request.POST.get('vendor')
        expense_date = request.POST.get('expense_date')
        
        # Create expense record
        user_info = get_user_info_dict(request)
        expense = Expense.objects.create(
            expense_user_id=user_info['user_id'],
            expense_username=user_info['username'],
            expense_email=user_info['email'],
            expense_full_name=user_info['full_name'],
            expense_type=expense_type,
            amount=amount,
            description=f"Exhibition: {description}",
            expense_date=expense_date,
            vendor=vendor,
            status='pending'
        )
        
        messages.success(request, f'Expense of ₹{amount} added successfully!')
        return redirect('marketing:exhibition_expense_tracking')
    
    # Get exhibitions
    exhibitions = Exhibition.objects.all().order_by('-event_date')
    
    # Get exhibition expenses
    exhibition_expenses = Expense.objects.filter(
        description__icontains='Exhibition'
    ).order_by('-expense_date')
    
    # Calculate statistics
    total_expenses = exhibition_expenses.aggregate(total=Sum('amount'))['total'] or 0
    pending_expenses = exhibition_expenses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
    approved_expenses = exhibition_expenses.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0
    
    # Group expenses by type
    expense_by_type = exhibition_expenses.values('expense_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    context = {
        'exhibitions': exhibitions,
        'exhibition_expenses': exhibition_expenses,
        'total_expenses': total_expenses,
        'pending_expenses': pending_expenses,
        'approved_expenses': approved_expenses,
        'expense_by_type': expense_by_type,
    }
    return render(request, 'marketing/exhibition_expense_tracking.html', context)

@login_required
def post_event_analysis(request):
    """Post Event Analysis"""
    # Get completed exhibitions
    completed_exhibitions = Exhibition.objects.filter(
        end_date__lt=timezone.now().date()
    ).order_by('-end_date')
    
    # Get exhibition performance data from database
    exhibition_performance = []
    for exhibition in completed_exhibitions:
        # Get leads generated from this exhibition (leads created during exhibition period)
        exhibition_leads = Lead.objects.filter(
            source='event',
            created_at__date__gte=exhibition.start_date,
            created_at__date__lte=exhibition.end_date
        )
        qualified_leads = exhibition_leads.filter(status__in=['qualified', 'proposal', 'negotiation', 'converted']).count()
        total_leads = exhibition_leads.count()
        
        # Calculate conversion rate
        conversion_rate = (qualified_leads / total_leads * 100) if total_leads > 0 else 0
        
        # Get expenses for this exhibition
        exhibition_expenses = Expense.objects.filter(
            expense_type__in=['travel', 'accommodation', 'entertainment', 'office'],
            date__gte=exhibition.start_date,
            date__lte=exhibition.end_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Calculate ROI (simplified - using exhibition budget vs expenses)
        roi = ((exhibition.budget - exhibition_expenses) / exhibition_expenses * 100) if exhibition_expenses > 0 else 0
        
        exhibition_performance.append({
            'exhibition': exhibition.name,
            'event_date': exhibition.start_date.strftime('%Y-%m-%d'),
            'total_visitors': exhibition.visitor_count,
            'qualified_leads': qualified_leads,
            'conversion_rate': round(conversion_rate, 1),
            'total_expenses': float(exhibition_expenses),
            'roi': round(float(roi), 1),
            'visitor_satisfaction': 0,  # Would need a separate model for this
            'objectives_achieved': 'N/A'  # Would need a separate model for this
        })
    
    # Calculate overall statistics
    total_exhibitions = len(exhibition_performance)
    total_visitors = sum(e['total_visitors'] for e in exhibition_performance)
    total_leads = sum(e['qualified_leads'] for e in exhibition_performance)
    avg_conversion = sum(e['conversion_rate'] for e in exhibition_performance) / total_exhibitions if total_exhibitions > 0 else 0
    total_expenses = sum(e['total_expenses'] for e in exhibition_performance)
    avg_roi = sum(e['roi'] for e in exhibition_performance) / total_exhibitions if total_exhibitions > 0 else 0
    
    context = {
        'completed_exhibitions': completed_exhibitions,
        'exhibition_performance': exhibition_performance,
        'total_exhibitions': total_exhibitions,
        'total_visitors': total_visitors,
        'total_leads': total_leads,
        'avg_conversion': round(avg_conversion, 1),
        'total_expenses': total_expenses,
        'avg_roi': round(avg_roi, 1),
    }
    return render(request, 'marketing/post_event_analysis.html', context)

# Customer Management Views
@login_required
def customer_form(request):
    """Customer Form Interface"""
    if request.method == 'POST':
        # Handle form submission (placeholder)
        messages.success(request, 'Customer updated successfully!')
        return redirect('marketing:customer_list')
    
    regions = Region.objects.all()
    context = {
        'regions': regions,
        'customer_types': Customer.CUSTOMER_TYPES,
    }
    return render(request, 'marketing/customer_form.html', context)


@login_required
def customer_registration(request):
    """Customer Registration with Multi-Location Support"""
    if request.method == 'POST':
        # Handle customer registration form
        customer_name = request.POST.get('customer_name')
        contact_person = request.POST.get('contact_person')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        region_id = request.POST.get('region')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        # Create customer
        customer = Customer.objects.create(
            name=customer_name,
            contact_person=contact_person,
            email=email,
            phone=phone,
            region_id=region_id,
            # Store HRMS user info instead of Django User
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
        )
        
        # Handle multiple locations
        locations_data = request.POST.getlist('location_address')
        cities = request.POST.getlist('location_city')
        states = request.POST.getlist('location_state')
        pincodes = request.POST.getlist('location_pincode')
        
        for i in range(len(locations_data)):
            if locations_data[i].strip():
                CustomerLocation.objects.create(
                    customer=customer,
                    address=locations_data[i],
                    city=cities[i] if i < len(cities) else '',
                    state=states[i] if i < len(states) else '',
                    pincode=pincodes[i] if i < len(pincodes) else '',
                    is_primary=(i == 0)  # First location is primary
                )
        
        messages.success(request, f'Customer "{customer_name}" registered successfully!')
        return redirect('marketing:customer_list')
    
    regions = Region.objects.all()
    context = {
        'regions': regions,
    }
    return render(request, 'marketing/customer_registration.html', context)

@login_required
def customer_list(request):
    """Region-wise Customer Listing with Search and Filtering"""
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    region_filter = request.GET.get('region', '')
    status_filter = request.GET.get('status', '')
    
    # Get all customers with locations
    customers = Customer.objects.prefetch_related('locations').all()
    
    # Apply search filter
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(locations__address__icontains=search_query)
        ).distinct()
    
    # Apply region filter
    if region_filter:
        customers = customers.filter(region_id=region_filter)
    
    # Apply status filter (using customer_type instead of is_active)
    if status_filter == 'active':
        customers = customers.filter(customer_type='existing')
    elif status_filter == 'inactive':
        customers = customers.filter(customer_type='lapsed')
    
    # Get customer statistics
    customer_stats = {
        'total_customers': Customer.objects.count(),
        'active_customers': Customer.objects.filter(customer_type='existing').count(),
        'customers_with_locations': Customer.objects.filter(locations__isnull=False).distinct().count(),
        'new_customers_this_month': Customer.objects.filter(
            created_at__month=datetime.now().month,
            created_at__year=datetime.now().year
        ).count(),
    }
    
    # Get region-wise distribution
    region_distribution = []
    regions = Region.objects.all()
    for region in regions:
        customer_count = customers.filter(region=region).count()
        if customer_count > 0:
            region_distribution.append({
                'region': region.name,
                'count': customer_count,
                'percentage': (customer_count / customer_stats['total_customers']) * 100 if customer_stats['total_customers'] > 0 else 0
            })
    
    # Pagination
    paginator = Paginator(customers, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'customer_stats': customer_stats,
        'region_distribution': region_distribution,
        'search_query': search_query,
        'region_filter': region_filter,
        'status_filter': status_filter,
        'regions': regions,
    }
    return render(request, 'marketing/customer_list.html', context)

@login_required
def customer_detail(request, customer_id):
    """Customer Detail View with Locations and History"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Get customer's leads (filter by email since Lead doesn't have customer FK)
    leads = Lead.objects.filter(email=customer.email).order_by('-created_at')
    
    # Get customer's visits
    visits = Visit.objects.filter(customer=customer).order_by('-scheduled_date')
    
    # Get customer's quotations
    quotations = Quotation.objects.filter(customer=customer).order_by('-created_at')
    
    # Get customer's purchase orders
    purchase_orders = PurchaseOrder.objects.filter(customer=customer).order_by('-created_at')
    
    context = {
        'customer': customer,
        'leads': leads,
        'visits': visits,
        'quotations': quotations,
        'purchase_orders': purchase_orders,
    }
    return render(request, 'marketing/customer_detail.html', context)

@login_required
def customer_regions(request):
    """Region-wise Customer View"""
    regions = Region.objects.prefetch_related('customer_set').all()
    
    # Get statistics for each region
    for region in regions:
        region.total_customers = region.customer_set.count()
        region.prospects = region.customer_set.filter(customer_type='prospect').count()
        region.existing_customers = region.customer_set.filter(customer_type='existing').count()
        region.lapsed_customers = region.customer_set.filter(customer_type='lapsed').count()
    
    context = {
        'regions': regions,
    }
    
    return render(request, 'marketing/customer_regions.html', context)

@login_required
def customer_import(request):
    """Customer Import View"""
    if request.method == 'POST':
        # Handle file upload and import
        uploaded_file = request.FILES.get('customer_file')
        if uploaded_file:
            try:
                # Process CSV/Excel file
                # This is a placeholder - implement actual import logic
                messages.success(request, 'Customer import completed successfully!')
                return redirect('marketing:customer_list')
            except Exception as e:
                messages.error(request, f'Error importing customers: {str(e)}')
    
    return render(request, 'marketing/customer_import.html')

@login_required
def lead_form(request):
    """Lead Form Interface"""
    if request.method == 'POST':
        # Handle form submission (placeholder)
        messages.success(request, 'Lead updated successfully!')
        return redirect('marketing:lead_list')
    
    context = {}
    return render(request, 'marketing/lead_form.html', context)


@login_required
def lead_generation(request):
    """Lead Generation Form with Source Selection"""
    if request.method == 'POST':
        # Handle lead generation form
        customer_id = request.POST.get('customer')
        source = request.POST.get('source')
        description = request.POST.get('description')
        potential_value = request.POST.get('potential_value')
        expected_closing_date = request.POST.get('expected_closing_date')
        
        # Create lead
        # Get customer to extract lead information
        try:
            customer = Customer.objects.get(id=customer_id)
        except Customer.DoesNotExist:
            messages.error(request, 'Customer not found')
            return redirect('marketing:lead_generation')
        
        # Get user info from HRMS session

        
        # Build notes with all information
        notes_parts = []
        if description:
            notes_parts.append(f"Description: {description}")
        if potential_value:
            notes_parts.append(f"Potential Value: {potential_value}")
        if expected_closing_date:
            notes_parts.append(f"Expected Closing Date: {expected_closing_date}")
        notes_parts.append(f"Lead generated from customer: {customer.name}")
        notes = "\n".join(notes_parts)
        
        # Split contact person name
        contact_name = customer.contact_person or 'Unknown'
        name_parts = contact_name.split()
        first_name = name_parts[0] if name_parts else 'Unknown'
        last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
        
        # Create lead with fields that exist in Lead model
        try:
            # Get user info from HRMS session
            user_info = get_user_info_dict(request)


            lead = Lead.objects.create(
                first_name=first_name,
                last_name=last_name,
                email=customer.email,
                phone=customer.phone,
                company=customer.name,
                source=source or 'other',
                notes=notes,
                # Store HRMS user info instead of Django User
                assigned_to_user_id=user_info['user_id'],
                assigned_to_username=user_info['username'],
                assigned_to_email=user_info['email'],
                assigned_to_full_name=user_info['full_name'],
            )
            messages.success(request, f'Lead "{lead.full_name}" generated successfully!')
            return redirect('marketing:lead_list')
        except Exception as e:
            messages.error(request, f'Error creating lead: {str(e)}')
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating lead: {str(e)}")
        
        messages.success(request, 'Lead generated successfully!')
        return redirect('marketing:lead_list')
    
    customers = Customer.objects.all()
    context = {
        'customers': customers,
        'lead_sources': Lead.SOURCE_CHOICES,
    }
    return render(request, 'marketing/lead_generation.html', context)

# Duplicate lead_list function removed - already defined earlier

@login_required
def customer_visit(request):
    """Customer Visit Form with GPS Integration"""
    if request.method == 'POST':
        # Handle customer visit form
        customer_id = request.POST.get('customer')
        visit_date = request.POST.get('visit_date')
        visit_time = request.POST.get('visit_time')
        purpose = request.POST.get('purpose')
        outcome = request.POST.get('outcome')
        next_follow_up = request.POST.get('next_follow_up')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        
        # Create visit
        visit = Visit.objects.create(
            customer_id=customer_id,
            scheduled_date=visit_date,
            purpose=purpose,
            outcome=outcome,
            next_follow_up_date=next_follow_up,
            gps_latitude=latitude,
            gps_longitude=longitude,
            assigned_to=get_django_user(request)
        )
        
        # Handle multiple participants
        participants = request.POST.getlist('participants')
        participant_roles = request.POST.getlist('participant_roles')
        participant_notes = request.POST.getlist('participant_notes')
        
        # Add the primary assigned user as a participant
        user_info = get_user_info_dict(request)
        VisitParticipant.objects.create(
            visit=visit,
            participant_user_id=user_info['user_id'],
            participant_username=user_info['username'],
            participant_email=user_info['email'],
            participant_full_name=user_info['full_name'],
            role='primary',
            is_primary=True
        )
        
        # Add additional participants
        for i, participant_id in enumerate(participants):
            if participant_id and participant_id != str(request.user.id):
                role = participant_roles[i] if i < len(participant_roles) else 'secondary'
                notes = participant_notes[i] if i < len(participant_notes) else ''
                
                VisitParticipant.objects.create(
                    visit=visit,
                    user_id=participant_id,
                    role=role,
                    is_primary=False,
                    notes=notes
                )
        
        messages.success(request, f'Visit recorded successfully with {visit.participant_count} participants!')
        return redirect('marketing:visit_list')
    
    customers = Customer.objects.all()
    users = User.objects.filter(is_active=True)
    context = {
        'customers': customers,
        'users': users,
    }
    return render(request, 'marketing/customer_visit.html', context)

@login_required
def visit_list(request):
    """Visit Listing with GPS Data"""
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    date_filter = request.GET.get('date_filter', '')
    user_filter = request.GET.get('user', '')
    
    # Get all visits with participants
    visits = Visit.objects.select_related('customer', 'assigned_to').prefetch_related('participants__user').all()
    
    # Apply search filter
    if search_query:
        visits = visits.filter(
            Q(customer__name__icontains=search_query) |
            Q(purpose__icontains=search_query) |
            Q(outcome__icontains=search_query)
        )
    
    # Apply date filter
    if date_filter == 'today':
        visits = visits.filter(scheduled_date__date=datetime.now().date())
    elif date_filter == 'this_week':
        visits = visits.filter(scheduled_date__week=datetime.now().isocalendar()[1])
    elif date_filter == 'this_month':
        visits = visits.filter(scheduled_date__month=datetime.now().month)
    
    # Apply user filter
    if user_filter:
        visits = visits.filter(assigned_to_id=user_filter)
    
    # Get visit statistics
    visit_stats = {
        'total_visits': Visit.objects.count(),
        'visits_today': Visit.objects.filter(scheduled_date__date=datetime.now().date()).count(),
        'visits_this_week': Visit.objects.filter(scheduled_date__week=datetime.now().isocalendar()[1]).count(),
        'visits_this_month': Visit.objects.filter(scheduled_date__month=datetime.now().month).count(),
    }
    
    # Pagination
    paginator = Paginator(visits, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'visit_stats': visit_stats,
        'search_query': search_query,
        'date_filter': date_filter,
        'user_filter': user_filter,
        'users': get_user_model().objects.filter(is_active=True),
    }
    return render(request, 'marketing/visit_list.html', context)

@login_required
def follow_up_reminders(request):
    """Follow-up Reminder System"""
    # Get upcoming follow-ups
    today = datetime.now().date()
    upcoming_follow_ups = Visit.objects.filter(
        next_follow_up_date__gte=today,
        next_follow_up_date__lte=today + timedelta(days=7)
    ).select_related('customer', 'assigned_to').order_by('next_follow_up_date')
    
    # Get overdue follow-ups
    overdue_follow_ups = Visit.objects.filter(
        next_follow_up_date__lt=today
    ).select_related('customer', 'assigned_to').order_by('next_follow_up_date')
    
    # Get follow-up statistics
    follow_up_stats = {
        'total_follow_ups': Visit.objects.filter(next_follow_up_date__isnull=False).count(),
        'upcoming_follow_ups': upcoming_follow_ups.count(),
        'overdue_follow_ups': overdue_follow_ups.count(),
        'completed_today': Visit.objects.filter(scheduled_date__date=today).count(),
    }
    
    context = {
        'upcoming_follow_ups': upcoming_follow_ups,
        'overdue_follow_ups': overdue_follow_ups,
        'follow_up_stats': follow_up_stats,
    }
    return render(request, 'marketing/follow_up_reminders.html', context)

@login_required
def daily_status_report(request):
    """Daily Status Report Generation"""
    today = datetime.now().date()
    
    # Get today's activities
    today_visits = Visit.objects.filter(scheduled_date__date=today).select_related('customer', 'assigned_to')
    today_leads = Lead.objects.filter(created_at__date=today).select_related('assigned_to', 'campaign')
    today_quotations = Quotation.objects.filter(created_at__date=today).select_related('customer', 'created_by')
    
    # Get user-wise summary (using RBAC usernames)
    user_summary = []
    # Get unique usernames from today's activities
    visit_usernames = set(today_visits.values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    lead_usernames = set(today_leads.values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    quotation_usernames = set(today_quotations.values_list('created_by_username', flat=True).exclude(created_by_username=''))
    all_usernames = visit_usernames | lead_usernames | quotation_usernames
    
    for username in all_usernames:
        user_visits = today_visits.filter(assigned_to_username=username).count()
        user_leads = today_leads.filter(assigned_to_username=username).count()
        user_quotations = today_quotations.filter(created_by_username=username).count()
        
        if user_visits > 0 or user_leads > 0 or user_quotations > 0:
            user_summary.append({
                'username': username,
                'visits': user_visits,
                'leads': user_leads,
                'quotations': user_quotations,
            })
    
    # Get region-wise summary
    region_summary = []
    regions = Region.objects.all()
    for region in regions:
        region_visits = today_visits.filter(customer__region=region).count()
        # For leads, we need to check if they have a company and if that company has a region
        # Since leads don't directly link to customers, we'll count leads by company name matching customer names
        region_leads = 0  # We'll need to implement this differently since leads don't have direct region relationship
        region_quotations = today_quotations.filter(customer__region=region).count()
        
        if region_visits > 0 or region_leads > 0 or region_quotations > 0:
            region_summary.append({
                'region': region,
                'visits': region_visits,
                'leads': region_leads,
                'quotations': region_quotations,
            })
    
    context = {
        'today': today,
        'today_visits': today_visits,
        'today_leads': today_leads,
        'today_quotations': today_quotations,
        'user_summary': user_summary,
        'region_summary': region_summary,
    }
    return render(request, 'marketing/daily_status_report.html', context)

# Sales Process Views
@login_required
def urs_evaluation(request):
    """URS (User Requirement Specification) Evaluation"""
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        project_name = request.POST.get('project_name')
        requirement_details = request.POST.get('requirement_details')
        technical_specs = request.POST.get('technical_specs')
        site_requirements = request.POST.get('site_requirements')
        timeline = request.POST.get('timeline')
        budget_range = request.POST.get('budget_range')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        urs = URS.objects.create(
            customer_id=customer_id,
            project_name=project_name,
            requirement_details=requirement_details,
            technical_specs=technical_specs,
            site_requirements=site_requirements,
            timeline=timeline,
            budget_range=budget_range,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
            status='pending'
        )
        
        messages.success(request, f'URS for "{project_name}" created successfully!')
        return redirect('marketing:urs_list')
    
    customers = Customer.objects.all()
    context = {
        'customers': customers,
    }
    return render(request, 'marketing/urs_evaluation.html', context)

@login_required
def urs_list(request):
    """List all URS evaluations"""
    urs_list = URS.objects.select_related('customer', 'created_by').all().order_by('-created_at')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        urs_list = urs_list.filter(
            Q(project_name__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(requirement_details__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        urs_list = urs_list.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(urs_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_urs = URS.objects.count()
    pending_urs = URS.objects.filter(status='pending').count()
    approved_urs = URS.objects.filter(status='approved').count()
    rejected_urs = URS.objects.filter(status='rejected').count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'total_urs': total_urs,
        'pending_urs': pending_urs,
        'approved_urs': approved_urs,
        'rejected_urs': rejected_urs,
    }
    return render(request, 'marketing/urs_list.html', context)

@login_required
def urs_detail(request, urs_id):
    """View URS details"""
    urs = get_object_or_404(URS, id=urs_id)
    related_quotations = Quotation.objects.filter(urs=urs)
    related_ga_drawings = GADrawing.objects.filter(urs=urs)
    
    context = {
        'urs': urs,
        'related_quotations': related_quotations,
        'related_ga_drawings': related_ga_drawings,
    }
    return render(request, 'marketing/urs_detail.html', context)

@login_required
def ga_drawing_create(request):
    """Create GA Drawing for URS"""
    if request.method == 'POST':
        urs_id = request.POST.get('urs')
        drawing_title = request.POST.get('drawing_title')
        drawing_file = request.FILES.get('drawing_file')
        drawing_details = request.POST.get('drawing_details')
        version = request.POST.get('version', '1.0')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        ga_drawing = GADrawing.objects.create(
            urs_id=urs_id,
            title=drawing_title,
            drawing_file=drawing_file,
            details=drawing_details,
            version=version,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
            status='draft'
        )
        
        messages.success(request, f'GA Drawing "{drawing_title}" created successfully!')
        return redirect('marketing:ga_drawing_list')
    
    urs_list = URS.objects.filter(status='approved')
    context = {
        'urs_list': urs_list,
    }
    return render(request, 'marketing/ga_drawing_create.html', context)

@login_required
def ga_drawing_list(request):
    """List all GA Drawings"""
    drawings = GADrawing.objects.select_related('urs', 'created_by').all().order_by('-created_at')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        drawings = drawings.filter(
            Q(title__icontains=search) |
            Q(urs__project_name__icontains=search) |
            Q(urs__customer__name__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        drawings = drawings.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(drawings, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_drawings = GADrawing.objects.count()
    draft_drawings = GADrawing.objects.filter(status='draft').count()
    shared_drawings = GADrawing.objects.filter(status='shared').count()
    approved_drawings = GADrawing.objects.filter(status='approved').count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'total_drawings': total_drawings,
        'draft_drawings': draft_drawings,
        'shared_drawings': shared_drawings,
        'approved_drawings': approved_drawings,
    }
    return render(request, 'marketing/ga_drawing_list.html', context)

@login_required
def quotation_create(request):
    """Create Quotation for URS"""
    if request.method == 'POST':
        urs_id = request.POST.get('urs')
        quotation_number = request.POST.get('quotation_number')
        total_amount = request.POST.get('total_amount')
        validity_period = request.POST.get('validity_period')
        payment_terms = request.POST.get('payment_terms')
        delivery_terms = request.POST.get('delivery_terms')
        technical_specs = request.POST.get('technical_specs')
        terms_conditions = request.POST.get('terms_conditions')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        quotation = Quotation.objects.create(
            urs_id=urs_id,
            quotation_number=quotation_number,
            total_amount=total_amount,
            validity_period=validity_period,
            payment_terms=payment_terms,
            delivery_terms=delivery_terms,
            technical_specs=technical_specs,
            terms_conditions=terms_conditions,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
            status='draft'
        )
        
        messages.success(request, f'Quotation "{quotation_number}" created successfully!')
        return redirect('marketing:quotation_list')
    
    urs_list = URS.objects.filter(status='approved')
    context = {
        'urs_list': urs_list,
    }
    return render(request, 'marketing/quotation_create.html', context)

@login_required
def quotation_list(request):
    """List all Quotations"""
    quotations = Quotation.objects.select_related('customer', 'created_by').all().order_by('-created_at')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        quotations = quotations.filter(
            Q(quotation_number__icontains=search) |
            Q(urs__project_name__icontains=search) |
            Q(urs__customer__name__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        quotations = quotations.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(quotations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_quotations = Quotation.objects.count()
    draft_quotations = Quotation.objects.filter(status='draft').count()
    shared_quotations = Quotation.objects.filter(status='shared').count()
    accepted_quotations = Quotation.objects.filter(status='accepted').count()
    rejected_quotations = Quotation.objects.filter(status='rejected').count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'total_quotations': total_quotations,
        'draft_quotations': draft_quotations,
        'shared_quotations': shared_quotations,
        'accepted_quotations': accepted_quotations,
        'rejected_quotations': rejected_quotations,
    }
    return render(request, 'marketing/quotation_list.html', context)

@login_required
def quotation_detail(request, quotation_id):
    """View Quotation details"""
    quotation = get_object_or_404(Quotation, id=quotation_id)
    related_negotiations = Negotiation.objects.filter(quotation=quotation)
    related_purchase_orders = PurchaseOrder.objects.filter(quotation=quotation)
    
    context = {
        'quotation': quotation,
        'related_negotiations': related_negotiations,
        'related_purchase_orders': related_purchase_orders,
    }
    return render(request, 'marketing/quotation_detail.html', context)

@login_required
def technical_discussion(request):
    """Record Technical Discussion"""
    if request.method == 'POST':
        quotation_id = request.POST.get('quotation')
        discussion_date = request.POST.get('discussion_date')
        discussion_type = request.POST.get('discussion_type')
        participants = request.POST.get('participants')
        discussion_points = request.POST.get('discussion_points')
        decisions_made = request.POST.get('decisions_made')
        action_items = request.POST.get('action_items')
        next_follow_up = request.POST.get('next_follow_up')
        
        user_info = get_user_info_dict(request)
        discussion = TechnicalDiscussion.objects.create(
            quotation_id=quotation_id,
            discussion_date=discussion_date,
            discussion_type=discussion_type,
            participants=participants,
            discussion_points=discussion_points,
            decisions_made=decisions_made,
            action_items=action_items,
            next_follow_up=next_follow_up,
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name']
        )
        
        messages.success(request, 'Technical discussion recorded successfully!')
        return redirect('marketing:technical_discussion_list')
    
    quotations = Quotation.objects.filter(status='shared')
    context = {
        'quotations': quotations,
    }
    return render(request, 'marketing/technical_discussion.html', context)

@login_required
def technical_discussion_list(request):
    """List all Technical Discussions"""
    discussions = TechnicalDiscussion.objects.select_related('quotation', 'recorded_by').all().order_by('-discussion_date')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        discussions = discussions.filter(
            Q(quotation__quotation_number__icontains=search) |
            Q(quotation__urs__customer__name__icontains=search) |
            Q(participants__icontains=search)
        )
    
    # Filter by type
    type_filter = request.GET.get('type', '')
    if type_filter:
        discussions = discussions.filter(discussion_type=type_filter)
    
    # Pagination
    paginator = Paginator(discussions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_discussions = TechnicalDiscussion.objects.count()
    telephonic_discussions = TechnicalDiscussion.objects.filter(discussion_type='telephonic').count()
    in_person_discussions = TechnicalDiscussion.objects.filter(discussion_type='in_person').count()
    video_discussions = TechnicalDiscussion.objects.filter(discussion_type='video').count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'type_filter': type_filter,
        'total_discussions': total_discussions,
        'telephonic_discussions': telephonic_discussions,
        'in_person_discussions': in_person_discussions,
        'video_discussions': video_discussions,
    }
    return render(request, 'marketing/technical_discussion_list.html', context)

@login_required
def negotiation_create(request):
    """Create Negotiation Record"""
    if request.method == 'POST':
        quotation_id = request.POST.get('quotation')
        negotiation_date = request.POST.get('negotiation_date')
        negotiation_type = request.POST.get('negotiation_type')
        participants = request.POST.get('participants')
        initial_offer = request.POST.get('initial_offer')
        customer_counter_offer = request.POST.get('customer_counter_offer')
        final_offer = request.POST.get('final_offer')
        discount_percentage = request.POST.get('discount_percentage')
        payment_terms = request.POST.get('payment_terms')
        delivery_terms = request.POST.get('delivery_terms')
        outcome = request.POST.get('outcome')
        notes = request.POST.get('notes')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        negotiation = Negotiation.objects.create(
            quotation_id=quotation_id,
            negotiation_date=negotiation_date,
            negotiation_type=negotiation_type,
            participants=participants,
            initial_offer=initial_offer,
            customer_counter_offer=customer_counter_offer,
            final_offer=final_offer,
            discount_percentage=discount_percentage,
            payment_terms=payment_terms,
            delivery_terms=delivery_terms,
            outcome=outcome,
            notes=notes,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
        )
        
        messages.success(request, 'Negotiation record created successfully!')
        return redirect('marketing:negotiation_list')
    
    quotations = Quotation.objects.filter(status='shared')
    context = {
        'quotations': quotations,
    }
    return render(request, 'marketing/negotiation_create.html', context)

@login_required
def negotiation_list(request):
    """List all Negotiations"""
    negotiations = Negotiation.objects.select_related('quotation', 'created_by').all().order_by('-negotiation_date')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        negotiations = negotiations.filter(
            Q(quotation__quotation_number__icontains=search) |
            Q(quotation__urs__customer__name__icontains=search) |
            Q(participants__icontains=search)
        )
    
    # Filter by outcome
    outcome_filter = request.GET.get('outcome', '')
    if outcome_filter:
        negotiations = negotiations.filter(outcome=outcome_filter)
    
    # Pagination
    paginator = Paginator(negotiations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_negotiations = Negotiation.objects.count()
    successful_negotiations = Negotiation.objects.filter(outcome='successful').count()
    pending_negotiations = Negotiation.objects.filter(outcome='pending').count()
    failed_negotiations = Negotiation.objects.filter(outcome='failed').count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'outcome_filter': outcome_filter,
        'total_negotiations': total_negotiations,
        'successful_negotiations': successful_negotiations,
        'pending_negotiations': pending_negotiations,
        'failed_negotiations': failed_negotiations,
    }
    return render(request, 'marketing/negotiation_list.html', context)


@login_required
def quotation_revision_timeline(request, quotation_id):
    """Show timeline of quotation revisions"""
    quotation = get_object_or_404(Quotation, id=quotation_id)
    revisions = QuotationRevision.objects.filter(quotation=quotation).order_by('-revision_date')
    
    # Get related negotiations
    negotiations = Negotiation.objects.filter(quotation=quotation).order_by('-negotiation_date')
    
    context = {
        'quotation': quotation,
        'revisions': revisions,
        'negotiations': negotiations,
        'page_title': f'Revision Timeline - {quotation.quotation_number}',
        'breadcrumb': ['Quotations', 'Revision Timeline'],
    }
    return render(request, 'marketing/quotation_revision_timeline.html', context)


@login_required
def create_quotation_revision(request, quotation_id):
    """Create a new quotation revision"""
    quotation = get_object_or_404(Quotation, id=quotation_id)
    
    if request.method == 'POST':
        revision_reason = request.POST.get('revision_reason')
        new_amount = request.POST.get('new_amount')
        changes_summary = request.POST.get('changes_summary')
        negotiation_id = request.POST.get('negotiation_id')
        
        # Get the next revision number
        last_revision = QuotationRevision.objects.filter(quotation=quotation).order_by('-revision_number').first()
        next_revision_number = (last_revision.revision_number + 1) if last_revision else 1
        
        # Create revision record
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        revision = QuotationRevision.objects.create(
            quotation=quotation,
            revision_number=next_revision_number,
            revision_reason=revision_reason,
            previous_amount=quotation.total_amount,
            new_amount=new_amount,
            changes_summary=changes_summary,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
            negotiation_id=negotiation_id if negotiation_id else None
        )
        
        # Update quotation
        quotation.total_amount = new_amount
        quotation.version = next_revision_number
        quotation.status = 'revised'
        quotation.updated_at = timezone.now()
        quotation.save()
        
        messages.success(request, f'Quotation revision {next_revision_number} created successfully!')
        return redirect('marketing:quotation_revision_timeline', quotation_id=quotation.id)
    
    # Get available negotiations for this quotation
    negotiations = Negotiation.objects.filter(quotation=quotation, outcome='pending')
    
    context = {
        'quotation': quotation,
        'negotiations': negotiations,
        'page_title': f'Create Revision - {quotation.quotation_number}',
        'breadcrumb': ['Quotations', 'Create Revision'],
    }
    return render(request, 'marketing/create_quotation_revision.html', context)


# Production Tracking Views
@login_required
def workorder_list(request):
    """Work Order List View"""
    work_orders = WorkOrder.objects.select_related('purchase_order__customer', 'assigned_to').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        work_orders = work_orders.filter(
            Q(work_order_number__icontains=search_query) |
            Q(purchase_order__po_number__icontains=search_query) |
            Q(purchase_order__customer__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        work_orders = work_orders.filter(status=status_filter)
    
    # Filter by department
    department_filter = request.GET.get('department', '')
    if department_filter:
        work_orders = work_orders.filter(department=department_filter)
    
    # Filter by assigned user
    assigned_filter = request.GET.get('assigned_to', '')
    if assigned_filter:
        work_orders = work_orders.filter(assigned_to_id=assigned_filter)
    
    # Pagination
    paginator = Paginator(work_orders, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'department_filter': department_filter,
        'assigned_filter': assigned_filter,
        'status_choices': WorkOrder.STATUS_CHOICES,
        'department_choices': WorkOrder.DEPARTMENT_CHOICES,
        'users': get_user_model().objects.all(),
    }
    
    return render(request, 'marketing/workorder_list.html', context)

@login_required
def workorder_create(request):
    """Create Work Order View"""
    if request.method == 'POST':
        try:
            # Get form data
            work_order_number = request.POST.get('work_order_number')
            purchase_order_id = request.POST.get('purchase_order')
            department = request.POST.get('department')
            priority = request.POST.get('priority', 'medium')
            assigned_to_id = request.POST.get('assigned_to', '')
            due_date = request.POST.get('due_date')
            description = request.POST.get('description', '')
            special_instructions = request.POST.get('special_instructions', '')
            
            # Get user info from HRMS session
            user_info = get_user_info_dict(request)
            
            # Create Work Order

            work_order = WorkOrder.objects.create(
                work_order_number=work_order_number,
                purchase_order_id=purchase_order_id,
                department=department,
                priority=priority,
                assigned_to_id=assigned_to_id if assigned_to_id else None,
                due_date=due_date,
                description=description,
                special_instructions=special_instructions,
                # Get user info from HRMS session


                

                # Store HRMS user info

                created_by_user_id=user_info['user_id'],

                created_by_username=user_info['username'],

                created_by_email=user_info['email'],

                created_by_full_name=user_info['full_name'],
            )
            
            messages.success(request, f'Work Order {work_order_number} created successfully!')
            return redirect('marketing:workorder_detail', workorder_id=work_order.id)
            
        except Exception as e:
            messages.error(request, f'Error creating Work Order: {str(e)}')
    
    # Get purchase orders and users for the form
    purchase_orders = PurchaseOrder.objects.filter(status__in=['received', 'verified', 'approved']).order_by('-created_at')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'purchase_orders': purchase_orders,
        'users': users,
        'page_title': 'Create Work Order',
        'breadcrumb': ['Work Orders', 'Create Work Order'],
    }
    return render(request, 'marketing/workorder_create.html', context)

@login_required
def workorder_detail(request, workorder_id):
    """Work Order Detail View"""
    work_order = get_object_or_404(WorkOrder.objects.select_related('purchase_order__customer', 'assigned_to'), id=workorder_id)
    
    # Get related manufacturing records
    manufacturing_records = Manufacturing.objects.filter(work_order=work_order).order_by('-created_at')
    
    context = {
        'work_order': work_order,
        'manufacturing_records': manufacturing_records,
    }
    
    return render(request, 'marketing/workorder_detail.html', context)

@login_required
def manufacturing_list(request):
    """Manufacturing List View"""
    manufacturing_records = Manufacturing.objects.select_related('work_order__purchase_order__customer').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        manufacturing_records = manufacturing_records.filter(
            Q(batch_number__icontains=search_query) |
            Q(work_order__work_order_number__icontains=search_query) |
            Q(work_order__purchase_order__customer__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        manufacturing_records = manufacturing_records.filter(status=status_filter)
    
    # Filter by machine
    machine_filter = request.GET.get('machine', '')
    if machine_filter:
        manufacturing_records = manufacturing_records.filter(machine=machine_filter)
    
    # Pagination
    paginator = Paginator(manufacturing_records, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'machine_filter': machine_filter,
        'status_choices': Manufacturing.STATUS_CHOICES,
    }
    
    return render(request, 'marketing/manufacturing_list.html', context)

@login_required
def manufacturing_detail(request, manufacturing_id):
    """Manufacturing Detail View"""
    manufacturing = get_object_or_404(Manufacturing.objects.select_related('work_order__purchase_order__customer'), id=manufacturing_id)
    
    context = {
        'manufacturing': manufacturing,
    }
    
    return render(request, 'marketing/manufacturing_detail.html', context)

@login_required
def qc_create(request):
    """Create QC Record View"""
    if request.method == 'POST':
        try:
            # Get form data
            qc_number = request.POST.get('qc_number')
            manufacturing_id = request.POST.get('manufacturing')
            inspection_type = request.POST.get('inspection_type')
            qc_date = request.POST.get('qc_date')
            inspector_id = request.POST.get('inspector', '')
            status = request.POST.get('status', 'pending')
            test_results = request.POST.get('test_results', '')
            defects_found = request.POST.get('defects_found', '')
            corrective_actions = request.POST.get('corrective_actions', '')
            notes = request.POST.get('notes', '')
            
            # Get user info from HRMS session
            user_info = get_user_info_dict(request)
            
            # Create QC Record
            qc_record = QCTracking.objects.create(
                qc_number=qc_number,
                manufacturing_id=manufacturing_id,
                inspection_type=inspection_type,
                qc_date=qc_date,
                inspector_id=inspector_id if inspector_id else None,
                status=status,
                test_results=test_results,
                defects_found=defects_found,
                corrective_actions=corrective_actions,
                notes=notes,
                # Get user info from HRMS session


                

                # Store HRMS user info

                created_by_user_id=user_info['user_id'],

                created_by_username=user_info['username'],

                created_by_email=user_info['email'],

                created_by_full_name=user_info['full_name'],
            )
            
            messages.success(request, f'QC Record {qc_number} created successfully!')
            return redirect('marketing:qc_tracking')
            
        except Exception as e:
            messages.error(request, f'Error creating QC Record: {str(e)}')
    
    # Get manufacturing records and users for the form
    manufacturing_records = Manufacturing.objects.filter(status__in=['started', 'in_progress', 'qc_started']).order_by('-created_at')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'manufacturing_records': manufacturing_records,
        'users': users,
        'page_title': 'Create QC Record',
        'breadcrumb': ['QC Tracking', 'Create QC Record'],
    }
    return render(request, 'marketing/qc_create.html', context)

@login_required
def qc_tracking(request):
    """QC Tracking Interface"""
    # Get QC records
    qc_records = QCTracking.objects.select_related('manufacturing__work_order__purchase_order__customer').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        qc_records = qc_records.filter(
            Q(qc_number__icontains=search_query) |
            Q(manufacturing__batch_number__icontains=search_query) |
            Q(manufacturing__work_order__purchase_order__customer__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        qc_records = qc_records.filter(status=status_filter)
    
    # Filter by inspection type
    inspection_filter = request.GET.get('inspection_type', '')
    if inspection_filter:
        qc_records = qc_records.filter(inspection_type=inspection_filter)
    
    # Calculate statistics
    total_qc = qc_records.count()
    pending_qc = qc_records.filter(status='pending').count()
    in_progress_qc = qc_records.filter(status='in_progress').count()
    passed_qc = qc_records.filter(status='passed').count()
    failed_qc = qc_records.filter(status='failed').count()
    rework_qc = qc_records.filter(status='rework').count()
    
    # Pagination
    paginator = Paginator(qc_records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'qc_records': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'inspection_filter': inspection_filter,
        'total_qc': total_qc,
        'pending_qc': pending_qc,
        'in_progress_qc': in_progress_qc,
        'passed_qc': passed_qc,
        'failed_qc': failed_qc,
        'rework_qc': rework_qc,
    }
    return render(request, 'marketing/qc_tracking.html', context)

@login_required
def qc_export(request):
    """Export QC Records to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    import io
    
    # Get QC records with same filters as the main view
    qc_records = QCTracking.objects.select_related('manufacturing__work_order__purchase_order__customer', 'inspector').order_by('-created_at')
    
    # Apply same filters as main view
    search_query = request.GET.get('search', '')
    if search_query:
        qc_records = qc_records.filter(
            Q(qc_number__icontains=search_query) |
            Q(manufacturing__batch_number__icontains=search_query) |
            Q(manufacturing__work_order__purchase_order__customer__name__icontains=search_query)
        )
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        qc_records = qc_records.filter(status=status_filter)
    
    inspection_filter = request.GET.get('inspection_type', '')
    if inspection_filter:
        qc_records = qc_records.filter(inspection_type=inspection_filter)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Records"
    
    # Define headers
    headers = [
        'QC Number', 'Batch Number', 'Customer', 'Inspection Type', 'Status',
        'QC Date', 'Inspector', 'Test Results', 'Defects Found', 'Corrective Actions',
        'Notes', 'Created By', 'Created At'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, qc in enumerate(qc_records, 2):
        ws.cell(row=row, column=1, value=qc.qc_number)
        ws.cell(row=row, column=2, value=qc.manufacturing.batch_number)
        ws.cell(row=row, column=3, value=qc.manufacturing.work_order.purchase_order.customer.name)
        ws.cell(row=row, column=4, value=qc.get_inspection_type_display())
        ws.cell(row=row, column=5, value=qc.get_status_display())
        ws.cell(row=row, column=6, value=qc.qc_date.strftime('%Y-%m-%d') if qc.qc_date else '')
        ws.cell(row=row, column=7, value=qc.inspector.get_full_name() if qc.inspector else 'Not Assigned')
        ws.cell(row=row, column=8, value=qc.test_results)
        ws.cell(row=row, column=9, value=qc.defects_found)
        ws.cell(row=row, column=10, value=qc.corrective_actions)
        ws.cell(row=row, column=11, value=qc.notes)
        ws.cell(row=row, column=12, value=qc.created_by.get_full_name() if qc.created_by else qc.created_by.username)
        ws.cell(row=row, column=13, value=qc.created_at.strftime('%Y-%m-%d %H:%M') if qc.created_at else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with current date
    from datetime import datetime
    current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'QC_Records_Export_{current_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def dispatch_list(request):
    """Dispatch List View"""
    dispatches = Dispatch.objects.select_related('work_order__purchase_order__customer').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        dispatches = dispatches.filter(
            Q(dispatch_number__icontains=search_query) |
            Q(work_order__work_order_number__icontains=search_query) |
            Q(work_order__purchase_order__customer__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        dispatches = dispatches.filter(status=status_filter)
    
    # Filter by transporter
    transporter_filter = request.GET.get('transporter', '')
    if transporter_filter:
        dispatches = dispatches.filter(transporter__icontains=transporter_filter)
    
    # Pagination
    paginator = Paginator(dispatches, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'transporter_filter': transporter_filter,
        'status_choices': Dispatch.STATUS_CHOICES,
    }
    
    return render(request, 'marketing/dispatch_list.html', context)

@login_required
def dispatch_detail(request, dispatch_id):
    """Dispatch Detail View"""
    dispatch = get_object_or_404(Dispatch.objects.select_related('work_order__purchase_order__customer'), id=dispatch_id)
    
    context = {
        'dispatch': dispatch,
    }
    
    return render(request, 'marketing/dispatch_detail.html', context)

@login_required
def exhibition_list(request):
    """Exhibition List View"""
    return render(request, 'marketing/exhibition_list.html')

@login_required
def exhibition_create(request):
    """Create Exhibition View"""
    return render(request, 'marketing/exhibition_create.html')

@login_required
def exhibition_detail(request, exhibition_id):
    """Exhibition Detail View"""
    return render(request, 'marketing/exhibition_detail.html')

@login_required
def exhibition_planning(request):
    """Exhibition Planning View"""
    return render(request, 'marketing/exhibition_planning.html')

@login_required
def visitor_database(request):
    """Visitor Database View"""
    return render(request, 'marketing/visitor_database.html')

@login_required
def visitor_create(request):
    """Create Visitor View"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            company = request.POST.get('company')
            designation = request.POST.get('designation', '')
            email = request.POST.get('email')
            phone = request.POST.get('phone', '')
            industry = request.POST.get('industry', '')
            exhibition_id = request.POST.get('exhibition', '')
            status = request.POST.get('status', 'new')
            notes = request.POST.get('notes', '')
            
            # Create Visitor (assuming we have a Visitor model)
            # For now, we'll create a simple success message
            messages.success(request, f'Visitor {name} created successfully!')
            return redirect('marketing:visitor_database')
            
        except Exception as e:
            messages.error(request, f'Error creating visitor: {str(e)}')
    
    # Get exhibitions for the form
    exhibitions = Exhibition.objects.all().order_by('-start_date')
    
    context = {
        'exhibitions': exhibitions,
        'page_title': 'Add Visitor',
        'breadcrumb': ['Visitor Database', 'Add Visitor'],
    }
    return render(request, 'marketing/visitor_create.html', context)

@login_required
def visitor_export(request):
    """Export Visitor Data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    import io
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitor Database"
    
    # Define headers
    headers = [
        'Name', 'Company', 'Designation', 'Email', 'Phone', 'Industry',
        'Exhibition', 'Status', 'Visit Date', 'Notes', 'Created At'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get actual visitor data from Lead model (leads from events/exhibitions)
    visitors = Lead.objects.filter(source='event').order_by('-created_at')
    
    # Write actual data
    for row, lead in enumerate(visitors, 2):
        # Get associated exhibition if any
        exhibition_name = 'N/A'
        if lead.campaign:
            # Try to find exhibition by date
            exhibition = Exhibition.objects.filter(
                start_date__lte=lead.created_at.date(),
                end_date__gte=lead.created_at.date()
            ).first()
            if exhibition:
                exhibition_name = exhibition.name
        
        visitor_data = [
            lead.full_name,
            lead.company or 'N/A',
            lead.position or 'N/A',
            lead.email,
            lead.phone or 'N/A',
            'N/A',  # Industry - not in Lead model
            exhibition_name,
            lead.get_status_display(),
            lead.created_at.strftime('%Y-%m-%d'),
            lead.notes or '',
            lead.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col, value in enumerate(visitor_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with current date
    from datetime import datetime
    current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Visitor_Database_Export_{current_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def expense_list(request):
    """Expense List View"""
    return render(request, 'marketing/expense_list.html')

@login_required
def expense_create(request):
    """Create Expense View"""
    return render(request, 'marketing/expense_form.html')

@login_required
def expense_detail(request, expense_id):
    """Expense Detail View"""
    return render(request, 'marketing/expense_detail.html')

@login_required
def expense_approval(request):
    """Expense Approval View"""
    return render(request, 'marketing/expense_approval.html')

@login_required
def expense_reports(request):
    """Expense Reports View"""
    return render(request, 'marketing/expense_reports.html')

# Reports Views
@login_required
def daily_reports(request):
    """Daily Reports Dashboard"""
    from datetime import datetime, timedelta
    
    # Get today's date
    today = datetime.now().date()
    
    # Get daily statistics
    daily_stats = {
        'new_customers': Customer.objects.filter(created_at__date=today).count(),
        'new_leads': Lead.objects.filter(created_at__date=today).count(),
        'new_quotations': Quotation.objects.filter(created_at__date=today).count(),
        'new_purchase_orders': PurchaseOrder.objects.filter(received_date=today).count(),
        'new_work_orders': WorkOrder.objects.filter(created_at__date=today).count(),
        'completed_manufacturing': Manufacturing.objects.filter(status='completed', completion_date=today).count(),
        'new_visits': Visit.objects.filter(scheduled_date__date=today).count(),
        'pending_expenses': Expense.objects.filter(status='pending').count(),
    }
    
    # Get recent activities for today
    recent_activities = []
    
    # Add recent customers
    recent_customers = Customer.objects.filter(created_at__date=today)[:5]
    for customer in recent_customers:
        recent_activities.append({
            'type': 'customer',
            'title': f'New Customer: {customer.name}',
            'description': f'Added by {customer.created_by.get_full_name() if customer.created_by else "System"}',
            'time': customer.created_at,
            'icon': 'users',
            'color': 'blue'
        })
    
    # Add recent leads
    recent_leads = Lead.objects.filter(created_at__date=today)[:5]
    for lead in recent_leads:
        recent_activities.append({
            'type': 'lead',
            'title': f'New Lead: {lead.first_name} {lead.last_name}',
            'description': f'From {lead.company} - {lead.get_source_display()}',
            'time': lead.created_at,
            'icon': 'user-plus',
            'color': 'green'
        })
    
    # Add recent quotations
    recent_quotations = Quotation.objects.filter(created_at__date=today)[:5]
    for quotation in recent_quotations:
        recent_activities.append({
            'type': 'quotation',
            'title': f'New Quotation: {quotation.quotation_number}',
            'description': f'For {quotation.customer.name} - ₹{quotation.total_amount}',
            'time': quotation.created_at,
            'icon': 'file-text',
            'color': 'yellow'
        })
    
    # Sort activities by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    
    context = {
        'daily_stats': daily_stats,
        'recent_activities': recent_activities[:10],
        'today': today,
    }
    
    return render(request, 'marketing/daily_reports.html', context)

@login_required
def monthly_reports(request):
    """Monthly Reports Dashboard"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Avg
    
    # Get current month
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    
    # Monthly statistics
    monthly_stats = {
        'total_customers': Customer.objects.filter(created_at__month=current_month, created_at__year=current_year).count(),
        'total_leads': Lead.objects.filter(created_at__month=current_month, created_at__year=current_year).count(),
        'total_quotations': Quotation.objects.filter(created_at__month=current_month, created_at__year=current_year).count(),
        'total_purchase_orders': PurchaseOrder.objects.filter(received_date__month=current_month, received_date__year=current_year).count(),
        'total_revenue': PurchaseOrder.objects.filter(
            received_date__month=current_month, 
            received_date__year=current_year,
            status__in=['received', 'verified', 'approved', 'in_production', 'completed', 'dispatched', 'delivered']
        ).aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_work_orders': WorkOrder.objects.filter(created_at__month=current_month, created_at__year=current_year).count(),
        'completed_manufacturing': Manufacturing.objects.filter(
            status='completed', 
            completion_date__month=current_month, 
            completion_date__year=current_year
        ).count(),
        'total_visits': Visit.objects.filter(scheduled_date__month=current_month, scheduled_date__year=current_year).count(),
    }
    
    # Regional performance
    regional_performance = []
    regions = Region.objects.all()
    for region in regions:
        customer_count = Customer.objects.filter(region=region, created_at__month=current_month, created_at__year=current_year).count()
        lead_count = Lead.objects.filter(created_at__month=current_month, created_at__year=current_year).count()  # Assuming leads are not region-specific
        regional_performance.append({
            'region': region,
            'customers': customer_count,
            'leads': lead_count,
        })
    
    # Top performing customers
    top_customers = Customer.objects.annotate(
        po_count=Count('purchaseorder'),
        total_spent=Sum('purchaseorder__total_amount')
    ).filter(
        purchaseorder__received_date__month=current_month,
        purchaseorder__received_date__year=current_year
    ).order_by('-total_spent')[:5]
    
    context = {
        'monthly_stats': monthly_stats,
        'regional_performance': regional_performance,
        'top_customers': top_customers,
        'current_month': current_month,
        'current_year': current_year,
    }
    
    return render(request, 'marketing/monthly_reports.html', context)

@login_required
def customer_reports(request):
    """Customer Reports and Analytics"""
    from django.db.models import Count, Sum, Avg
    from datetime import datetime, timedelta
    
    # Get date range filter
    date_filter = request.GET.get('date_range', 'all')
    today = datetime.now().date()
    
    if date_filter == 'week':
        start_date = today - timedelta(days=7)
        customers = Customer.objects.filter(created_at__date__gte=start_date)
    elif date_filter == 'month':
        start_date = today - timedelta(days=30)
        customers = Customer.objects.filter(created_at__date__gte=start_date)
    elif date_filter == 'quarter':
        start_date = today - timedelta(days=90)
        customers = Customer.objects.filter(created_at__date__gte=start_date)
    else:
        customers = Customer.objects.all()
    
    # Customer statistics
    customer_stats = {
        'total_customers': customers.count(),
        'active_customers': customers.filter(customer_type='existing').count(),
        'new_customers': customers.filter(created_at__date=today).count(),
        'customers_with_orders': customers.filter(purchaseorder__isnull=False).distinct().count(),
    }
    
    # Regional distribution
    regional_distribution = customers.values('region__name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Customer type distribution
    customer_type_distribution = customers.values('customer_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Top customers by order value
    top_customers = customers.annotate(
        total_orders=Count('purchaseorder'),
        total_value=Sum('purchaseorder__total_amount')
    ).filter(total_value__isnull=False).order_by('-total_value')[:10]
    
    # Customer growth trend (last 12 months)
    growth_trend = []
    for i in range(12):
        month_date = today - timedelta(days=30*i)
        month_customers = Customer.objects.filter(
            created_at__month=month_date.month,
            created_at__year=month_date.year
        ).count()
        growth_trend.append({
            'month': month_date.strftime('%b %Y'),
            'customers': month_customers
        })
    growth_trend.reverse()
    
    context = {
        'customer_stats': customer_stats,
        'regional_distribution': regional_distribution,
        'customer_type_distribution': customer_type_distribution,
        'top_customers': top_customers,
        'growth_trend': growth_trend,
        'date_filter': date_filter,
    }
    
    return render(request, 'marketing/customer_reports.html', context)

@login_required
def performance_analytics(request):
    """Performance Analytics Dashboard"""
    from django.db.models import Count, Sum, Avg
    from datetime import datetime, timedelta
    
    # Get date range filter
    date_filter = request.GET.get('date_range', 'month')
    today = datetime.now().date()
    
    if date_filter == 'week':
        start_date = today - timedelta(days=7)
    elif date_filter == 'month':
        start_date = today - timedelta(days=30)
    elif date_filter == 'quarter':
        start_date = today - timedelta(days=90)
    else:
        start_date = today - timedelta(days=365)
    
    # Sales performance
    sales_performance = {
        'total_quotations': Quotation.objects.filter(created_at__date__gte=start_date).count(),
        'accepted_quotations': Quotation.objects.filter(status='accepted', created_at__date__gte=start_date).count(),
        'conversion_rate': 0,
        'total_revenue': PurchaseOrder.objects.filter(
            received_date__gte=start_date,
            status__in=['received', 'verified', 'approved', 'in_production', 'completed', 'dispatched', 'delivered']
        ).aggregate(total=Sum('total_amount'))['total'] or 0,
        'avg_order_value': 0,
    }
    
    if sales_performance['total_quotations'] > 0:
        sales_performance['conversion_rate'] = (sales_performance['accepted_quotations'] / sales_performance['total_quotations']) * 100
    
    # Production performance
    production_performance = {
        'total_work_orders': WorkOrder.objects.filter(created_at__date__gte=start_date).count(),
        'completed_work_orders': WorkOrder.objects.filter(status='completed', created_at__date__gte=start_date).count(),
        'on_time_delivery': 0,
        'avg_production_time': 0,
    }
    
    # Lead performance
    lead_performance = {
        'total_leads': Lead.objects.filter(created_at__date__gte=start_date).count(),
        'qualified_leads': Lead.objects.filter(status='qualified', created_at__date__gte=start_date).count(),
        'converted_leads': Lead.objects.filter(status='converted', created_at__date__gte=start_date).count(),
        'lead_conversion_rate': 0,
    }
    
    if lead_performance['total_leads'] > 0:
        lead_performance['lead_conversion_rate'] = (lead_performance['converted_leads'] / lead_performance['total_leads']) * 100
    
    # User performance (using RBAC usernames)
    user_performance = []
    # Get unique usernames from all activities in the date range
    customer_usernames = set(Customer.objects.filter(created_at__date__gte=start_date).values_list('created_by_username', flat=True).exclude(created_by_username=''))
    lead_usernames = set(Lead.objects.filter(created_at__date__gte=start_date).values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    quotation_usernames = set(Quotation.objects.filter(created_at__date__gte=start_date).values_list('created_by_username', flat=True).exclude(created_by_username=''))
    visit_usernames = set(Visit.objects.filter(scheduled_date__date__gte=start_date).values_list('assigned_to_username', flat=True).exclude(assigned_to_username=''))
    all_usernames = customer_usernames | lead_usernames | quotation_usernames | visit_usernames
    
    for username in all_usernames:
        user_stats = {
            'username': username,
            'customers_created': Customer.objects.filter(created_by_username=username, created_at__date__gte=start_date).count(),
            'leads_created': Lead.objects.filter(assigned_to_username=username, created_at__date__gte=start_date).count(),
            'quotations_created': Quotation.objects.filter(created_by_username=username, created_at__date__gte=start_date).count(),
            'visits_conducted': Visit.objects.filter(assigned_to_username=username, scheduled_date__date__gte=start_date).count(),
        }
        user_performance.append(user_stats)
    
    # Performance trends (last 6 periods)
    trends = []
    for i in range(6):
        period_start = start_date - timedelta(days=30*i)
        period_end = period_start + timedelta(days=30)
        
        period_stats = {
            'period': period_start.strftime('%b %Y'),
            'quotations': Quotation.objects.filter(created_at__date__gte=period_start, created_at__date__lt=period_end).count(),
            'orders': PurchaseOrder.objects.filter(received_date__gte=period_start, received_date__lt=period_end).count(),
            'revenue': PurchaseOrder.objects.filter(
                received_date__gte=period_start,
                received_date__lt=period_end,
                status__in=['received', 'verified', 'approved', 'in_production', 'completed', 'dispatched', 'delivered']
            ).aggregate(total=Sum('total_amount'))['total'] or 0,
        }
        trends.append(period_stats)
    trends.reverse()
    
    context = {
        'sales_performance': sales_performance,
        'production_performance': production_performance,
        'lead_performance': lead_performance,
        'user_performance': user_performance,
        'trends': trends,
        'date_filter': date_filter,
    }
    
    return render(request, 'marketing/performance_analytics.html', context)

@login_required
def export_reports(request):
    """Export Reports Interface"""
    from datetime import datetime
    
    # Get available report types
    report_types = [
        {'id': 'daily', 'name': 'Daily Report', 'description': 'Daily activity summary'},
        {'id': 'monthly', 'name': 'Monthly Report', 'description': 'Monthly performance summary'},
        {'id': 'customer', 'name': 'Customer Report', 'description': 'Customer analysis and statistics'},
        {'id': 'sales', 'name': 'Sales Report', 'description': 'Sales performance and pipeline'},
        {'id': 'production', 'name': 'Production Report', 'description': 'Production and manufacturing status'},
        {'id': 'financial', 'name': 'Financial Report', 'description': 'Revenue and financial metrics'},
    ]
    
    # Get available formats
    export_formats = [
        {'id': 'pdf', 'name': 'PDF', 'icon': 'file-text'},
        {'id': 'excel', 'name': 'Excel', 'icon': 'table'},
        {'id': 'csv', 'name': 'CSV', 'icon': 'file'},
    ]
    
    context = {
        'report_types': report_types,
        'export_formats': export_formats,
    }
    
    return render(request, 'marketing/export_reports.html', context)

@login_required
def export_data(request):
    """Export Data View"""
    return render(request, 'marketing/export_data.html')

# Settings & User Management Views
@login_required
def user_management(request):
    """User Management Dashboard"""
    from django.db.models import Count, Q
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    
    # Get all users
    users = get_user_model().objects.all()
    
    # Apply search filter
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Apply role filter (assuming you have a role field or groups)
    if role_filter:
        users = users.filter(groups__name=role_filter)
    
    # Apply status filter
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    # Get user statistics
    user_stats = {
        'total_users': get_user_model().objects.count(),
        'active_users': get_user_model().objects.filter(is_active=True).count(),
        'inactive_users': get_user_model().objects.filter(is_active=False).count(),
        'new_users_this_month': get_user_model().objects.filter(
            date_joined__month=datetime.now().month,
            date_joined__year=datetime.now().year
        ).count(),
    }
    
    # Get role distribution
    from django.contrib.auth.models import Group
    role_distribution = []
    groups = Group.objects.all()
    for group in groups:
        user_count = users.filter(groups=group).count()
        if user_count > 0:
            role_distribution.append({
                'role': group.name,
                'count': user_count,
                'percentage': (user_count / user_stats['total_users']) * 100 if user_stats['total_users'] > 0 else 0
            })
    
    # Pagination
    paginator = Paginator(users, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'user_stats': user_stats,
        'role_distribution': role_distribution,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'groups': groups,
    }
    
    return render(request, 'marketing/user_management.html', context)

@login_required
def user_create(request):
    """Create New User"""
    from django.contrib.auth.forms import UserCreationForm
    from django.contrib.auth.models import Group
    
    if request.method == 'POST':
        try:
            # Create user
            user = get_user_model().objects.create_user(
                username=request.POST.get('username'),
                email=request.POST.get('email'),
                password=request.POST.get('password1'),
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                is_active=request.POST.get('is_active') == 'on'
            )
            
            # Assign groups/roles
            selected_groups = request.POST.getlist('groups')
            for group_id in selected_groups:
                group = Group.objects.get(id=group_id)
                user.groups.add(group)
            
            messages.success(request, f'User "{user.username}" created successfully!')
            return redirect('marketing:user_management')
        except Exception as e:
            messages.error(request, f'Error creating user: {str(e)}')
    
    context = {
        'groups': Group.objects.all(),
    }
    
    return render(request, 'marketing/user_form.html', context)

@login_required
def user_edit(request, user_id):
    """Edit User"""
    from django.contrib.auth.models import Group
    
    try:
        user = get_user_model().objects.get(id=user_id)
    except get_user_model().DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('marketing:user_management')
    
    if request.method == 'POST':
        try:
            # Update user fields
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name', '')
            user.last_name = request.POST.get('last_name', '')
            user.is_active = request.POST.get('is_active') == 'on'
            
            # Update password if provided
            if request.POST.get('password1'):
                user.set_password(request.POST.get('password1'))
            
            user.save()
            
            # Update groups/roles
            user.groups.clear()
            selected_groups = request.POST.getlist('groups')
            for group_id in selected_groups:
                group = Group.objects.get(id=group_id)
                user.groups.add(group)
            
            messages.success(request, f'User "{user.username}" updated successfully!')
            return redirect('marketing:user_management')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')
    
    context = {
        'user': user,
        'groups': Group.objects.all(),
    }
    
    return render(request, 'marketing/user_form.html', context)

@login_required
def user_delete(request, user_id):
    """Delete User"""
    if request.method == 'POST':
        try:
            user = get_user_model().objects.get(id=user_id)
            username = user.username
            
            # Prevent deleting superusers or the current user
            if user.is_superuser:
                messages.error(request, 'Cannot delete superuser accounts.')
                return redirect('marketing:user_management')
            
            if user == request.user:
                messages.error(request, 'You cannot delete your own account.')
                return redirect('marketing:user_management')
            
            # Delete the user
            user.delete()
            messages.success(request, f'User "{username}" has been deleted successfully!')
            
        except get_user_model().DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'Error deleting user: {str(e)}')
    
    return redirect('marketing:user_management')

@login_required
def user_toggle_status(request, user_id):
    """Toggle User Active Status"""
    if request.method == 'POST':
        try:
            user = get_user_model().objects.get(id=user_id)
            
            # Get the activate parameter
            activate_param = request.POST.get('activate', '')
            print(f"DEBUG: activate parameter received: '{activate_param}'")  # Debug line
            
            # Prevent deactivating superusers or the current user
            if user.is_superuser and activate_param == 'false':
                messages.error(request, 'Cannot deactivate superuser accounts.')
                return redirect('marketing:user_management')
            
            if user == request.user and activate_param == 'false':
                messages.error(request, 'You cannot deactivate your own account.')
                return redirect('marketing:user_management')
            
            # Toggle status
            activate = activate_param == 'true'
            print(f"DEBUG: Current user.is_active: {user.is_active}, Setting to: {activate}")  # Debug line
            user.is_active = activate
            user.save()
            
            status = 'activated' if activate else 'deactivated'
            messages.success(request, f'User "{user.username}" has been {status} successfully!')
            
        except get_user_model().DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'Error updating user status: {str(e)}')
            print(f"DEBUG: Exception occurred: {str(e)}")  # Debug line
    
    return redirect('marketing:user_management')

@login_required
def user_profile(request):
    """User Profile Management"""
    user = request.user
    
    if request.method == 'POST':
        try:
            # Update profile information
            user.first_name = request.POST.get('first_name', '')
            user.last_name = request.POST.get('last_name', '')
            user.email = request.POST.get('email')
            
            # Update password if provided
            if request.POST.get('new_password1'):
                if user.check_password(request.POST.get('current_password')):
                    user.set_password(request.POST.get('new_password1'))
                else:
                    messages.error(request, 'Current password is incorrect.')
                    return render(request, 'marketing/user_profile.html', {'user': user})
            
            user.save()
            messages.success(request, 'Profile updated successfully!')
        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
    
    context = {
        'user': user,
    }
    
    return render(request, 'marketing/user_profile.html', context)



@login_required
def notification_settings(request):
    """Notification Settings Management"""
    # Get current notification settings (placeholder)
    notification_settings = {
        'email_notifications': True,
        'sms_notifications': False,
        'daily_reports': True,
        'weekly_reports': True,
        'monthly_reports': True,
        'lead_notifications': True,
        'quotation_notifications': True,
        'order_notifications': True,
        'expense_notifications': True,
        'system_alerts': True,
    }
    
    if request.method == 'POST':
        # Update notification settings
        notification_settings.update({
            'email_notifications': request.POST.get('email_notifications') == 'on',
            'sms_notifications': request.POST.get('sms_notifications') == 'on',
            'daily_reports': request.POST.get('daily_reports') == 'on',
            'weekly_reports': request.POST.get('weekly_reports') == 'on',
            'monthly_reports': request.POST.get('monthly_reports') == 'on',
            'lead_notifications': request.POST.get('lead_notifications') == 'on',
            'quotation_notifications': request.POST.get('quotation_notifications') == 'on',
            'order_notifications': request.POST.get('order_notifications') == 'on',
            'expense_notifications': request.POST.get('expense_notifications') == 'on',
            'system_alerts': request.POST.get('system_alerts') == 'on',
        })
        messages.success(request, 'Notification settings updated successfully!')
    
    context = {
        'notification_settings': notification_settings,
    }
    
    return render(request, 'marketing/notification_settings.html', context)

@login_required
def visit_detail(request, visit_id):
    """Visit Detail View"""
    visit = get_object_or_404(Visit, id=visit_id)
    
    context = {
        'visit': visit,
    }
    return render(request, 'marketing/visit_detail.html', context)

@login_required
def visit_tracking(request):
    """Live Visit Tracking Dashboard"""
    # Get current active visits
    active_visits = Visit.objects.filter(
        scheduled_date__date=timezone.now().date(),
        status='in_progress'
    ).select_related('customer', 'location')
    
    # Get today's completed visits
    completed_visits = Visit.objects.filter(
        scheduled_date__date=timezone.now().date(),
        status='completed'
    ).select_related('customer', 'location')
    
    # Get upcoming visits
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date__gt=timezone.now().date()
    ).select_related('customer', 'location').order_by('scheduled_date')[:10]
    
    context = {
        'active_visits': active_visits,
        'completed_visits': completed_visits,
        'upcoming_visits': upcoming_visits,
    }
    return render(request, 'marketing/visit_tracking.html', context)

@login_required
def visit_reports(request):
    """Visit Reports and Analytics"""
    # Get date range from request
    start_date = request.GET.get('start_date', (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))
    
    # Convert to date objects
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get visits in date range
    visits = Visit.objects.filter(
        scheduled_date__date__range=[start_date, end_date]
    ).select_related('customer', 'location')
    
    # Calculate statistics
    total_visits = visits.count()
    completed_visits = visits.filter(status='completed').count()
    pending_visits = visits.filter(status='scheduled').count()  # Changed from 'pending' to 'scheduled'
    cancelled_visits = visits.filter(status='cancelled').count()
    
    # Visits by type
    visits_by_type = visits.values('visit_type').annotate(count=Count('id'))
    
    # Visits by region
    visits_by_region = visits.values('customer__region__name').annotate(count=Count('id'))
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_visits': total_visits,
        'completed_visits': completed_visits,
        'pending_visits': pending_visits,
        'cancelled_visits': cancelled_visits,
        'visits_by_type': visits_by_type,
        'visits_by_region': visits_by_region,
        'visits': visits,
    }
    return render(request, 'marketing/visit_reports.html', context)

@login_required
def follow_up_reminders(request):
    """Follow-up Reminders Management"""
    # Get upcoming follow-ups
    upcoming_follow_ups = Visit.objects.filter(
        next_follow_up_date__date__gte=timezone.now().date(),
        next_follow_up_date__date__lte=timezone.now().date() + timedelta(days=7)
    ).select_related('customer', 'location').order_by('next_follow_up_date')
    
    # Get overdue follow-ups
    overdue_follow_ups = Visit.objects.filter(
        next_follow_up_date__date__lt=timezone.now().date(),
        status='completed'
    ).select_related('customer', 'location').order_by('next_follow_up_date')
    
    context = {
        'upcoming_follow_ups': upcoming_follow_ups,
        'overdue_follow_ups': overdue_follow_ups,
    }
    return render(request, 'marketing/follow_up_reminders.html', context)

@login_required
def customer_regions(request):
    """Customer Regions Management"""
    # Prefetch customers and their locations for each region
    regions = Region.objects.prefetch_related('customer_set', 'customer_set__locations').all()
    
    # Calculate statistics for each region
    for region in regions:
        region.customer_count = region.customer_set.count()
        region.total_customers = region.customer_set.count()
        # Get customer types breakdown
        region.prospects = region.customer_set.filter(customer_type='prospect').count()
        region.existing_customers = region.customer_set.filter(customer_type='existing').count()
        region.lapsed_customers = region.customer_set.filter(customer_type='lapsed').count()
    
    context = {
        'regions': regions,
    }
    return render(request, 'marketing/customer_regions.html', context)

@login_required
def customer_import(request):
    """Customer Import Interface"""
    if request.method == 'POST':
        # Handle file upload and import
        messages.success(request, 'Customer data imported successfully!')
        return redirect('marketing:customer_list')
    
    return render(request, 'marketing/customer_import.html')

@login_required
def lead_import(request):
    """Lead Import Interface"""
    if request.method == 'POST':
        # Handle file upload and import
        messages.success(request, 'Lead data imported successfully!')
        return redirect('marketing:lead_list')
    
    return render(request, 'marketing/lead_import.html')

@login_required
def lead_scoring(request):
    """Lead Scoring Interface"""
    leads = Lead.objects.all().order_by('-score')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        leads = leads.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(company__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(leads, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'leads': page_obj,
        'search_query': search_query,
    }
    return render(request, 'marketing/lead_scoring.html', context)

@login_required
def workorder_list(request):
    """Work Order List View"""
    work_orders = WorkOrder.objects.select_related('purchase_order__customer').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        work_orders = work_orders.filter(
            Q(work_order_number__icontains=search_query) |
            Q(purchase_order__po_number__icontains=search_query) |
            Q(purchase_order__customer__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        work_orders = work_orders.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(work_orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'work_orders': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'marketing/workorder_list.html', context)

# Duplicate functions removed - these are already defined earlier in the file

@login_required
def exhibition_planning(request):
    """Exhibition Planning View"""
    context = {}
    return render(request, 'marketing/exhibition_planning.html', context)

@login_required
def visitor_database(request):
    """Visitor Database View"""
    context = {}
    return render(request, 'marketing/visitor_database.html', context)

@login_required
def expense_create(request):
    """Expense Create View"""
    if request.method == 'POST':
        # Handle expense creation
        messages.success(request, 'Expense created successfully!')
        return redirect('marketing:expense_list')
    
    return render(request, 'marketing/expense_create.html')

@login_required
def expense_detail(request, expense_id):
    """Expense Detail View"""
    expense = get_object_or_404(Expense, id=expense_id)
    
    context = {
        'expense': expense,
    }
    return render(request, 'marketing/expense_detail.html', context)

@login_required
def expense_approval(request):
    """Expense Approval View"""
    context = {}
    return render(request, 'marketing/expense_approval.html', context)

@login_required
def expense_reports(request):
    """Expense Reports View"""
    context = {}
    return render(request, 'marketing/expense_reports.html', context)

@login_required
def daily_reports(request):
    """Daily Reports View"""
    context = {}
    return render(request, 'marketing/daily_reports.html', context)

@login_required
def monthly_reports(request):
    """Monthly Reports View"""
    context = {}
    return render(request, 'marketing/monthly_reports.html', context)

@login_required
def performance_analytics(request):
    """Performance Analytics View"""
    context = {}
    return render(request, 'marketing/performance_analytics.html', context)

@login_required
def export_data(request):
    """Export Data View"""
    context = {}
    return render(request, 'marketing/export_data.html', context)

@login_required
def email_templates(request):
    """Email Templates View"""
    context = {}
    return render(request, 'marketing/email_templates.html', context)

@login_required
def landing_pages(request):
    """Landing Pages View"""
    context = {}
    return render(request, 'marketing/landing_pages.html', context)

@login_required
def social_posts(request):
    """Social Posts View"""
    context = {}
    return render(request, 'marketing/social_posts.html', context)

@login_required
def campaign_performance(request):
    """Campaign Performance View"""
    context = {}
    return render(request, 'marketing/campaign_performance.html', context)

@login_required
def roi_analysis(request):
    """ROI Analysis View"""
    context = {}
    return render(request, 'marketing/roi_analysis.html', context)

@login_required
def conversion_funnel(request):
    """Conversion Funnel View"""
    context = {}
    return render(request, 'marketing/conversion_funnel.html', context)

@login_required
def workflows(request):
    """Workflows View"""
    context = {}
    return render(request, 'marketing/workflows.html', context)

@login_required
def po_list(request):
    """Purchase Order List View"""
    purchase_orders = PurchaseOrder.objects.select_related('customer').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        purchase_orders = purchase_orders.filter(
            Q(po_number__icontains=search_query) |
            Q(customer__name__icontains=search_query) |
            Q(quotation__quotation_number__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        purchase_orders = purchase_orders.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(purchase_orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'purchase_orders': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'marketing/po_list.html', context)

@login_required
def po_create(request):
    """Create Purchase Order View"""
    if request.method == 'POST':
        try:
            # Get form data
            po_number = request.POST.get('po_number')
            customer_id = request.POST.get('customer')
            quotation_id = request.POST.get('quotation', '')
            total_amount = request.POST.get('total_amount')
            received_date = request.POST.get('received_date')
            delivery_date = request.POST.get('delivery_date')
            payment_terms = request.POST.get('payment_terms', '')
            payment_method = request.POST.get('payment_method', '')
            payment_terms_declared = request.POST.get('payment_terms_declared', '')
            special_requirements = request.POST.get('special_requirements', '')
            
            # Get user info from HRMS session
            user_info = get_user_info_dict(request)
            
            # Create Purchase Order
            purchase_order = PurchaseOrder.objects.create(
                po_number=po_number,
                customer_id=customer_id,
                quotation_id=quotation_id if quotation_id else None,
                total_amount=total_amount,
                received_date=received_date,
                delivery_date=delivery_date,
                payment_terms=payment_terms,
                payment_method=payment_method,
                payment_terms_declared=payment_terms_declared,
                special_requirements=special_requirements,
                # Get user info from HRMS session


                

                # Store HRMS user info

                created_by_user_id=user_info['user_id'],

                created_by_username=user_info['username'],

                created_by_email=user_info['email'],

                created_by_full_name=user_info['full_name'],
            )
            
            messages.success(request, f'Purchase Order {po_number} created successfully!')
            return redirect('marketing:po_detail', po_id=purchase_order.id)
            
        except Exception as e:
            messages.error(request, f'Error creating Purchase Order: {str(e)}')
    
    # Get customers and quotations for the form
    customers = Customer.objects.all().order_by('name')
    quotations = Quotation.objects.filter(status='approved').order_by('-created_at')
    
    context = {
        'customers': customers,
        'quotations': quotations,
        'page_title': 'Create Purchase Order',
        'breadcrumb': ['Purchase Orders', 'Create PO'],
    }
    return render(request, 'marketing/po_create.html', context)

@login_required
def po_detail(request, po_id):
    """Purchase Order Detail View"""
    purchase_order = get_object_or_404(PurchaseOrder, id=po_id)
    
    context = {
        'purchase_order': purchase_order,
    }
    return render(request, 'marketing/po_detail.html', context)

# Duplicate expense_list function removed - already defined earlier

@login_required
def triggers(request):
    """Triggers View"""
    context = {}
    return render(request, 'marketing/triggers.html', context)

@login_required
def sequences(request):
    """Sequences View"""
    context = {}
    return render(request, 'marketing/sequences.html', context)

@login_required
def post_event_analysis(request):
    """Post Event Analysis"""
    # Get completed exhibitions
    completed_exhibitions = Exhibition.objects.filter(
        end_date__lt=timezone.now().date()
    ).order_by('-end_date')
    
    # Get exhibition performance data from database
    exhibition_performance = []
    for exhibition in completed_exhibitions:
        # Get leads generated from this exhibition (leads created during exhibition period)
        exhibition_leads = Lead.objects.filter(
            source='event',
            created_at__date__gte=exhibition.start_date,
            created_at__date__lte=exhibition.end_date
        )
        qualified_leads = exhibition_leads.filter(status__in=['qualified', 'proposal', 'negotiation', 'converted']).count()
        total_leads = exhibition_leads.count()
        
        # Calculate conversion rate
        conversion_rate = (qualified_leads / total_leads * 100) if total_leads > 0 else 0
        
        # Get expenses for this exhibition
        exhibition_expenses = Expense.objects.filter(
            expense_type__in=['travel', 'accommodation', 'entertainment', 'office'],
            date__gte=exhibition.start_date,
            date__lte=exhibition.end_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Calculate ROI (simplified - using exhibition budget vs expenses)
        roi = ((exhibition.budget - exhibition_expenses) / exhibition_expenses * 100) if exhibition_expenses > 0 else 0
        
        exhibition_performance.append({
            'exhibition': exhibition.name,
            'event_date': exhibition.start_date.strftime('%Y-%m-%d'),
            'total_visitors': exhibition.visitor_count,
            'qualified_leads': qualified_leads,
            'conversion_rate': round(conversion_rate, 1),
            'total_expenses': float(exhibition_expenses),
            'roi': round(float(roi), 1),
            'visitor_satisfaction': 0,  # Would need a separate model for this
            'objectives_achieved': 'N/A'  # Would need a separate model for this
        })
    
    # Calculate overall statistics
    total_exhibitions = len(exhibition_performance)
    total_visitors = sum(e['total_visitors'] for e in exhibition_performance)
    total_leads = sum(e['qualified_leads'] for e in exhibition_performance)
    avg_conversion = sum(e['conversion_rate'] for e in exhibition_performance) / total_exhibitions if total_exhibitions > 0 else 0
    total_expenses = sum(e['total_expenses'] for e in exhibition_performance)
    avg_roi = sum(e['roi'] for e in exhibition_performance) / total_exhibitions if total_exhibitions > 0 else 0
    
    context = {
        'completed_exhibitions': completed_exhibitions,
        'exhibition_performance': exhibition_performance,
        'total_exhibitions': total_exhibitions,
        'total_visitors': total_visitors,
        'total_leads': total_leads,
        'avg_conversion': round(avg_conversion, 1),
        'total_expenses': total_expenses,
        'avg_roi': round(avg_roi, 1),
    }
    return render(request, 'marketing/post_event_analysis.html', context)

# Phase 8: Advanced Features
@login_required
def email_automation(request):
    """Email Automation Management"""
    if request.method == 'POST':
        template_name = request.POST.get('template_name')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        trigger_type = request.POST.get('trigger_type')
        target_audience = request.POST.get('target_audience')
        schedule_type = request.POST.get('schedule_type')
        
        # Create email template (placeholder for now)
        email_data = {
            'template_name': template_name,
            'subject': subject,
            'content': content,
            'trigger_type': trigger_type,
            'target_audience': target_audience,
            'schedule_type': schedule_type,
            'created_by': request.user.username,
            'created_at': timezone.now().isoformat(),
        }
        
        messages.success(request, f'Email template "{template_name}" created successfully!')
        return redirect('marketing:email_automation')
    
    # Get email templates from database
    email_templates_qs = EmailTemplate.objects.all().order_by('-created_at')
    
    # Convert to list format for template compatibility
    email_templates = []
    for template in email_templates_qs:
        email_templates.append({
            'name': template.name,
            'subject': template.subject,
            'trigger': 'Manual',  # EmailTemplate doesn't have trigger field
            'status': 'Active' if template.is_active else 'Inactive',
            'sent_count': 0,  # Would need EmailLog model to track this
            'open_rate': 0,  # Would need EmailLog model to track this
            'click_rate': 0,  # Would need EmailLog model to track this
            'last_sent': 'N/A'  # Would need EmailLog model to track this
        })
    
    # Calculate statistics
    total_templates = len(email_templates)
    active_templates = len([t for t in email_templates if t['status'] == 'Active'])
    total_sent = sum(t['sent_count'] for t in email_templates)
    avg_open_rate = sum(t['open_rate'] for t in email_templates) / total_templates if total_templates > 0 else 0
    
    context = {
        'email_templates': email_templates,
        'total_templates': total_templates,
        'active_templates': active_templates,
        'total_sent': total_sent,
        'avg_open_rate': round(avg_open_rate, 1),
    }
    return render(request, 'marketing/email_automation.html', context)

@login_required
def sms_notifications(request):
    """SMS Notifications Management"""
    if request.method == 'POST':
        notification_name = request.POST.get('notification_name')
        message_content = request.POST.get('message_content')
        trigger_event = request.POST.get('trigger_event')
        recipients = request.POST.get('recipients')
        schedule = request.POST.get('schedule')
        
        # Create SMS notification (placeholder for now)
        sms_data = {
            'notification_name': notification_name,
            'message_content': message_content,
            'trigger_event': trigger_event,
            'recipients': recipients,
            'schedule': schedule,
            'created_by': request.user.username,
            'created_at': timezone.now().isoformat(),
        }
        
        messages.success(request, f'SMS notification "{notification_name}" created successfully!')
        return redirect('marketing:sms_notifications')
    
    # Get SMS notifications - using EmailTemplate as reference (no SMS model exists yet)
    # In a real implementation, you'd have an SMSNotification model
    sms_notifications = []
    # For now, return empty list or use EmailTemplate as reference
    # email_templates = EmailTemplate.objects.all()
    # You could map email templates to SMS notifications if needed
    
    # Calculate statistics
    total_notifications = len(sms_notifications)
    active_notifications = len([n for n in sms_notifications if n.get('status') == 'Active'])
    total_sent = sum(n.get('sent_count', 0) for n in sms_notifications)
    avg_delivery_rate = sum(n.get('delivery_rate', 0) for n in sms_notifications) / total_notifications if total_notifications > 0 else 0
    
    context = {
        'sms_notifications': sms_notifications,
        'total_notifications': total_notifications,
        'active_notifications': active_notifications,
        'total_sent': total_sent,
        'avg_delivery_rate': round(avg_delivery_rate, 1),
    }
    return render(request, 'marketing/sms_notifications.html', context)

@login_required
def file_upload_management(request):
    """File Upload and Management System"""
    if request.method == 'POST':
        file_title = request.POST.get('file_title')
        file_category = request.POST.get('file_category')
        file_description = request.POST.get('file_description')
        file_tags = request.POST.get('file_tags')
        
        # Handle file upload (placeholder for now)
        uploaded_file = request.FILES.get('file')
        
        if uploaded_file:
            # Process file upload
            file_data = {
                'title': file_title,
                'category': file_category,
                'description': file_description,
                'tags': file_tags,
                'filename': uploaded_file.name,
                'size': uploaded_file.size,
                'uploaded_by': request.user.username,
                'uploaded_at': timezone.now().isoformat(),
            }
            
            messages.success(request, f'File "{file_title}" uploaded successfully!')
        else:
            messages.error(request, 'Please select a file to upload.')
        
        return redirect('marketing:file_upload_management')
    
    # Get uploaded files from models that have FileField
    uploaded_files = []
    
    # Get files from Quotation (GA drawings)
    quotations_with_files = Quotation.objects.exclude(ga_drawing='').select_related('customer', 'created_by')
    for quotation in quotations_with_files:
        if quotation.ga_drawing:
            file_size = quotation.ga_drawing.size / (1024 * 1024)  # Convert to MB
            uploaded_files.append({
                'title': f'GA Drawing - {quotation.quotation_number}',
                'category': 'Drawings',
                'filename': quotation.ga_drawing.name.split('/')[-1],
                'size': f'{file_size:.2f} MB',
                'uploaded_by': quotation.created_by_username or 'Unknown',
                'uploaded_at': quotation.created_at.strftime('%Y-%m-%d'),
                'downloads': 0,  # Would need tracking model
                'tags': f'quotation, {quotation.customer.name}',
                'file_url': quotation.ga_drawing.url
            })
    
    # Get files from GADrawing model
    ga_drawings = GADrawing.objects.exclude(drawing_file='').select_related('urs', 'created_by')
    for drawing in ga_drawings:
        if drawing.drawing_file:
            file_size = drawing.drawing_file.size / (1024 * 1024)  # Convert to MB
            uploaded_files.append({
                'title': drawing.title,
                'category': 'GA Drawings',
                'filename': drawing.drawing_file.name.split('/')[-1],
                'size': f'{file_size:.2f} MB',
                'uploaded_by': drawing.created_by_username or 'Unknown',
                'uploaded_at': drawing.created_at.strftime('%Y-%m-%d'),
                'downloads': 0,
                'tags': f'ga-drawing, {drawing.urs.project_name}',
                'file_url': drawing.drawing_file.url
            })
    
    # Get files from Expense (receipts)
    expenses_with_receipts = Expense.objects.exclude(receipt='').select_related('user')
    for expense in expenses_with_receipts:
        if expense.receipt:
            file_size = expense.receipt.size / (1024 * 1024)  # Convert to MB
            uploaded_files.append({
                'title': f'Receipt - {expense.description[:30]}',
                'category': 'Expense Receipts',
                'filename': expense.receipt.name.split('/')[-1],
                'size': f'{file_size:.2f} MB',
                'uploaded_by': expense.expense_username or 'Unknown',
                'uploaded_at': expense.created_at.strftime('%Y-%m-%d'),
                'downloads': 0,
                'tags': f'expense, {expense.expense_type}',
                'file_url': expense.receipt.url
            })
    
    # Calculate statistics
    total_files = len(uploaded_files)
    total_size = sum(float(f['size'].split()[0]) for f in uploaded_files)
    total_downloads = sum(f['downloads'] for f in uploaded_files)
    
    # Group by category
    categories = {}
    for file in uploaded_files:
        cat = file['category']
        if cat not in categories:
            categories[cat] = 0
        categories[cat] += 1
    
    context = {
        'uploaded_files': uploaded_files,
        'total_files': total_files,
        'total_size': round(total_size, 1),
        'total_downloads': total_downloads,
        'categories': categories,
    }
    return render(request, 'marketing/file_upload_management.html', context)

@login_required
def calendar_integration(request):
    """Calendar Integration and Event Management"""
    if request.method == 'POST':
        event_title = request.POST.get('event_title')
        event_type = request.POST.get('event_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        location = request.POST.get('location')
        description = request.POST.get('description')
        attendees = request.POST.get('attendees')
        
        # Create calendar event (placeholder for now)
        event_data = {
            'title': event_title,
            'type': event_type,
            'start_date': start_date,
            'end_date': end_date,
            'start_time': start_time,
            'end_time': end_time,
            'location': location,
            'description': description,
            'attendees': attendees,
            'created_by': request.user.username,
            'created_at': timezone.now().isoformat(),
        }
        
        messages.success(request, f'Event "{event_title}" created successfully!')
        return redirect('marketing:calendar_integration')
    
    # Get calendar events from Visit and Exhibition models
    calendar_events = []
    today = timezone.now().date()
    
    # Get upcoming visits
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date__gte=today
    ).select_related('customer', 'location', 'assigned_to').order_by('scheduled_date')[:50]
    
    for visit in upcoming_visits:
        location_str = f"{visit.location.city}, {visit.location.state}" if visit.location else "Location TBD"
        calendar_events.append({
            'title': f'Visit - {visit.customer.name}',
            'type': 'Customer Visit',
            'start_date': visit.scheduled_date.date().strftime('%Y-%m-%d'),
            'end_date': visit.scheduled_date.date().strftime('%Y-%m-%d'),
            'start_time': visit.scheduled_date.time().strftime('%H:%M'),
            'end_time': (visit.scheduled_date + timedelta(hours=2)).time().strftime('%H:%M') if visit.scheduled_date else '17:00',
            'location': location_str,
            'attendees': visit.assigned_to_full_name or 'TBD',
            'status': visit.status.title()
        })
    
    # Get upcoming exhibitions
    upcoming_exhibitions = Exhibition.objects.filter(
        start_date__gte=today
    ).order_by('start_date')[:20]
    
    for exhibition in upcoming_exhibitions:
        calendar_events.append({
            'title': f'Exhibition - {exhibition.name}',
            'type': 'Exhibition',
            'start_date': exhibition.start_date.strftime('%Y-%m-%d'),
            'end_date': exhibition.end_date.strftime('%Y-%m-%d'),
            'start_time': '09:00',
            'end_time': '17:00',
            'location': exhibition.venue,
            'attendees': 'Team',
            'status': exhibition.status.title()
        })
    
    # Get visits with follow-up dates
    follow_up_visits = Visit.objects.filter(
        next_follow_up_date__isnull=False,
        next_follow_up_date__date__gte=today
    ).select_related('customer').order_by('next_follow_up_date')[:20]
    
    for visit in follow_up_visits:
        calendar_events.append({
            'title': f'Follow-up - {visit.customer.name}',
            'type': 'Follow-up',
            'start_date': visit.next_follow_up_date.date().strftime('%Y-%m-%d'),
            'end_date': visit.next_follow_up_date.date().strftime('%Y-%m-%d'),
            'start_time': '10:00',
            'end_time': '10:30',
            'location': 'Phone Call / Visit',
            'attendees': visit.assigned_to_full_name or 'TBD',
            'status': 'Pending'
        })
    
    # Sort by start_date
    calendar_events.sort(key=lambda x: x['start_date'])
    
    # Calculate statistics
    total_events = len(calendar_events)
    upcoming_events = len([e for e in calendar_events if e['start_date'] >= today.strftime('%Y-%m-%d')])
    confirmed_events = len([e for e in calendar_events if e['status'] in ['Confirmed', 'Scheduled', 'Completed']])
    
    # Group by event type
    event_types = {}
    for event in calendar_events:
        event_type = event['type']
        if event_type not in event_types:
            event_types[event_type] = 0
        event_types[event_type] += 1
    
    context = {
        'calendar_events': calendar_events,
        'total_events': total_events,
        'upcoming_events': upcoming_events,
        'confirmed_events': confirmed_events,
        'event_types': event_types,
    }
    return render(request, 'marketing/calendar_integration.html', context)

@login_required
def real_time_notifications(request):
    """Real-time Notifications Dashboard"""
    notifications = []
    today = timezone.now().date()
    tomorrow = today + timedelta(days=1)
    
    # Get upcoming visits (tomorrow)
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date=tomorrow,
        status='scheduled'
    ).select_related('customer')[:10]
    
    for visit in upcoming_visits:
        visit_time = visit.scheduled_date.time().strftime('%I:%M %p')
        notifications.append({
            'type': 'visit',
            'title': 'New Customer Visit Scheduled',
            'message': f'Visit scheduled for {visit.customer.name} tomorrow at {visit_time}',
            'timestamp': visit.created_at.strftime('%Y-%m-%d %H:%M'),
            'priority': 'high',
            'read': False
        })
    
    # Get overdue follow-ups
    overdue_followups = Visit.objects.filter(
        next_follow_up_date__date__lt=today,
        status__in=['completed', 'in_progress']
    ).select_related('customer')[:10]
    
    for visit in overdue_followups:
        days_overdue = (today - visit.next_follow_up_date.date()).days
        notifications.append({
            'type': 'follow_up',
            'title': 'Follow-up Overdue',
            'message': f'Follow-up for {visit.customer.name} is {days_overdue} day(s) overdue',
            'timestamp': visit.updated_at.strftime('%Y-%m-%d %H:%M') if hasattr(visit, 'updated_at') else visit.created_at.strftime('%Y-%m-%d %H:%M'),
            'priority': 'medium',
            'read': False
        })
    
    # Get upcoming exhibitions (within 7 days)
    upcoming_exhibitions = Exhibition.objects.filter(
        start_date__lte=today + timedelta(days=7),
        start_date__gte=today,
        status__in=['confirmed', 'planning']
    )[:10]
    
    for exhibition in upcoming_exhibitions:
        days_until = (exhibition.start_date - today).days
        notifications.append({
            'type': 'exhibition',
            'title': 'Exhibition Reminder',
            'message': f'{exhibition.name} starts in {days_until} day(s)',
            'timestamp': exhibition.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(exhibition, 'created_at') else timezone.now().strftime('%Y-%m-%d %H:%M'),
            'priority': 'low',
            'read': False
        })
    
    # Get expenses pending approval
    pending_expenses = Expense.objects.filter(
        status='prepared'
    ).select_related('user')[:10]
    
    for expense in pending_expenses:
        notifications.append({
            'type': 'expense',
            'title': 'Expense Approval Required',
            'message': f'Expense claim of ₹{expense.amount} from {expense.expense_full_name or "User"} requires approval',
            'timestamp': expense.created_at.strftime('%Y-%m-%d %H:%M'),
            'priority': 'medium',
            'read': False
        })
    
    # Sort by timestamp (most recent first)
    notifications.sort(key=lambda x: x['timestamp'], reverse=True)
    
    # Calculate statistics
    total_notifications = len(notifications)
    unread_notifications = len([n for n in notifications if not n['read']])
    high_priority = len([n for n in notifications if n['priority'] == 'high'])
    
    # Group by type
    notification_types = {}
    for notification in notifications:
        notif_type = notification['type']
        if notif_type not in notification_types:
            notification_types[notif_type] = 0
        notification_types[notif_type] += 1
    
    context = {
        'notifications': notifications,
        'total_notifications': total_notifications,
        'unread_notifications': unread_notifications,
        'high_priority': high_priority,
        'notification_types': notification_types,
    }
    return render(request, 'marketing/real_time_notifications.html', context)

@login_required
def audit_trail_system(request):
    """Audit Trail and Activity Logging"""
    # Get audit trail data (placeholder data)
    audit_logs = [
        {
            'user': 'admin',
            'action': 'Created Customer',
            'details': 'New customer "ABC Corp" added to database',
            'timestamp': '2024-01-20 16:45',
            'ip_address': '192.168.1.100',
            'user_agent': 'Chrome/120.0.0.0',
            'status': 'Success'
        },
        {
            'user': 'sales_rep',
            'action': 'Updated Visit',
            'details': 'Visit status changed from "Scheduled" to "Completed"',
            'timestamp': '2024-01-20 15:30',
            'ip_address': '192.168.1.101',
            'user_agent': 'Mobile Safari',
            'status': 'Success'
        },
        {
            'user': 'manager',
            'action': 'Approved Expense',
            'details': 'Expense claim of ₹3,500 approved for John Doe',
            'timestamp': '2024-01-20 14:20',
            'ip_address': '192.168.1.102',
            'user_agent': 'Firefox/121.0',
            'status': 'Success'
        },
        {
            'user': 'designer',
            'action': 'Uploaded File',
            'details': 'File "booth_design_v2.ai" uploaded to system',
            'timestamp': '2024-01-20 13:15',
            'ip_address': '192.168.1.103',
            'user_agent': 'Chrome/120.0.0.0',
            'status': 'Success'
        },
        {
            'user': 'admin',
            'action': 'Failed Login',
            'details': 'Failed login attempt for user "test_user"',
            'timestamp': '2024-01-20 12:00',
            'ip_address': '192.168.1.104',
            'user_agent': 'Unknown',
            'status': 'Failed'
        }
    ]
    
    # Calculate statistics
    total_logs = len(audit_logs)
    successful_actions = len([log for log in audit_logs if log['status'] == 'Success'])
    failed_actions = len([log for log in audit_logs if log['status'] == 'Failed'])
    
    # Group by action type
    action_types = {}
    for log in audit_logs:
        action = log['action']
        if action not in action_types:
            action_types[action] = 0
        action_types[action] += 1
    
    # Group by user
    user_activities = {}
    for log in audit_logs:
        user = log['user']
        if user not in user_activities:
            user_activities[user] = 0
        user_activities[user] += 1
    
    context = {
        'audit_logs': audit_logs,
        'total_logs': total_logs,
        'successful_actions': successful_actions,
        'failed_actions': failed_actions,
        'action_types': action_types,
        'user_activities': user_activities,
    }
    return render(request, 'marketing/audit_trail_system.html', context)

@login_required
def region_management(request):
    """Region Management Interface"""
    if request.method == 'POST':
        action = request.POST.get('action', 'create')

        if action == 'delete':
            region_id = request.POST.get('region_id')
            try:
                region = get_object_or_404(Region, id=region_id)
                region_name = region.name
                
                # Check if region has customers
                customer_count = region.customer_set.count()
                if customer_count > 0:
                    messages.error(request, f'Cannot delete region "{region_name}" because it has {customer_count} customer(s) assigned. Please reassign customers first.')
                    return redirect('marketing:region_management')
                
                # Delete the region
                region.delete()
                messages.success(request, f'Region "{region_name}" has been deleted successfully!')
                
            except Exception as e:
                messages.error(request, f'Error deleting region: {str(e)}')
            
            return redirect('marketing:region_management')

        if action == 'update':
            region_id = request.POST.get('region_id')
            region = get_object_or_404(Region, id=region_id)

            monthly_target_raw = request.POST.get('monthly_target', '').strip()
            quarterly_target_raw = request.POST.get('quarterly_target', '').strip()
            manager_id = request.POST.get('manager_id')

            try:
                monthly_target = Decimal(monthly_target_raw or '0')
            except Exception:
                messages.error(request, 'Please provide a valid monthly target amount.')
                return redirect('marketing:region_management')

            try:
                quarterly_target = Decimal(quarterly_target_raw or '0')
            except Exception:
                messages.error(request, 'Please provide a valid quarterly target amount.')
                return redirect('marketing:region_management')

            manager = None
            if manager_id:
                manager = get_object_or_404(User, id=manager_id)

            region.monthly_target = monthly_target
            region.quarterly_target = quarterly_target
            region.manager = manager
            region.save(update_fields=['monthly_target', 'quarterly_target', 'manager'])

            messages.success(request, f'Targets updated for region "{region.name}".')
            return redirect('marketing:region_management')

        # create action (default)
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        monthly_target_raw = request.POST.get('monthly_target', '').strip()
        quarterly_target_raw = request.POST.get('quarterly_target', '').strip()
        manager_id = request.POST.get('manager_id')

        if not name:
            messages.error(request, 'Region name is required.')
            return redirect('marketing:region_management')

        if Region.objects.filter(name__iexact=name).exists():
            messages.warning(request, f'Region "{name}" already exists.')
            return redirect('marketing:region_management')

        try:
            monthly_target = Decimal(monthly_target_raw or '0')
        except Exception:
            messages.error(request, 'Please provide a valid monthly target amount.')
            return redirect('marketing:region_management')

        try:
            quarterly_target = Decimal(quarterly_target_raw or '0')
        except Exception:
            messages.error(request, 'Please provide a valid quarterly target amount.')
            return redirect('marketing:region_management')

        manager = None
        if manager_id:
            manager = get_object_or_404(User, id=manager_id)

        Region.objects.create(
            name=name,
            description=description,
            monthly_target=monthly_target,
            quarterly_target=quarterly_target,
            manager=manager,
        )

        messages.success(request, f'Region "{name}" created successfully!')
        return redirect('marketing:region_management')

    regions = Region.objects.select_related('manager').annotate(
        customer_count=Count('customer', distinct=True),
        lead_count=Value(0, output_field=IntegerField()),
    ).order_by('name')

    context = {
        'regions': regions,
        'total_regions': regions.count(),
        'total_customers': Customer.objects.count(),
        'total_leads': Lead.objects.count(),
        'managers': User.objects.order_by('first_name', 'last_name', 'username'),
    }
    return render(request, 'marketing/region_management.html', context)





@login_required
def notification_settings(request):
    """Notification Settings Interface"""
    context = {
        'email_settings': {
            'smtp_enabled': True,
            'smtp_host': 'smtp.gmail.com',
            'smtp_port': 587,
            'from_email': 'noreply@company.com'
        },
        'sms_settings': {
            'sms_enabled': False,
            'provider': 'Twilio',
            'api_key': '***hidden***'
        },
        'notification_types': {
            'report_notifications': {
                'daily': True,
                'weekly': True,
                'monthly': True
            },
            'business_notifications': {
                'new_leads': True,
                'quotation_updates': True,
                'order_status': True,
                'expense_approvals': True
            },
            'system_alerts': {
                'backup_completion': True,
                'system_errors': True,
                'maintenance': True
            }
        }
    }
    return render(request, 'marketing/notification_settings.html', context)


@login_required
def export_reports(request):
    """Export Reports Interface"""
    context = {
        'export_history': [
            {
                'id': 1,
                'filename': 'customer_report_2024_01_20.xlsx',
                'type': 'Customer Report',
                'size': '1.2 MB',
                'created_at': '2024-01-20 15:30:00',
                'status': 'Completed'
            },
            {
                'id': 2,
                'filename': 'lead_report_2024_01_19.xlsx',
                'type': 'Lead Report',
                'size': '856 KB',
                'created_at': '2024-01-19 16:45:00',
                'status': 'Completed'
            }
        ],
        'export_templates': [
            {
                'id': 1,
                'name': 'Customer Summary',
                'description': 'Complete customer information with locations',
                'format': 'Excel',
                'last_used': '2024-01-20'
            },
            {
                'id': 2,
                'name': 'Lead Analysis',
                'description': 'Lead performance and conversion metrics',
                'format': 'Excel',
                'last_used': '2024-01-19'
            },
            {
                'id': 3,
                'name': 'Visit Report',
                'description': 'Customer visit details and outcomes',
                'format': 'PDF',
                'last_used': '2024-01-18'
            }
        ]
    }
    return render(request, 'marketing/export_reports.html', context)


# User Profile Management Views
@login_required
def user_profile(request):
    """User Profile View"""
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Get user statistics (using RBAC username)
    user_stats = {
        'total_customers': Customer.objects.filter(created_by_username=username).count(),
        'total_leads': Lead.objects.filter(assigned_to_username=username).count(),
        'total_visits': Visit.objects.filter(assigned_to_username=username).count(),
        'total_expenses': Expense.objects.filter(expense_username=username).count(),
    }
    
    # Get recent activities
    recent_activities = []
    
    # Recent visits
    recent_visits = Visit.objects.filter(assigned_to_username=username).select_related('customer').order_by('-created_at')[:5]
    for visit in recent_visits:
        recent_activities.append({
            'type': 'visit',
            'title': f'Visit to {visit.customer.name}',
            'description': f'{visit.get_visit_type_display()} - {visit.outcome}',
            'time': visit.created_at,
            'icon': 'map-pin',
            'color': 'blue',
            'status': visit.get_status_display()
        })
    
    # Recent expenses
    recent_expenses = Expense.objects.filter(expense_username=username).order_by('-created_at')[:5]
    for expense in recent_expenses:
        recent_activities.append({
            'type': 'expense',
            'title': f'Expense: {expense.description}',
            'description': f'₹{expense.amount} - {expense.get_status_display()}',
            'time': expense.created_at,
            'icon': 'receipt',
            'color': 'green'
        })
    
    # Sort activities by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    
    context = {
        'user': user,
        'user_stats': user_stats,
        'recent_activities': recent_activities[:10],
    }
    return render(request, 'marketing/user_profile.html', context)


@login_required
def edit_profile(request):
    """Edit User Profile"""
    user = request.user
    
    if request.method == 'POST':
        # Update user information
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('marketing:user_profile')
    
    context = {
        'user': user,
    }
    return render(request, 'marketing/edit_profile.html', context)


@login_required
def change_password(request):
    """Change User Password"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate current password
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('marketing:change_password')
        
        # Validate new password
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('marketing:change_password')
        
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('marketing:change_password')
        
        # Update password
        request.user.set_password(new_password)
        request.user.save()
        
        messages.success(request, 'Password changed successfully! Please log in again.')
        return redirect('login')

    return render(request, 'marketing/change_password.html')


# =============================================================================
# MIS REPORTS VIEWS
# =============================================================================

@login_required
def visitor_attendance(request):
    """Visitor Attendance Report"""
    from django.db import models
    from datetime import datetime
    
    # Get visitor attendance data
    today = datetime.now().date()
    this_month = today.replace(day=1)
    
    # Get visits for current month
    monthly_visits = Visit.objects.filter(
        scheduled_date__date__gte=this_month
    ).select_related('customer', 'assigned_to').order_by('-scheduled_date')
    
    # Get daily attendance summary
    daily_attendance = Visit.objects.filter(
        scheduled_date__date__gte=this_month
    ).extra(select={'visit_date': 'DATE(scheduled_date)'}).values('visit_date').annotate(
        total_visits=models.Count('id'),
        unique_visitors=models.Count('customer', distinct=True)
    ).order_by('visit_date')
    
    # Get top visitors
    top_visitors = Visit.objects.filter(
        scheduled_date__date__gte=this_month
    ).values('customer__name', 'customer__contact_person').annotate(
        visit_count=models.Count('id')
    ).order_by('-visit_count')[:10]
    
    context = {
        'page_title': 'Visitor Attendance',
        'breadcrumb': ['MIS Reports', 'Visitor Attendance'],
        'monthly_visits': monthly_visits,
        'daily_attendance': daily_attendance,
        'top_visitors': top_visitors,
        'current_month': this_month.strftime('%B %Y'),
    }
    return render(request, 'marketing/visitor_attendance.html', context)


@login_required
def ongoing_projects(request):
    """Ongoing Projects Report"""
    from django.db import models
    
    # Get ongoing projects data
    ongoing_work_orders = WorkOrder.objects.filter(
        status__in=['in_progress', 'allocated', 'on_hold']
    ).select_related('purchase_order__customer', 'allocated_to').order_by('-created_at')
    
    # Get project status summary
    project_status = WorkOrder.objects.values('status').annotate(
        count=models.Count('id')
    ).order_by('status')
    
    # Get projects by region
    region_projects = WorkOrder.objects.filter(
        status__in=['in_progress', 'allocated', 'on_hold']
    ).values('purchase_order__customer__region__name').annotate(
        count=models.Count('id')
    ).order_by('-count')
    
    context = {
        'page_title': 'Ongoing Projects',
        'breadcrumb': ['MIS Reports', 'Ongoing Projects'],
        'ongoing_work_orders': ongoing_work_orders,
        'project_status': project_status,
        'region_projects': region_projects,
    }
    return render(request, 'marketing/ongoing_projects.html', context)


@login_required
def region_employee_overview(request):
    """Region wise team and target overview UI"""
    # Helper function to get user info from HRMS session
    def get_user_info_dict(request):
        """Get user info from HRMS session"""
        return request.session.get('hrms_user_info', {})
    
    # Determine user role - Updated for HRMS compatibility
    user_role = 'marketing_head'
    
    # Get user info from HRMS session instead of Django groups
    user_info = get_user_info_dict(request)
    user_roles = user_info.get('roles', [])
    
    # Check HRMS roles instead of Django groups
    if any('regional head' in str(role).lower() for role in user_roles):
        user_role = 'regional_head'
    elif any('marketing head' in str(role).lower() for role in user_roles):
        user_role = 'marketing_head'
    elif user_info.get('is_superuser', False):
        user_role = 'marketing_head'

    # Determine accessible regions
    regions_qs = Region.objects.select_related('manager').order_by('name')
    if user_role == 'regional_head':
        # For HRMS users, filter by username instead of user object
        current_username = user_info.get('username', request.session.get('username', ''))
        if current_username:
            regions_qs = regions_qs.filter(manager__username=current_username)
        else:
            # Fallback: show no regions if username not found
            regions_qs = regions_qs.none()

    today = timezone.now().date()
    last_day = calendar.monthrange(today.year, today.month)[1]
    default_deadline = today.replace(day=last_day)

    region_summaries = []

    for region in regions_qs:
        region_target = region.monthly_target or Decimal('0')

        region_achievement = Quotation.objects.filter(
            customer__region=region
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0')

        target_remaining = region_target - region_achievement
        if target_remaining < 0:
            target_remaining = Decimal('0')

        progress_percent = 0.0
        if region_target and region_target > 0:
            progress_percent = float((region_achievement / region_target) * 100)

        team_queryset = Quotation.objects.filter(
            customer__region=region,
            created_by__isnull=False
        ).values(
            'created_by',
            'created_by__first_name',
            'created_by__last_name',
            'created_by__email'
        ).annotate(
            total_sales=Sum('total_amount'),
            deal_count=Count('id'),
            last_activity=Max('updated_at')
        ).order_by('-total_sales')

        team_members = []
        team_size = team_queryset.count()
        individual_target = region_target / team_size if team_size else Decimal('0')

        for member in team_queryset:
            achieved_amount = member['total_sales'] or Decimal('0')
            member_remaining = individual_target - achieved_amount
            if member_remaining < 0:
                member_remaining = Decimal('0')

            member_progress = 0.0
            if individual_target and individual_target > 0:
                member_progress = float((achieved_amount / individual_target) * 100)

            team_members.append({
                'id': member['created_by'],
                'name': (f"{member['created_by__first_name']} {member['created_by__last_name']}".strip()
                         or 'Unnamed User'),
                'email': member['created_by__email'],
                'role': 'Sales Executive',
                'target': individual_target,
                'achieved': achieved_amount,
                'remaining': member_remaining,
                'progress_percent': member_progress,
                'deal_count': member['deal_count'],
                'last_activity': member['last_activity'],
                'deadline': default_deadline,
            })

        region_summaries.append({
            'id': region.id,
            'name': region.name,
            'description': region.description,
            'head': region.manager,
            'employee_count': team_size,
            'target_total': region_target,
            'achieved_amount': region_achievement,
            'remaining_amount': target_remaining,
            'progress_percent': progress_percent,
            'deadline': default_deadline,
            'team_members': team_members,
        })

    total_regions = len(region_summaries)
    total_employees = sum(region['employee_count'] for region in region_summaries)
    total_target = sum(region['target_total'] for region in region_summaries) or Decimal('0')
    total_achieved = sum(region['achieved_amount'] for region in region_summaries) or Decimal('0')
    total_remaining = sum(region['remaining_amount'] for region in region_summaries) or Decimal('0')

    overall_progress = 0.0
    if total_target and total_target > 0:
        overall_progress = float((total_achieved / total_target) * 100)

    context = {
        'page_title': 'Region Teams & Targets',
        'breadcrumb': ['Customers', 'Region Teams'],
        'user_role': user_role,
        'regions': region_summaries,
        'global_stats': {
            'total_regions': total_regions,
            'total_employees': total_employees,
            'total_target': total_target,
            'total_achieved': total_achieved,
            'total_remaining': total_remaining,
            'progress_percent': overall_progress,
            'deadline': default_deadline,
        },
    }

    return render(request, 'marketing/region_employee_overview.html', context)


@login_required
def region_targets(request):
    """Region-wise Targets with Machine-wise Sales"""
    from django.db import models
    
    # Get region-wise targets and sales data
    regions = Region.objects.all().order_by('name')
    
    # Get machine-wise sales data by region
    machine_sales = []
    for region in regions:
        region_data = {
            'region': region,
            'machines': []
        }
        
        # Get sales data for this region from quotations
        region_quotations = Quotation.objects.filter(
            customer__region=region
        ).aggregate(
            total_amount=models.Sum('total_amount'),
            count=models.Count('id')
        )
        
        region_data['total_sales'] = region_quotations['total_amount'] or Decimal('0')
        region_data['quotation_count'] = region_quotations['count'] or 0
        
        # Calculate region performance percentage
        region_target = region.monthly_target or Decimal('0')
        region_performance = (region_data['total_sales'] / region_target * 100) if region_target > 0 else 0
        region_data['performance_percentage'] = round(region_performance, 1)
        
        # Get project breakdown from URS for this region
        region_urs = URS.objects.filter(customer__region=region).values('project_name').annotate(
            count=models.Count('id')
        ).order_by('-count')[:10]
        
        machines = []
        for urs_item in region_urs:
            project_name = urs_item['project_name'] or 'Other'
            count = urs_item['count']
            
            # Get quotations for customers with this project type
            project_quotations = Quotation.objects.filter(
                customer__region=region,
                customer__urs__project_name=project_name
            ).aggregate(
                total_sales=models.Sum('total_amount'),
                quotation_count=models.Count('id')
            )
            
            sales_amount = float(project_quotations['total_sales'] or 0)
            quotation_count = project_quotations['quotation_count'] or 0
            avg_price = sales_amount / quotation_count if quotation_count > 0 else 0
            
            machines.append({
                'name': project_name,
                'sales': sales_amount,
                'count': quotation_count,
                'avg_price': avg_price,
                'target': 0,  # Would need machine-specific targets
                'performance': 0  # Would need machine-specific targets
            })
        
        # If no URS data, use general quotation data as fallback
        if not machines and region_data['quotation_count'] > 0:
            machines.append({
                'name': 'General Quotations',
                'sales': float(region_data['total_sales']),
                'count': region_data['quotation_count'],
                'avg_price': float(region_data['total_sales']) / region_data['quotation_count'],
                'target': 0,
                'performance': 0
            })
        
        region_data['machines'] = machines
        machine_sales.append(region_data)
    
    # Get overall targets vs achievements from region monthly targets
    total_target = sum(region.monthly_target or Decimal('0') for region in regions)
    total_achieved = Quotation.objects.aggregate(
        total=models.Sum('total_amount')
    )['total'] or Decimal('0')
    
    achievement_percentage = (total_achieved / total_target * 100) if total_target > 0 else 0
    
    context = {
        'page_title': 'Region-wise Targets',
        'breadcrumb': ['MIS Reports', 'Region-wise Targets'],
        'machine_sales': machine_sales,
        'total_target': total_target,
        'total_achieved': total_achieved,
        'achievement_percentage': achievement_percentage,
    }
    return render(request, 'marketing/region_targets.html', context)


# Payment Follow-up Views
@login_required
def payment_followup_list(request):
    """List all payment follow-ups"""
    followups = PaymentFollowUp.objects.select_related('purchase_order__customer', 'created_by').all()
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        followups = followups.filter(
            Q(purchase_order__po_number__icontains=search) |
            Q(purchase_order__customer__name__icontains=search) |
            Q(notes__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        followups = followups.filter(status=status_filter)
    
    # Filter by payment method
    method_filter = request.GET.get('method', '')
    if method_filter:
        followups = followups.filter(payment_method=method_filter)
    
    # Get statistics
    stats = {
        'total_followups': PaymentFollowUp.objects.count(),
        'pending_followups': PaymentFollowUp.objects.filter(status='pending').count(),
        'overdue_followups': PaymentFollowUp.objects.filter(
            follow_up_date__lt=timezone.now().date(),
            status='pending'
        ).count(),
        'completed_followups': PaymentFollowUp.objects.filter(status='completed').count(),
    }
    
    # Pagination
    paginator = Paginator(followups, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search': search,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'page_title': 'Payment Follow-ups',
        'breadcrumb': ['Purchase Orders', 'Payment Follow-ups'],
    }
    return render(request, 'marketing/payment_followup_list.html', context)


@login_required
def payment_followup_create(request, po_id):
    """Create payment follow-up for a purchase order"""
    purchase_order = get_object_or_404(PurchaseOrder, id=po_id)
    
    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')
        payment_terms_declared = request.POST.get('payment_terms_declared')
        follow_up_date = request.POST.get('follow_up_date')
        notes = request.POST.get('notes', '')
        
        # Create payment follow-up
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        followup = PaymentFollowUp.objects.create(
            purchase_order=purchase_order,
            payment_method=payment_method,
            payment_terms_declared=payment_terms_declared,
            follow_up_date=follow_up_date,
            notes=notes,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
        )
        
        # Update purchase order with payment method if not set
        if not purchase_order.payment_method:
            purchase_order.payment_method = payment_method
            purchase_order.payment_terms_declared = payment_terms_declared
            purchase_order.save()
        
        messages.success(request, f'Payment follow-up created successfully for PO {purchase_order.po_number}!')
        return redirect('marketing:payment_followup_list')
    
    context = {
        'purchase_order': purchase_order,
        'page_title': f'Create Payment Follow-up - {purchase_order.po_number}',
        'breadcrumb': ['Purchase Orders', 'Payment Follow-up'],
    }
    return render(request, 'marketing/payment_followup_create.html', context)


@login_required
def payment_followup_update(request, followup_id):
    """Update payment follow-up status and details"""
    followup = get_object_or_404(PaymentFollowUp, id=followup_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        followup.status = status
        followup.notes = notes
        followup.save()
        
        messages.success(request, 'Payment follow-up updated successfully!')
        return redirect('marketing:payment_followup_list')
    
    context = {
        'followup': followup,
        'page_title': f'Update Payment Follow-up - {followup.purchase_order.po_number}',
        'breadcrumb': ['Purchase Orders', 'Payment Follow-up', 'Update'],
    }
    return render(request, 'marketing/payment_followup_update.html', context)


@login_required
def payment_followup_dashboard(request):
    """Payment follow-up dashboard with overview"""
    today = timezone.now().date()
    
    # Get follow-ups by status
    pending_followups = PaymentFollowUp.objects.filter(status='pending').select_related('purchase_order__customer')
    overdue_followups = PaymentFollowUp.objects.filter(
        follow_up_date__lt=today,
        status='pending'
    ).select_related('purchase_order__customer')
    
    # Get upcoming follow-ups (next 7 days)
    upcoming_followups = PaymentFollowUp.objects.filter(
        follow_up_date__gte=today,
        follow_up_date__lte=today + timedelta(days=7),
        status='pending'
    ).select_related('purchase_order__customer')
    
    # Get statistics
    stats = {
        'total_followups': PaymentFollowUp.objects.count(),
        'pending_followups': pending_followups.count(),
        'overdue_followups': overdue_followups.count(),
        'upcoming_followups': upcoming_followups.count(),
        'completed_followups': PaymentFollowUp.objects.filter(status='completed').count(),
    }
    
    # Get payment methods distribution
    payment_methods = PaymentFollowUp.objects.values('payment_method').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'pending_followups': pending_followups[:10],
        'overdue_followups': overdue_followups[:10],
        'upcoming_followups': upcoming_followups[:10],
        'stats': stats,
        'payment_methods': payment_methods,
        'page_title': 'Payment Follow-up Dashboard',
        'breadcrumb': ['Purchase Orders', 'Payment Follow-up Dashboard'],
    }
    return render(request, 'marketing/payment_followup_dashboard.html', context)


# Annual Exhibition Budget Views
@login_required
def annual_budget_list(request):
    """List all annual exhibition budgets"""
    budgets = AnnualExhibitionBudget.objects.all().order_by('-year')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        budgets = budgets.filter(year__icontains=search)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        budgets = budgets.filter(status=status_filter)
    
    # Get statistics
    stats = {
        'total_budgets': AnnualExhibitionBudget.objects.count(),
        'approved_budgets': AnnualExhibitionBudget.objects.filter(status='approved').count(),
        'pending_budgets': AnnualExhibitionBudget.objects.filter(status__in=['draft', 'submitted', 'under_review']).count(),
        'current_year_budget': AnnualExhibitionBudget.objects.filter(year=timezone.now().year).first(),
    }
    
    # Pagination
    paginator = Paginator(budgets, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search': search,
        'status_filter': status_filter,
        'page_title': 'Annual Exhibition Budgets',
        'breadcrumb': ['Exhibitions', 'Annual Budgets'],
    }
    return render(request, 'marketing/annual_budget_list.html', context)


@login_required
def annual_budget_create(request):
    """Create new annual exhibition budget"""
    if request.method == 'POST':
        year = request.POST.get('year')
        total_budget = request.POST.get('total_budget')
        notes = request.POST.get('notes', '')
        
        # Create annual budget
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)

        budget = AnnualExhibitionBudget.objects.create(
            year=year,
            total_budget=total_budget,
            notes=notes,
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],

            created_by_email=user_info['email'],

            created_by_full_name=user_info['full_name'],
        )
        
        # Create budget allocations for each category
        categories = BudgetCategory.objects.filter(is_active=True)
        for category in categories:
            allocation_amount = request.POST.get(f'allocation_{category.id}', 0)
            if allocation_amount and float(allocation_amount) > 0:
                BudgetAllocation.objects.create(
                    annual_budget=budget,
                    category=category,
                    allocated_amount=allocation_amount
                )
        
        # Update allocated budget
        budget.allocated_budget = sum(
            allocation.allocated_amount for allocation in budget.allocations.all()
        )
        budget.save()
        
        messages.success(request, f'Annual budget for {year} created successfully!')
        return redirect('marketing:annual_budget_list')
    
    categories = BudgetCategory.objects.filter(is_active=True)
    context = {
        'categories': categories,
        'page_title': 'Create Annual Budget',
        'breadcrumb': ['Exhibitions', 'Annual Budgets', 'Create'],
    }
    return render(request, 'marketing/annual_budget_create.html', context)


@login_required
def annual_budget_detail(request, budget_id):
    """View annual budget details with allocations and performance"""
    budget = get_object_or_404(AnnualExhibitionBudget, id=budget_id)
    allocations = budget.allocations.all().order_by('category__name')
    approvals = budget.approvals.all().order_by('-created_at')
    
    # Get exhibitions for this budget year
    exhibitions = Exhibition.objects.filter(
        annual_budget=budget
    ).order_by('-start_date')
    
    # Calculate performance metrics
    total_exhibitions = exhibitions.count()
    completed_exhibitions = exhibitions.filter(status='completed').count()
    total_visitors = sum(ex.visitor_count for ex in exhibitions)
    
    # Get monthly spending data
    monthly_data = []
    for month in range(1, 13):
        month_exhibitions = exhibitions.filter(start_date__month=month)
        month_spending = sum(ex.total_expense for ex in month_exhibitions)
        monthly_data.append({
            'month': month,
            'spending': month_spending,
            'exhibitions': month_exhibitions.count()
        })
    
    context = {
        'budget': budget,
        'allocations': allocations,
        'approvals': approvals,
        'exhibitions': exhibitions,
        'total_exhibitions': total_exhibitions,
        'completed_exhibitions': completed_exhibitions,
        'total_visitors': total_visitors,
        'monthly_data': monthly_data,
        'page_title': f'Annual Budget {budget.year}',
        'breadcrumb': ['Exhibitions', 'Annual Budgets', str(budget.year)],
    }
    return render(request, 'marketing/annual_budget_detail.html', context)


@login_required
def annual_budget_approve(request, budget_id):
    """Approve annual budget"""
    budget = get_object_or_404(AnnualExhibitionBudget, id=budget_id)
    
    if request.method == 'POST':
        approval_level = request.POST.get('approval_level')
        status = request.POST.get('status')
        comments = request.POST.get('comments', '')
        
        # Create approval record
        approval = BudgetApproval.objects.create(
            annual_budget=budget,
            approval_level=approval_level,
            status=status,
            comments=comments,
            # Get user info from HRMS session


            

            # Store HRMS user info

            approved_by_user_id=user_info['user_id'],

            approved_by_username=user_info['username'],

            approved_by_email=user_info['email'],

            approved_by_full_name=user_info['full_name'] if status == 'approved' else '',
            approved_at=timezone.now() if status == 'approved' else None
        )
        
        # Update budget status
        if status == 'approved':
            budget.status = 'approved'
            budget.approved_by = request.user
            budget.approved_at = timezone.now()
            budget.save()
        
        messages.success(request, f'Budget approval {status} successfully!')
        return redirect('marketing:annual_budget_detail', budget_id=budget.id)
    
    context = {
        'budget': budget,
        'page_title': f'Approve Budget {budget.year}',
        'breadcrumb': ['Exhibitions', 'Annual Budgets', 'Approve'],
    }
    return render(request, 'marketing/annual_budget_approve.html', context)


@login_required
def budget_dashboard(request):
    """Annual budget dashboard with overview and analytics"""
    current_year = timezone.now().year
    
    # Get current year budget
    current_budget = AnnualExhibitionBudget.objects.filter(year=current_year).first()
    
    # Get budget statistics
    stats = {
        'total_budgets': AnnualExhibitionBudget.objects.count(),
        'approved_budgets': AnnualExhibitionBudget.objects.filter(status='approved').count(),
        'pending_budgets': AnnualExhibitionBudget.objects.filter(status__in=['draft', 'submitted', 'under_review']).count(),
        'current_year_budget': current_budget,
    }
    
    # Get recent budgets
    recent_budgets = AnnualExhibitionBudget.objects.all().order_by('-year')[:5]
    
    # Get category-wise spending for current year
    category_spending = []
    if current_budget:
        for allocation in current_budget.allocations.all():
            category_spending.append({
                'category': allocation.category.name,
                'allocated': allocation.allocated_amount,
                'spent': allocation.spent_amount,
                'remaining': allocation.remaining_amount,
                'utilization': allocation.utilization_percentage
            })
    
    # Get monthly spending trend for current year
    monthly_spending = []
    if current_budget:
        exhibitions = Exhibition.objects.filter(annual_budget=current_budget)
        for month in range(1, 13):
            month_exhibitions = exhibitions.filter(start_date__month=month)
            month_spending = sum(ex.total_expense for ex in month_exhibitions)
            monthly_spending.append({
                'month': month,
                'spending': month_spending
            })
    
    context = {
        'stats': stats,
        'recent_budgets': recent_budgets,
        'category_spending': category_spending,
        'monthly_spending': monthly_spending,
        'current_year': current_year,
        'page_title': 'Budget Dashboard',
        'breadcrumb': ['Exhibitions', 'Budget Dashboard'],
    }
    return render(request, 'marketing/budget_dashboard.html', context)


@login_required
def budget_category_manage(request):
    """Manage budget categories"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name')
            category_type = request.POST.get('category_type')
            description = request.POST.get('description', '')
            is_active = request.POST.get('is_active') == 'on'
            
            BudgetCategory.objects.create(
                name=name,
                category_type=category_type,
                description=description,
                is_active=is_active
            )
            messages.success(request, f'Budget category "{name}" created successfully!')
            
        elif action == 'update':
            category_id = request.POST.get('category_id')
            name = request.POST.get('name')
            category_type = request.POST.get('category_type')
            description = request.POST.get('description', '')
            is_active = request.POST.get('is_active') == 'on'
            
            category = BudgetCategory.objects.get(id=category_id)
            category.name = name
            category.category_type = category_type
            category.description = description
            category.is_active = is_active
            category.save()
            messages.success(request, f'Budget category "{name}" updated successfully!')
            
        elif action == 'toggle_status':
            category_id = request.POST.get('category_id')
            is_active = request.POST.get('is_active') == 'true'
            
            category = BudgetCategory.objects.get(id=category_id)
            category.is_active = is_active
            category.save()
            status = 'activated' if is_active else 'deactivated'
            messages.success(request, f'Budget category "{category.name}" {status} successfully!')
        
        return redirect('marketing:budget_category_manage')
    
    categories = BudgetCategory.objects.all().order_by('category_type', 'name')
    
    context = {
        'categories': categories,
        'page_title': 'Manage Budget Categories',
        'breadcrumb': ['Exhibitions', 'Budget Categories'],
    }
    return render(request, 'marketing/budget_category_manage.html', context)

@login_required
def expense_export(request):
    """Export Expense Data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    import io
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Management"
    
    # Define headers
    headers = [
        'Employee', 'Date', 'Type', 'Amount', 'Status', 'Description', 'Receipt', 'Created At'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data (replace with actual expense data when model is available)
    sample_expenses = [
        ['John Doe', '2024-01-20', 'Travel', '₹2,500', 'Approved', 'Flight tickets to Mumbai', 'receipt_001.pdf', '2024-01-20 10:30'],
        ['Jane Smith', '2024-01-22', 'Meals', '₹1,200', 'Pending', 'Client lunch meeting', 'receipt_002.pdf', '2024-01-22 14:15'],
        ['Mike Brown', '2024-01-25', 'Accommodation', '₹3,500', 'Approved', 'Hotel stay in Delhi', 'receipt_003.pdf', '2024-01-25 09:45'],
        ['Alice Johnson', '2024-01-28', 'Transport', '₹800', 'Rejected', 'Taxi fare', 'receipt_004.pdf', '2024-01-28 16:20'],
        ['David Wilson', '2024-01-30', 'Entertainment', '₹2,000', 'Pending', 'Client dinner', 'receipt_005.pdf', '2024-01-30 19:30'],
    ]
    
    # Write sample data
    for row, expense_data in enumerate(sample_expenses, 2):
        for col, value in enumerate(expense_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with current date
    from datetime import datetime
    current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Expense_Management_Export_{current_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ==================== INQUIRY LOG VIEWS ====================

@login_required
def inquiry_log_list(request):
    """List all inquiry logs with filtering and pagination"""
    inquiries = InquiryLog.objects.all().order_by('-enquiry_date', '-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        inquiries = inquiries.filter(
            Q(company_name__icontains=search_query) |
            Q(enquiry_number__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    # Filter by month
    month_filter = request.GET.get('month', '')
    if month_filter:
        inquiries = inquiries.filter(month=month_filter)
    
    # Filter by quote status
    quote_send_filter = request.GET.get('quote_send', '')
    if quote_send_filter:
        inquiries = inquiries.filter(quote_send=quote_send_filter)
    
    # Filter by offer category
    category_filter = request.GET.get('offer_category', '')
    if category_filter:
        inquiries = inquiries.filter(offer_category=category_filter)
    
    # Pagination
    paginator = Paginator(inquiries, 20)  # Show 20 inquiries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    # Get user info from HRMS session
# Removed - user_info not needed here

    total_inquiries = InquiryLog.objects.count()
    quotes_sent = InquiryLog.objects.filter(quote_send='yes').count()
    # Get user info from HRMS session
# Removed - user_info not needed here

    pending_followups = InquiryLog.objects.filter(follow_up='').count()
    total_value = InquiryLog.objects.aggregate(
        total=Sum('quote_price')
    )['total'] or 0
    
    # Get unique months for filter dropdown
    months = InquiryLog.objects.values_list('month', flat=True).distinct().order_by('month')
    
    # Get offer categories for filter dropdown
    offer_categories = InquiryLog.OFFER_CATEGORIES
    
    context = {
        'inquiries': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_inquiries': total_inquiries,
        'quotes_sent': quotes_sent,
        'pending_followups': pending_followups,
        'total_value': total_value,
        'months': months,
        'offer_categories': offer_categories,
    }
    
    return render(request, 'marketing/inquiry_log_list.html', context)


@login_required
def inquiry_log_detail(request, pk):
    """View detailed information of a specific inquiry log"""
    inquiry = get_object_or_404(InquiryLog, pk=pk)
    
    context = {
        'inquiry': inquiry,
    }
    
    return render(request, 'marketing/inquiry_log_detail.html', context)


@login_required
def inquiry_log_create(request):
    """Create a new inquiry log entry"""
    if request.method == 'POST':
        # Create form data from POST request
        form_data = request.POST.copy()
        form_data['created_by'] = request.user.id
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        # Create InquiryLog instance
        inquiry = InquiryLog(
            month=form_data.get('month'),
            enquiry_date=form_data.get('enquiry_date'),
            enquiry_number=form_data.get('enquiry_number') or None,
            location=form_data.get('location'),
            enquiry_mail=form_data.get('enquiry_mail'),
            enquiry_through=form_data.get('enquiry_through'),
            quote_number=form_data.get('quote_number') or None,
            quote_date=form_data.get('quote_date') or None,
            offer_category=form_data.get('offer_category'),
            company_name=form_data.get('company_name'),
            company_address=form_data.get('company_address'),
            contact_person=form_data.get('contact_person'),
            contact_number=form_data.get('contact_number'),
            email_id=form_data.get('email_id'),
            requirement_details=form_data.get('requirement_details'),
            quote_send=form_data.get('quote_send'),
            quote_price=form_data.get('quote_price') or None,
            discounted_price=form_data.get('discounted_price') or None,
            follow_up_status=form_data.get('follow_up_status') or None,
            follow_up=form_data.get('follow_up') or '',
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
        )
        
        try:
            inquiry.save()
            messages.success(request, f'Inquiry log {inquiry.enquiry_number} created successfully!')
    # Get user info from HRMS session
# Removed - user_info not needed here

            return redirect('marketing:inquiry_log_detail', pk=inquiry.pk)
        except Exception as e:
            messages.error(request, f'Error creating inquiry log: {str(e)}')
    
    # For GET request, show empty form
    context = {
        'form': None,  # We'll handle form rendering in template
    }
    
    return render(request, 'marketing/inquiry_log_form.html', context)


@login_required
def inquiry_log_edit(request, pk):
    """Edit an existing inquiry log entry"""
    inquiry = get_object_or_404(InquiryLog, pk=pk)
    
    if request.method == 'POST':
        # Update inquiry with form data
        inquiry.month = request.POST.get('month')
        inquiry.enquiry_date = request.POST.get('enquiry_date')
        inquiry.enquiry_number = request.POST.get('enquiry_number') or None
        inquiry.location = request.POST.get('location')
        inquiry.enquiry_mail = request.POST.get('enquiry_mail')
        inquiry.enquiry_through = request.POST.get('enquiry_through')
        inquiry.quote_number = request.POST.get('quote_number') or None
        inquiry.quote_date = request.POST.get('quote_date') or None
        inquiry.offer_category = request.POST.get('offer_category')
        inquiry.company_name = request.POST.get('company_name')
        inquiry.company_address = request.POST.get('company_address')
        inquiry.contact_person = request.POST.get('contact_person')
        inquiry.contact_number = request.POST.get('contact_number')
        inquiry.email_id = request.POST.get('email_id')
        inquiry.requirement_details = request.POST.get('requirement_details')
        inquiry.quote_send = request.POST.get('quote_send')
        inquiry.quote_price = request.POST.get('quote_price') or None
        inquiry.discounted_price = request.POST.get('discounted_price') or None
        inquiry.follow_up_status = request.POST.get('follow_up_status') or None
        inquiry.follow_up = request.POST.get('follow_up') or ''
        
    # Get user info from HRMS session
# Removed - user_info not needed here

        try:
            inquiry.save()
            messages.success(request, f'Inquiry log {inquiry.enquiry_number} updated successfully!')
            return redirect('marketing:inquiry_log_detail', pk=inquiry.pk)
        except Exception as e:
            messages.error(request, f'Error updating inquiry log: {str(e)}')
    
    context = {
        'object': inquiry,
        'form': None,  # We'll handle form rendering in template
    }
    
    return render(request, 'marketing/inquiry_log_form.html', context)


@login_required
def inquiry_log_delete(request, pk):
    # Get user info from HRMS session
# Removed - user_info not needed here


    """Delete an inquiry log entry"""
    inquiry = get_object_or_404(InquiryLog, pk=pk)
    
    if request.method == 'POST':
        enquiry_number = inquiry.enquiry_number
        inquiry.delete()
        messages.success(request, f'Inquiry log {enquiry_number} deleted successfully!')
    # Get user info from HRMS session
# Removed - user_info not needed here

        return redirect('marketing:inquiry_log_list')
    
    context = {
        'inquiry': inquiry,
    }
    
    return render(request, 'marketing/inquiry_log_confirm_delete.html', context)


# MIS System Views
@login_required
def mis_dashboard(request):
    """MIS Dashboard - Main overview of all sheets"""
    # Get statistics for each sheet
    follow_up_count = FollowUpStatus.objects.count()
    project_today_count = ProjectToday.objects.count()
    order_expected_count = OrderExpectedNextMonth.objects.count()
    purchase_order_count = MISPurchaseOrder.objects.count()
    new_data_count = NewData.objects.count()
    new_data_details_count = NewDataDetails.objects.count()
    od_plan_count = ODPlan.objects.count()
    inquiry_log_count = InquiryLog.objects.count()
    
    # Recent activities
    recent_follow_ups = FollowUpStatus.objects.order_by('-created_at')[:5]
    recent_projects = ProjectToday.objects.order_by('-created_at')[:5]
    recent_orders = OrderExpectedNextMonth.objects.order_by('-created_at')[:5]
    
    # Build saved sheets list from various MIS models
    saved_sheets = []
    
    # Add recent inquiry logs
    recent_inquiries = InquiryLog.objects.order_by('-created_at')[:5]
    for inquiry in recent_inquiries:
        saved_sheets.append({
            'type': 'Inquiry Log',
            'title': f'{inquiry.company_name} - {inquiry.enquiry_number}',
            'status': 'saved',
            'modified_at': inquiry.updated_at,
            'edit_url': reverse('marketing:inquiry_log_detail', kwargs={'pk': inquiry.pk}),
        })
    
    # Add recent follow-up statuses
    for follow_up in recent_follow_ups:
        saved_sheets.append({
            'type': 'Follow-Up Status',
            'title': f'{follow_up.company_group} Follow-up',
            'status': 'saved',
            'modified_at': follow_up.updated_at,
            'edit_url': reverse('marketing:follow_up_status_detail', kwargs={'pk': follow_up.pk}),
        })
    
    # Sort by modified_at (most recent first)
    saved_sheets.sort(key=lambda x: x['modified_at'], reverse=True)
    saved_sheets = saved_sheets[:10]  # Limit to 10 most recent
    
    context = {
        'follow_up_count': follow_up_count,
        'project_today_count': project_today_count,
        'order_expected_count': order_expected_count,
        'purchase_order_count': purchase_order_count,
        'new_data_count': new_data_count,
        'new_data_details_count': new_data_details_count,
        'od_plan_count': od_plan_count,
        'inquiry_log_count': inquiry_log_count,
        'recent_follow_ups': recent_follow_ups,
        'recent_projects': recent_projects,
        'recent_orders': recent_orders,
        'saved_sheets': saved_sheets,
    }
    
    return render(request, 'marketing/mis_dashboard.html', context)


@login_required
def follow_up_status_list(request):
    """List all follow-up status entries"""
    follow_ups = FollowUpStatus.objects.all().order_by('-follow_up_date', '-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        follow_ups = follow_ups.filter(
            Q(company_group__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(quote_no__icontains=search_query)
        )

    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        follow_ups = follow_ups.filter(follow_up_status=status_filter)
    
    # Pagination
    paginator = Paginator(follow_ups, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_follow_ups = FollowUpStatus.objects.count()
    pending_follow_ups = FollowUpStatus.objects.filter(follow_up_status__in=['qtn_submitted', 'qtn_followup']).count()
    completed_follow_ups = FollowUpStatus.objects.filter(follow_up_status__in=['order_finalization', 'po_release']).count()
    
    context = {
        'follow_ups': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_follow_ups': total_follow_ups,
        'pending_follow_ups': pending_follow_ups,
        'completed_follow_ups': completed_follow_ups,
        'status_choices': FollowUpStatus.FOLLOW_UP_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/follow_up_status_list.html', context)


@login_required
def follow_up_status_create(request):
    """Create a new follow-up status entry"""
    if request.method == 'POST':
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        follow_up = FollowUpStatus(
            month=request.POST.get('month'),
            date=request.POST.get('date'),
            quote_no=request.POST.get('quote_no') or None,
            responsible_person=request.POST.get('responsible_person'),
            company_group=request.POST.get('company_group'),
            address=request.POST.get('address'),
            contact_person=request.POST.get('contact_person'),
            contact_no=request.POST.get('contact_no'),
            mail_id=request.POST.get('mail_id'),
            requirements=request.POST.get('requirements'),
            follow_up_date=request.POST.get('follow_up_date'),
            follow_up_status=request.POST.get('follow_up_status'),
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
            # Legacy field set to None
            created_by=None,
        )
        
        try:
            follow_up.save()
            messages.success(request, f'Follow-up status entry created successfully!')
            return redirect('marketing:follow_up_status_detail', pk=follow_up.pk)
        except Exception as e:
            messages.error(request, f'Error creating follow-up status: {str(e)}')
    
    context = {
        'status_choices': FollowUpStatus.FOLLOW_UP_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/follow_up_status_form.html', context)


@login_required
def follow_up_status_detail(request, pk):
    """View detailed information of a specific follow-up status entry"""
    follow_up = get_object_or_404(FollowUpStatus, pk=pk)
    
    context = {
        'follow_up': follow_up,
    }
# Removed - user_info not needed here

    
    return render(request, 'marketing/follow_up_status_detail.html', context)


@login_required
def follow_up_status_edit(request, pk):
    """Edit an existing follow-up status entry"""
    follow_up = get_object_or_404(FollowUpStatus, pk=pk)
    
    if request.method == 'POST':
        follow_up.month = request.POST.get('month')
        follow_up.date = request.POST.get('date')
        follow_up.quote_no = request.POST.get('quote_no') or None
        follow_up.responsible_person = request.POST.get('responsible_person')
        follow_up.company_group = request.POST.get('company_group')
        follow_up.address = request.POST.get('address')
        follow_up.contact_person = request.POST.get('contact_person')
        follow_up.contact_no = request.POST.get('contact_no')
        follow_up.mail_id = request.POST.get('mail_id')
        follow_up.requirements = request.POST.get('requirements')
        follow_up.follow_up_date = request.POST.get('follow_up_date')
        follow_up.follow_up_status = request.POST.get('follow_up_status')
        
        try:
            follow_up.save()
            messages.success(request, f'Follow-up status entry updated successfully!')
            return redirect('marketing:follow_up_status_detail', pk=follow_up.pk)
        except Exception as e:
            messages.error(request, f'Error updating follow-up status: {str(e)}')
    
    context = {
        'follow_up': follow_up,
        'status_choices': FollowUpStatus.FOLLOW_UP_STATUS_CHOICES,
    }
    # Get user info from HRMS session
# Removed - user_info not needed here

    
    # Get user info from HRMS session
# Removed - user_info not needed here

    return render(request, 'marketing/follow_up_status_form.html', context)


@login_required
def follow_up_status_delete(request, pk):
    """Delete a follow-up status entry"""
    follow_up = get_object_or_404(FollowUpStatus, pk=pk)
    
    if request.method == 'POST':
        company_name = follow_up.company_group
        follow_up.delete()
        messages.success(request, f'Follow-up status entry for {company_name} deleted successfully!')
        return redirect('marketing:follow_up_status_list')
    
    context = {
        'follow_up': follow_up,
    }
    
    return render(request, 'marketing/follow_up_status_confirm_delete.html', context)


@login_required
def mis_sheets(request):
    """MIS Sheets - Tabbed interface for all sheets"""
    if request.method == 'POST':
        form_type = request.POST.get('form_type') or request.POST.get('data-form-type')
        
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        try:
            if form_type == 'project-today':
                # Create ProjectToday instance
                project = ProjectToday(
                    location=request.POST.get('location'),
                    district=request.POST.get('district'),
                    state=request.POST.get('state'),
                    product1=request.POST.get('product1'),
                    promoter_name=request.POST.get('promoter_name'),
                    promoter_office_add=request.POST.get('promoter_office_add'),
                    promoter_contact_person_name=request.POST.get('promoter_contact_person_name'),
                    promoter_contact_person_designation=request.POST.get('promoter_contact_person_designation'),
                    promoter_contact_person_direct_contact=request.POST.get('promoter_contact_person_direct_contact'),
                    promoter_contact_person_email=request.POST.get('promoter_contact_person_email'),
                    architect_name=request.POST.get('architect_name') or '',
                    consultant_name=request.POST.get('consultant_name') or '',
                    contractor_name=request.POST.get('contractor_name') or '',
                    followup_date=request.POST.get('followup_date'),
                    followup_status=request.POST.get('followup_status'),
                    # Store HRMS user info
                    created_by_user_id=user_info['user_id'],
                    created_by_username=user_info['username'],
                    created_by_email=user_info['email'],
                    created_by_full_name=user_info['full_name'],
                    created_by=None,
                )
                project.save()
                messages.success(request, 'Project Today entry created successfully!')
                
            elif form_type == 'order-expected':
                # Create OrderExpectedNextMonth instance
                order = OrderExpectedNextMonth(
                    region=request.POST.get('region'),
                    from_month=request.POST.get('from_month'),
                    company_name=request.POST.get('company_name'),
                    requirement=request.POST.get('requirement'),
                    location=request.POST.get('location'),
                    contact_person=request.POST.get('contact_person'),
                    contact_no=request.POST.get('contact_no'),
                    ap_quote_price=request.POST.get('ap_quote_price') or None,
                    discounted_price=request.POST.get('discounted_price') or None,
                    total_price=request.POST.get('total_price') or None,
                    last_status_date=request.POST.get('last_status_date'),
                    order_status=request.POST.get('order_status'),
                    expected_in_month=request.POST.get('expected_in_month'),
                    # Store HRMS user info
                    created_by_user_id=user_info['user_id'],
                    created_by_username=user_info['username'],
                    created_by_email=user_info['email'],
                    created_by_full_name=user_info['full_name'],
                    created_by=None,
                )
                order.save()
                messages.success(request, 'Order Expected entry created successfully!')
                
            elif form_type == 'purchase-order':
                # Create MISPurchaseOrder instance
                po = MISPurchaseOrder(
                    person_name=request.POST.get('person_name'),
                    purchase_order_no=request.POST.get('purchase_order_no'),
                    po_date=request.POST.get('po_date'),
                    company_name=request.POST.get('company_name'),
                    location=request.POST.get('location'),
                    contact_person_details=request.POST.get('contact_person_details'),
                    contact_number=request.POST.get('contact_number'),
                    enquiry_log_number=request.POST.get('enquiry_log_number') or '',
                    quote_no=request.POST.get('quote_no') or '',
                    work_order_number=request.POST.get('work_order_number') or '',
                    product_name=request.POST.get('product_name'),
                    capacity=request.POST.get('capacity') or '',
                    model_number=request.POST.get('model_number') or '',
                    machine_details=request.POST.get('machine_details') or '',
                    equipment_sr_number=request.POST.get('equipment_sr_number') or '',
                    po_amount=request.POST.get('po_amount'),
                    ap_quote_price=request.POST.get('ap_quote_price') or None,
                    percentage_order=request.POST.get('percentage_order') or None,
                    # Store HRMS user info
                    created_by_user_id=user_info['user_id'],
                    created_by_username=user_info['username'],
                    created_by_email=user_info['email'],
                    created_by_full_name=user_info['full_name'],
                    created_by=None,
                )
                po.save()
                messages.success(request, 'Purchase Order entry created successfully!')
                
            elif form_type == 'new-data':
                # Create NewData instance
                new_data = NewData(
                    category=request.POST.get('category'),
                    april=request.POST.get('april') or 0,
                    may=request.POST.get('may') or 0,
                    total=request.POST.get('total') or 0,
                    # Store HRMS user info
                    created_by_user_id=user_info['user_id'],
                    created_by_username=user_info['username'],
                    created_by_email=user_info['email'],
                    created_by_full_name=user_info['full_name'],
                    created_by=None,
                )
                new_data.save()
                messages.success(request, 'New Data entry created successfully!')
                
            elif form_type == 'od-plan':
                # Create ODPlan instance
                od_plan = ODPlan(
                    region=request.POST.get('region'),
                    month=request.POST.get('month'),
                    name=request.POST.get('name'),
                    from_date=request.POST.get('from_date'),
                    to_date=request.POST.get('to_date'),
                    location=request.POST.get('location'),
                    total_days=request.POST.get('total_days') or 0,
                    company_visits=request.POST.get('company_visits') or 0,
                    # Store HRMS user info
                    created_by_user_id=user_info['user_id'],
                    created_by_username=user_info['username'],
                    created_by_email=user_info['email'],
                    created_by_full_name=user_info['full_name'],
                    created_by=None,
                )
                od_plan.save()
                messages.success(request, 'OD Plan entry created successfully!')
                
            else:
                messages.error(request, 'Unknown form type. Please try again.')
                
        except Exception as e:
            messages.error(request, f'Error creating entry: {str(e)}')
    
    context = {
        'FOLLOW_UP_STATUS_CHOICES': FollowUpStatus.FOLLOW_UP_STATUS_CHOICES,
        'PHARMA_CATEGORIES': ProjectToday.PHARMA_CATEGORIES,
        'NON_PHARMA_CATEGORIES': ProjectToday.NON_PHARMA_CATEGORIES,
        'ORDER_STATUS_CHOICES': OrderExpectedNextMonth.ORDER_STATUS_CHOICES,
        'CATEGORY_CHOICES': NewData.CATEGORY_CHOICES,
        'PHARMA_CHOICES': NewDataDetails.PHARMA_CHOICES,
        'GROUP_CHOICES': NewDataDetails.GROUP_CHOICES,
    }
    
    return render(request, 'marketing/mis_sheets.html', context)


# OD Plan and Visit Report System Views
@login_required
def od_plan_dashboard(request):
    """OD Plan Dashboard - Main overview of all OD Plan activities"""
    # Get statistics for each sheet
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    visit_reports_count = ODPlanVisitReport.objects.filter(created_by_username=username).count()
    remarks_count = ODPlanRemarks.objects.filter(created_by_username=username).count()
    
    # Recent activities
    recent_reports = ODPlanVisitReport.objects.filter(created_by_username=username).order_by('-created_at')[:5]
    recent_remarks = ODPlanRemarks.objects.filter(created_by_username=username).order_by('-created_at')[:3]
    
    # Status breakdown
    planned_visits = ODPlanVisitReport.objects.filter(created_by_username=username, visit_status='planned').count()
    completed_visits = ODPlanVisitReport.objects.filter(created_by_username=username, visit_status='completed').count()
    cancelled_visits = ODPlanVisitReport.objects.filter(created_by_username=username, visit_status='cancelled').count()
    
    context = {
        'visit_reports_count': visit_reports_count,
        'remarks_count': remarks_count,
        'recent_reports': recent_reports,
        'recent_remarks': recent_remarks,
        'planned_visits': planned_visits,
        'completed_visits': completed_visits,
        'cancelled_visits': cancelled_visits,
    }
    
    return render(request, 'marketing/od_plan_dashboard.html', context)


@login_required
def od_plan_sheets(request):
    """OD Plan Sheets - Tabbed interface for all sheets"""
    context = {
        'REASON_FOR_VISIT_CHOICES': ODPlanVisitReport.REASON_FOR_VISIT_CHOICES,
        'APPOINTMENT_STATUS_CHOICES': ODPlanVisitReport.APPOINTMENT_STATUS_CHOICES,
        'VISIT_STATUS_CHOICES': ODPlanVisitReport.VISIT_STATUS_CHOICES,
        'MAIL_STATUS_CHOICES': ODPlanVisitReport.MAIL_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/od_plan_sheets.html', context)


@login_required
def od_plan_guidelines(request):
    """OD Plan Guidelines - Instructions for using the template"""
    return render(request, 'marketing/od_plan_guidelines.html')


@login_required
def od_plan_visit_report_list(request):
    """List all OD Plan Visit Reports"""
    search_query = request.GET.get('search', '')
    region_filter = request.GET.get('region', '')
    status_filter = request.GET.get('status', '')
    
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Temporarily show all reports to check if any exist without user info
    # TODO: Remove this after fixing existing reports
    reports = ODPlanVisitReport.objects.all()
    
    # Filter by user - but also include reports without created_by_username for now
    reports = reports.filter(
        Q(created_by_username=username) | Q(created_by_username__isnull=True) | Q(created_by_username='')
    )
    
    if search_query:
        reports = reports.filter(
            Q(company_name__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(region__icontains=search_query)
        )
    
    if region_filter:
        reports = reports.filter(region__icontains=region_filter)
    
    if status_filter:
        reports = reports.filter(visit_status=status_filter)
    
    # Pagination
    paginator = Paginator(reports, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'region_filter': region_filter,
        'status_filter': status_filter,
        'VISIT_STATUS_CHOICES': ODPlanVisitReport.VISIT_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/od_plan_visit_report_list.html', context)


@login_required
def od_plan_visit_report_create(request):
    """Create new OD Plan Visit Report"""
    if request.method == 'POST':
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        # Handle form submission
        report = ODPlanVisitReport(
            month=request.POST.get('month'),
            region=request.POST.get('region'),
            date=request.POST.get('date'),
            name=request.POST.get('name'),
            visit_plan=request.POST.get('visit_plan'),
            location=request.POST.get('location'),
            company_name=request.POST.get('company_name'),
            contact_person=request.POST.get('contact_person'),
            contact_no=request.POST.get('contact_no'),
            mail_id=request.POST.get('mail_id'),
            reason_for_visit=request.POST.get('reason_for_visit'),
            appointment_status=request.POST.get('appointment_status'),
            visit_status=request.POST.get('visit_status', 'planned'),
            visited_on_date=request.POST.get('visited_on_date') or None,
            meeting_output=request.POST.get('meeting_output'),
            next_action_needed=request.POST.get('next_action_needed'),
            next_follow_up_visit=request.POST.get('next_follow_up_visit') or None,
            mail_status_about_visit=request.POST.get('mail_status_about_visit', 'not_sent'),
            comments=request.POST.get('comments'),
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
            # Legacy field set to None
            created_by=None,
        )
        report.save()
        messages.success(request, 'OD Plan Visit Report created successfully!')
        return redirect('marketing:od_plan_visit_report_list')
    
    context = {
        'REASON_FOR_VISIT_CHOICES': ODPlanVisitReport.REASON_FOR_VISIT_CHOICES,
        'APPOINTMENT_STATUS_CHOICES': ODPlanVisitReport.APPOINTMENT_STATUS_CHOICES,
        'VISIT_STATUS_CHOICES': ODPlanVisitReport.VISIT_STATUS_CHOICES,
        'MAIL_STATUS_CHOICES': ODPlanVisitReport.MAIL_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/od_plan_visit_report_form.html', context)


@login_required
def od_plan_visit_report_detail(request, pk):
    """View OD Plan Visit Report details"""
    report = get_object_or_404(ODPlanVisitReport, pk=pk)
    
    context = {
        'report': report,
    }
    
    return render(request, 'marketing/od_plan_visit_report_detail.html', context)


@login_required
def od_plan_visit_report_edit(request, pk):
    """Edit OD Plan Visit Report"""
 
    
    if request.method == 'POST':
        report.month = request.POST.get('month')
        report.region = request.POST.get('region')
        report.date = request.POST.get('date')
        report.name = request.POST.get('name')
        report.visit_plan = request.POST.get('visit_plan')
        report.location = request.POST.get('location')
        report.company_name = request.POST.get('company_name')
        report.contact_person = request.POST.get('contact_person')
        report.contact_no = request.POST.get('contact_no')
        report.mail_id = request.POST.get('mail_id')
        report.reason_for_visit = request.POST.get('reason_for_visit')
        report.appointment_status = request.POST.get('appointment_status')
    # Get user info from HRMS session
# Removed - user_info not needed here

        report.visit_status = request.POST.get('visit_status')
        report.visited_on_date = request.POST.get('visited_on_date') or None
        report.meeting_output = request.POST.get('meeting_output')
        report.next_action_needed = request.POST.get('next_action_needed')
        report.next_follow_up_visit = request.POST.get('next_follow_up_visit') or None
        report.mail_status_about_visit = request.POST.get('mail_status_about_visit')
        report.comments = request.POST.get('comments')
        report.save()
        
        messages.success(request, 'OD Plan Visit Report updated successfully!')
        return redirect('marketing:od_plan_visit_report_detail', pk=report.pk)
    
    context = {
        'report': report,
        'REASON_FOR_VISIT_CHOICES': ODPlanVisitReport.REASON_FOR_VISIT_CHOICES,

        'APPOINTMENT_STATUS_CHOICES': ODPlanVisitReport.APPOINTMENT_STATUS_CHOICES,
        'VISIT_STATUS_CHOICES': ODPlanVisitReport.VISIT_STATUS_CHOICES,
        'MAIL_STATUS_CHOICES': ODPlanVisitReport.MAIL_STATUS_CHOICES,
    }
    
    return render(request, 'marketing/od_plan_visit_report_form.html', context)


@login_required
def od_plan_visit_report_delete(request, pk):
    """Delete OD Plan Visit Report"""
 
    
    if request.method == 'POST':
        report.delete()
        messages.success(request, 'OD Plan Visit Report deleted successfully!')
        return redirect('marketing:od_plan_visit_report_list')
    
    # Get user info from HRMS session
# Removed - user_info not needed here

    context = {
        'report': report,
    }
    
    return render(request, 'marketing/od_plan_visit_report_confirm_delete.html', context)


@login_required
def od_plan_remarks_list(request):
    """List all OD Plan Remarks"""
    user_info = get_user_info_dict(request)
    username = user_info['username']
    remarks = ODPlanRemarks.objects.filter(created_by_username=username)
    
    # Pagination
    paginator = Paginator(remarks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'marketing/od_plan_remarks_list.html', context)
    # Get user info from HRMS session
# Removed - user_info not needed here



@login_required
def od_plan_remarks_create(request):
    """Create new OD Plan Remarks"""
    if request.method == 'POST':
        # Get user info from HRMS session
        user_info = get_user_info_dict(request)
        
        remark = ODPlanRemarks(
            remarks=request.POST.get('remarks'),
            # Store HRMS user info
            created_by_user_id=user_info['user_id'],
            created_by_username=user_info['username'],
            created_by_email=user_info['email'],
            created_by_full_name=user_info['full_name'],
            # Legacy field set to None
            created_by=None,
        )
        remark.save()
        messages.success(request, 'OD Plan Remarks added successfully!')
        return redirect('marketing:od_plan_dashboard')
    
    return render(request, 'marketing/od_plan_remarks_form.html')


@login_required
def od_plan_remarks_edit(request, pk):
    """Edit OD Plan Remarks"""
    odplanremarks = get_object_or_404(ODPlanRemarks, pk=pk)
    if request.method == 'POST':
        remark.remarks = request.POST.get('remarks')
        remark.save()
        messages.success(request, 'OD Plan Remarks updated successfully!')
        return redirect('marketing:od_plan_remarks_list')
    
    context = {
        'remark': remark,
    }
    # Get user info from HRMS session
# Removed - user_info not needed here

    
    return render(request, 'marketing/od_plan_remarks_form.html', context)


@login_required
def od_plan_remarks_delete(request, pk):
    """Delete OD Plan Remarks"""
    odplanremarks = get_object_or_404(ODPlanRemarks, pk=pk)
    if request.method == 'POST':
        remark.delete()
        messages.success(request, 'OD Plan Remarks deleted successfully!')
        return redirect('marketing:od_plan_remarks_list')
    
    context = {
        'remark': remark,
    }
    
    return render(request, 'marketing/od_plan_remarks_confirm_delete.html', context)



# Purchase Order Details System Views
@login_required
def po_details_dashboard(request):
    """PO Details Dashboard - Main overview of all PO Details"""
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Get statistics
    po_details_count = PODetails.objects.filter(created_by_username=username).count()
    
    # Recent activities
    recent_po_details = PODetails.objects.filter(created_by_username=username).order_by('-created_at')[:5]
    
    context = {
        'po_details_count': po_details_count,
        'recent_po_details': recent_po_details,
    }
    
    return render(request, 'marketing/po_details_dashboard.html', context)


@login_required
def po_details_sheets(request):
    """PO Details Sheets - Tabbed interface for all sheets"""
    context = {
        'PACKING_FORWARDING_CHOICES': PODetails.PACKING_FORWARDING_CHOICES,
        'TRANSPORTATION_CHOICES': PODetails.TRANSPORTATION_CHOICES,
    }
    
    return render(request, 'marketing/po_details_sheets.html', context)


@login_required
def po_details_list(request):
    """List all PO Details"""
    search_query = request.GET.get('search', '')
    customer_filter = request.GET.get('customer', '')
    
    user_info = get_user_info_dict(request)
    username = user_info['username']
    po_details = PODetails.objects.filter(created_by_username=username)
    
    if search_query:
        po_details = po_details.filter(
            Q(customer_name__icontains=search_query) |
            Q(po_no__icontains=search_query) |
            Q(wo_no__icontains=search_query)
        )
    
    if customer_filter:
        po_details = po_details.filter(customer_name__icontains=customer_filter)
    
    # Pagination
    # Get user info from HRMS session
# Removed - user_info not needed here

    paginator = Paginator(po_details, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'customer_filter': customer_filter,
    }
    
    return render(request, 'marketing/po_details_list.html', context)


@login_required
def po_details_create(request):
    """Create new PO Details"""
    # Get user info from HRMS session
# Removed - user_info not needed here

    if request.method == 'POST':
        po_detail = PODetails(
            customer_name=request.POST.get('customer_name'),
            po_no=request.POST.get('po_no'),
            po_date=request.POST.get('po_date'),
            wo_no=request.POST.get('wo_no'),
            contact_name=request.POST.get('contact_name'),
            contact_details=request.POST.get('contact_details'),
            tel_mob_no=request.POST.get('tel_mob_no'),
            email_id=request.POST.get('email_id'),
            discount=request.POST.get('discount') or None,
            advance_percentage=request.POST.get('advance_percentage') or None,
            against_pi_percentage=request.POST.get('against_pi_percentage') or None,
            against_fat_percentage=request.POST.get('against_fat_percentage') or None,
            after_delivery_percentage=request.POST.get('after_delivery_percentage') or None,
            after_installation_percentage=request.POST.get('after_installation_percentage') or None,
            packing_forwarding=request.POST.get('packing_forwarding'),
            transportation=request.POST.get('transportation'),
            marketing_dept=request.POST.get('marketing_dept', 'Miss. Pooja Kolse'),
            accounts_dept=request.POST.get('accounts_dept', 'Mr. Jitendra Tajanpure'),
            additional_contact=request.POST.get('additional_contact', 'Mr. Harshal Ghoge'),
            # Get user info from HRMS session


            

            # Store HRMS user info

            created_by_user_id=user_info['user_id'],

            created_by_username=user_info['username'],


        )
        po_detail.save()
        messages.success(request, 'PO Details created successfully!')
        return redirect('marketing:po_details_list')
    
    context = {
        'PACKING_FORWARDING_CHOICES': PODetails.PACKING_FORWARDING_CHOICES,
        'TRANSPORTATION_CHOICES': PODetails.TRANSPORTATION_CHOICES,
    # Get user info from HRMS session
# Removed - user_info not needed here

    }
    
    return render(request, 'marketing/po_details_form.html', context)


@login_required
def po_details_detail(request, pk):
    """View PO Details"""
    podetails = get_object_or_404(PODetails, pk=pk)
    context = {
        'po_detail': po_detail,
    }
    
    return render(request, 'marketing/po_details_detail.html', context)


@login_required
def po_details_edit(request, pk):
    # Get user info from HRMS session
# Removed - user_info not needed here


    """Edit PO Details"""
    podetails = get_object_or_404(PODetails, pk=pk)
    if request.method == 'POST':
        po_detail.customer_name = request.POST.get('customer_name')
        po_detail.po_no = request.POST.get('po_no')
        po_detail.po_date = request.POST.get('po_date')
        po_detail.wo_no = request.POST.get('wo_no')
        po_detail.contact_name = request.POST.get('contact_name')
        po_detail.contact_details = request.POST.get('contact_details')
        po_detail.tel_mob_no = request.POST.get('tel_mob_no')
        po_detail.email_id = request.POST.get('email_id')
        po_detail.discount = request.POST.get('discount') or None
        po_detail.advance_percentage = request.POST.get('advance_percentage') or None
        po_detail.against_pi_percentage = request.POST.get('against_pi_percentage') or None
        po_detail.against_fat_percentage = request.POST.get('against_fat_percentage') or None
        po_detail.after_delivery_percentage = request.POST.get('after_delivery_percentage') or None
        po_detail.after_installation_percentage = request.POST.get('after_installation_percentage') or None
        po_detail.packing_forwarding = request.POST.get('packing_forwarding')
        po_detail.transportation = request.POST.get('transportation')
        po_detail.marketing_dept = request.POST.get('marketing_dept', 'Miss. Pooja Kolse')

        po_detail.accounts_dept = request.POST.get('accounts_dept', 'Mr. Jitendra Tajanpure')
        po_detail.additional_contact = request.POST.get('additional_contact', 'Mr. Harshal Ghoge')
        po_detail.save()
        
        messages.success(request, 'PO Details updated successfully!')
        return redirect('marketing:po_details_detail', pk=po_detail.pk)
    
    context = {
        'po_detail': po_detail,
        'PACKING_FORWARDING_CHOICES': PODetails.PACKING_FORWARDING_CHOICES,
        'TRANSPORTATION_CHOICES': PODetails.TRANSPORTATION_CHOICES,
    }
    
    return render(request, 'marketing/po_details_form.html', context)


@login_required
def po_details_delete(request, pk):
    """Delete PO Details"""
    podetails = get_object_or_404(PODetails, pk=pk)
    if request.method == 'POST':
        po_detail.delete()
        messages.success(request, 'PO Details deleted successfully!')
        return redirect('marketing:po_details_list')
    
    context = {
        'po_detail': po_detail,
    }
    
    return render(request, 'marketing/po_details_confirm_delete.html', context)


# Purchase Order Status System Views
@login_required
def po_status_list(request):
    """List all PO Status entries"""
    search_query = request.GET.get('search', '')
    company_filter = request.GET.get('company', '')
    
    user_info = get_user_info_dict(request)
    username = user_info['username']
    po_statuses = POStatus.objects.filter(created_by_username=username)
    
    if search_query:
        po_statuses = po_statuses.filter(
            Q(company__icontains=search_query) |
            Q(po_number__icontains=search_query) |
            Q(responsible_marketing_person__icontains=search_query)
        )
    
    if company_filter:
        po_statuses = po_statuses.filter(company__icontains=company_filter)
    
    # Pagination
    paginator = Paginator(po_statuses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'company_filter': company_filter,
    }
    
    return render(request, 'marketing/po_status_list.html', context)
    # Get user info from HRMS session
# Removed - user_info not needed here



@login_required
def po_status_create(request):
    """Create new PO Status entry"""
    if request.method == 'POST':
        po_status = POStatus(
            month=request.POST.get('month'),
            region=request.POST.get('region'),
            company=request.POST.get('company'),
            order_is_for=request.POST.get('order_is_for'),
            po_number=request.POST.get('po_number'),
            responsible_marketing_person=request.POST.get('responsible_marketing_person'),
            coordinator=request.POST.get('coordinator'),
            po_date=request.POST.get('po_date'),
            po_value_without_gst=request.POST.get('po_value_without_gst'),
    # Get user info from HRMS session
# Removed - user_info not needed here

            gst=request.POST.get('gst'),
            po_acceptance_date=request.POST.get('po_acceptance_date') or None,
            wo_date=request.POST.get('wo_date') or None,
            # PayR-01
            payr01_agreed_percentage=request.POST.get('payr01_agreed_percentage') or None,
            payr01_agreed_amount=request.POST.get('payr01_agreed_amount') or None,
            payr01_received_percentage=request.POST.get('payr01_received_percentage') or None,
            payr01_received_amount=request.POST.get('payr01_received_amount') or None,
            payr01_received_date=request.POST.get('payr01_received_date') or None,
            # PayR-02
            payr02_agreed_percentage=request.POST.get('payr02_agreed_percentage') or None,
            payr02_agreed_amount=request.POST.get('payr02_agreed_amount') or None,
            payr02_received_percentage=request.POST.get('payr02_received_percentage') or None,
            payr02_received_amount=request.POST.get('payr02_received_amount') or None,
            payr02_received_date=request.POST.get('payr02_received_date') or None,
            # PayR-03
            payr03_agreed_percentage=request.POST.get('payr03_agreed_percentage') or None,
            payr03_agreed_amount=request.POST.get('payr03_agreed_amount') or None,
            payr03_received_percentage=request.POST.get('payr03_received_percentage') or None,
            payr03_received_amount=request.POST.get('payr03_received_amount') or None,
            payr03_received_date=request.POST.get('payr03_received_date') or None,
            # PayR-04
            payr04_agreed_percentage=request.POST.get('payr04_agreed_percentage') or None,
            payr04_agreed_amount=request.POST.get('payr04_agreed_amount') or None,
            payr04_received_percentage=request.POST.get('payr04_received_percentage') or None,
            payr04_received_amount=request.POST.get('payr04_received_amount') or None,
            payr04_received_date=request.POST.get('payr04_received_date') or None,
            # PayR-05
            payr05_agreed_percentage=request.POST.get('payr05_agreed_percentage') or None,
            payr05_agreed_amount=request.POST.get('payr05_agreed_amount') or None,
            payr05_received_percentage=request.POST.get('payr05_received_percentage') or None,
            payr05_received_amount=request.POST.get('payr05_received_amount') or None,
            payr05_received_date=request.POST.get('payr05_received_date') or None,
            # Get user info from HRMS session


            





        )
        po_status.save()
        messages.success(request, 'PO Status created successfully!')
        return redirect('marketing:po_status_list')
    
    context = {
        'ORDER_TYPE_CHOICES': POStatus.ORDER_TYPE_CHOICES,
    }
    
    return render(request, 'marketing/po_status_form.html', context)


@login_required
def po_status_detail(request, pk):
    """View PO Status details"""
    postatus = get_object_or_404(POStatus, pk=pk)
    context = {
        'po_status': po_status,
    }
    
    return render(request, 'marketing/po_status_detail.html', context)


@login_required
def po_status_edit(request, pk):
    """Edit PO Status entry"""
    postatus = get_object_or_404(POStatus, pk=pk)
    if request.method == 'POST':
        po_status.month = request.POST.get('month')
        po_status.region = request.POST.get('region')
    # Get user info from HRMS session
# Removed - user_info not needed here

        po_status.company = request.POST.get('company')
        po_status.order_is_for = request.POST.get('order_is_for')
        po_status.po_number = request.POST.get('po_number')
        po_status.responsible_marketing_person = request.POST.get('responsible_marketing_person')
        po_status.coordinator = request.POST.get('coordinator')
        po_status.po_date = request.POST.get('po_date')
        po_status.po_value_without_gst = request.POST.get('po_value_without_gst')
        po_status.gst = request.POST.get('gst')
        po_status.po_acceptance_date = request.POST.get('po_acceptance_date') or None
        po_status.wo_date = request.POST.get('wo_date') or None
        # PayR-01
        po_status.payr01_agreed_percentage = request.POST.get('payr01_agreed_percentage') or None
        po_status.payr01_agreed_amount = request.POST.get('payr01_agreed_amount') or None
        po_status.payr01_received_percentage = request.POST.get('payr01_received_percentage') or None
        po_status.payr01_received_amount = request.POST.get('payr01_received_amount') or None
        po_status.payr01_received_date = request.POST.get('payr01_received_date') or None
        # PayR-02
        po_status.payr02_agreed_percentage = request.POST.get('payr02_agreed_percentage') or None
        po_status.payr02_agreed_amount = request.POST.get('payr02_agreed_amount') or None
        po_status.payr02_received_percentage = request.POST.get('payr02_received_percentage') or None
        po_status.payr02_received_amount = request.POST.get('payr02_received_amount') or None
        po_status.payr02_received_date = request.POST.get('payr02_received_date') or None
        # PayR-03
        po_status.payr03_agreed_percentage = request.POST.get('payr03_agreed_percentage') or None
        po_status.payr03_agreed_amount = request.POST.get('payr03_agreed_amount') or None
    # Get user info from HRMS session
# Removed - user_info not needed here

        po_status.payr03_received_percentage = request.POST.get('payr03_received_percentage') or None
        po_status.payr03_received_amount = request.POST.get('payr03_received_amount') or None
        po_status.payr03_received_date = request.POST.get('payr03_received_date') or None
        # PayR-04
        po_status.payr04_agreed_percentage = request.POST.get('payr04_agreed_percentage') or None
        po_status.payr04_agreed_amount = request.POST.get('payr04_agreed_amount') or None
        po_status.payr04_received_percentage = request.POST.get('payr04_received_percentage') or None
        po_status.payr04_received_amount = request.POST.get('payr04_received_amount') or None
        po_status.payr04_received_date = request.POST.get('payr04_received_date') or None
        # PayR-05
        po_status.payr05_agreed_percentage = request.POST.get('payr05_agreed_percentage') or None
        po_status.payr05_agreed_amount = request.POST.get('payr05_agreed_amount') or None
        po_status.payr05_received_percentage = request.POST.get('payr05_received_percentage') or None
        po_status.payr05_received_amount = request.POST.get('payr05_received_amount') or None
        po_status.payr05_received_date = request.POST.get('payr05_received_date') or None
        po_status.save()
        
        messages.success(request, 'PO Status updated successfully!')
        return redirect('marketing:po_status_detail', pk=po_status.pk)
    
    context = {
        'po_status': po_status,
        'ORDER_TYPE_CHOICES': POStatus.ORDER_TYPE_CHOICES,
    }
    
    return render(request, 'marketing/po_status_form.html', context)


    # Get user info from HRMS session
# Removed - user_info not needed here

@login_required
def po_status_delete(request, pk):
    """Delete PO Status entry"""
    postatus = get_object_or_404(POStatus, pk=pk)
    if request.method == 'POST':
        po_status.delete()
        messages.success(request, 'PO Status deleted successfully!')
        return redirect('marketing:po_status_list')
    
    context = {
        'po_status': po_status,
    }
    
    return render(request, 'marketing/po_status_confirm_delete.html', context)


# PO Status System Dashboard and Sheets Views
@login_required
def po_status_dashboard(request):
    """PO Status Dashboard - Main overview of all PO Status"""

    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Get statistics
    po_status_count = POStatus.objects.filter(created_by_username=username).count()
    
    # Recent activities
    recent_po_status = POStatus.objects.filter(created_by_username=username).order_by('-created_at')[:5]
    
    context = {
        'po_status_count': po_status_count,
        'recent_po_status': recent_po_status,
    }
    
    return render(request, 'marketing/po_status_dashboard.html', context)


@login_required
def po_status_sheets(request):
    """PO Status Sheets - Tabbed interface for all sheets"""
    context = {
        'ORDER_TYPE_CHOICES': POStatus.ORDER_TYPE_CHOICES,
    }
    
    return render(request, 'marketing/po_status_sheets.html', context)


# Work Order System Views
@login_required
def work_order_format_dashboard(request):
    """Work Order Format Dashboard - Main overview of all Work Orders"""
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Get statistics
    work_order_count = WorkOrderFormat.objects.filter(created_by_username=username).count()
    
    # Recent activities
    recent_work_orders = WorkOrderFormat.objects.filter(created_by_username=username).order_by('-created_at')[:5]
    
    context = {
        'work_order_count': work_order_count,
        'recent_work_orders': recent_work_orders,
    }
    
    return render(request, 'marketing/work_order_format_dashboard.html', context)


@login_required
def work_order_format_sheets(request):
    """Work Order Format Sheets - Tabbed interface for all sheets"""
    context = {
        'CONTROLLER_CHOICES': WorkOrderFormat.CONTROLLER_CHOICES,
        'HMI_CHOICES': WorkOrderFormat.HMI_CHOICES,
        'DOOR_ACCESS_CHOICES': WorkOrderFormat.DOOR_ACCESS_CHOICES,
        'HOOTER_CHOICES': WorkOrderFormat.HOOTER_CHOICES,
        'PACKAGING_CHOICES': WorkOrderFormat.PACKAGING_CHOICES,
        'FAT_CHOICES': WorkOrderFormat.FAT_CHOICES,
    }
    
    return render(request, 'marketing/work_order_format_sheets.html', context)


@login_required
def work_order_format_list(request):
    """List all Work Order Format entries"""
    search_query = request.GET.get('search', '')
    equipment_filter = request.GET.get('equipment', '')
    
    user_info = get_user_info_dict(request)
    username = user_info['username']
    work_orders = WorkOrderFormat.objects.filter(created_by_username=username)
    
    if search_query:
        work_orders = work_orders.filter(
            Q(work_order_no__icontains=search_query) |
            Q(equipment_type__icontains=search_query) |
            Q(equipment_no__icontains=search_query)
        )
    
    if equipment_filter:
        work_orders = work_orders.filter(equipment_type__icontains=equipment_filter)
    
    # Pagination
    paginator = Paginator(work_orders, 20)
    page_number = request.GET.get('page')
    # Get user info from HRMS session
# Removed - user_info not needed here

    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'equipment_filter': equipment_filter,
    }
    
    return render(request, 'marketing/work_order_format_list.html', context)


@login_required
def work_order_format_create(request):
    """Create new Work Order Format entry"""
    if request.method == 'POST':
        work_order = WorkOrderFormat(
    # Get user info from HRMS session
# Removed - user_info not needed here

            date=request.POST.get('date'),
            work_order_no=request.POST.get('work_order_no'),
            equipment_no=request.POST.get('equipment_no'),
            delivery_date=request.POST.get('delivery_date'),
            equipment_type=request.POST.get('equipment_type'),
            capacity=request.POST.get('capacity'),
            model=request.POST.get('model'),
            inner_body_moc=request.POST.get('inner_body_moc'),
            outer_body_moc=request.POST.get('outer_body_moc'),
            inside_dimensions=request.POST.get('inside_dimensions'),
            outer_size=request.POST.get('outer_size'),
            temp_range=request.POST.get('temp_range'),
            accuracy=request.POST.get('accuracy'),
            uniformity=request.POST.get('uniformity'),
            controller_system=request.POST.get('controller_system'),
            hmi_system=request.POST.get('hmi_system'),
            gsm_system=request.POST.get('gsm_system') == 'on',
            scanner=request.POST.get('scanner', ''),
            software=request.POST.get('software', ''),
            door_access_system=request.POST.get('door_access_system'),
            hooter_system=request.POST.get('hooter_system'),
            castor_wheels=request.POST.get('castor_wheels') == 'on',
            pipe_length=request.POST.get('pipe_length', ''),
            packaging=request.POST.get('packaging'),
            fat=request.POST.get('fat'),
            refrigeration_system=request.POST.get('refrigeration_system'),
            sg_system=request.POST.get('sg_system'),
            sensor=request.POST.get('sensor'),
            tray_qty=request.POST.get('tray_qty'),
            tray_type=request.POST.get('tray_type'),
            tray_moc=request.POST.get('tray_moc'),
            tray_dimension=request.POST.get('tray_dimension'),
            rack_qty=request.POST.get('rack_qty'),
            protocol_documents=request.POST.get('protocol_documents') == 'on',
            calibration_duration=request.POST.get('calibration_duration'),
            validation_duration=request.POST.get('validation_duration'),
            validation_compressor=request.POST.get('validation_compressor'),
            extra_validation_charge=request.POST.get('extra_validation_charge') == 'on',
            calibration_probes=request.POST.get('calibration_probes'),
            plc_validation=request.POST.get('plc_validation') == 'on',
            training_handover=request.POST.get('training_handover') == 'on',
            special_instructions=request.POST.get('special_instructions', ''),
            advance_percentage=request.POST.get('advance_percentage'),
            against_pi_percentage=request.POST.get('against_pi_percentage'),
            after_material_percentage=request.POST.get('after_material_percentage'),
            # Get user info from HRMS session


            





        )
        work_order.save()
        messages.success(request, 'Work Order created successfully!')
        return redirect('marketing:work_order_format_list')
    
    context = {
        'CONTROLLER_CHOICES': WorkOrderFormat.CONTROLLER_CHOICES,
        'HMI_CHOICES': WorkOrderFormat.HMI_CHOICES,
        'DOOR_ACCESS_CHOICES': WorkOrderFormat.DOOR_ACCESS_CHOICES,
        'HOOTER_CHOICES': WorkOrderFormat.HOOTER_CHOICES,
        'PACKAGING_CHOICES': WorkOrderFormat.PACKAGING_CHOICES,
        'FAT_CHOICES': WorkOrderFormat.FAT_CHOICES,
    }
    
    # Get user info from HRMS session
# Removed - user_info not needed here

    return render(request, 'marketing/work_order_format_form.html', context)


@login_required
def work_order_format_detail(request, pk):
    """View Work Order Format details"""
    workorderformat = get_object_or_404(WorkOrderFormat, pk=pk)
    context = {
        'work_order': work_order,
    }
    
    return render(request, 'marketing/work_order_format_detail.html', context)


@login_required
def work_order_format_edit(request, pk):
    """Edit Work Order Format entry"""
    workorderformat = get_object_or_404(WorkOrderFormat, pk=pk)
    if request.method == 'POST':
        work_order.date = request.POST.get('date')
        work_order.work_order_no = request.POST.get('work_order_no')
        work_order.equipment_no = request.POST.get('equipment_no')
        work_order.delivery_date = request.POST.get('delivery_date')
        work_order.equipment_type = request.POST.get('equipment_type')
        work_order.capacity = request.POST.get('capacity')
        work_order.model = request.POST.get('model')
        work_order.inner_body_moc = request.POST.get('inner_body_moc')
        work_order.outer_body_moc = request.POST.get('outer_body_moc')
        work_order.inside_dimensions = request.POST.get('inside_dimensions')
        work_order.outer_size = request.POST.get('outer_size')
        work_order.temp_range = request.POST.get('temp_range')
        work_order.accuracy = request.POST.get('accuracy')
        work_order.uniformity = request.POST.get('uniformity')
        work_order.controller_system = request.POST.get('controller_system')
        work_order.hmi_system = request.POST.get('hmi_system')
        work_order.gsm_system = request.POST.get('gsm_system') == 'on'
        work_order.scanner = request.POST.get('scanner', '')
        work_order.software = request.POST.get('software', '')
        work_order.door_access_system = request.POST.get('door_access_system')
        work_order.hooter_system = request.POST.get('hooter_system')
        work_order.castor_wheels = request.POST.get('castor_wheels') == 'on'
        work_order.pipe_length = request.POST.get('pipe_length', '')
        work_order.packaging = request.POST.get('packaging')
        work_order.fat = request.POST.get('fat')
        work_order.refrigeration_system = request.POST.get('refrigeration_system')
        work_order.sg_system = request.POST.get('sg_system')
        work_order.sensor = request.POST.get('sensor')
        work_order.tray_qty = request.POST.get('tray_qty')
        work_order.tray_type = request.POST.get('tray_type')
        work_order.tray_moc = request.POST.get('tray_moc')
        work_order.tray_dimension = request.POST.get('tray_dimension')
        work_order.rack_qty = request.POST.get('rack_qty')
        work_order.protocol_documents = request.POST.get('protocol_documents') == 'on'
        work_order.calibration_duration = request.POST.get('calibration_duration')
        work_order.validation_duration = request.POST.get('validation_duration')
        work_order.validation_compressor = request.POST.get('validation_compressor')
        work_order.extra_validation_charge = request.POST.get('extra_validation_charge') == 'on'
        work_order.calibration_probes = request.POST.get('calibration_probes')
        work_order.plc_validation = request.POST.get('plc_validation') == 'on'
        work_order.training_handover = request.POST.get('training_handover') == 'on'
        work_order.special_instructions = request.POST.get('special_instructions', '')
        work_order.advance_percentage = request.POST.get('advance_percentage')
        work_order.against_pi_percentage = request.POST.get('against_pi_percentage')
        work_order.after_material_percentage = request.POST.get('after_material_percentage')
        work_order.save()
        
        messages.success(request, 'Work Order updated successfully!')
        return redirect('marketing:work_order_format_detail', pk=work_order.pk)
    
    context = {
        'work_order': work_order,
        'CONTROLLER_CHOICES': WorkOrderFormat.CONTROLLER_CHOICES,
        'HMI_CHOICES': WorkOrderFormat.HMI_CHOICES,
        'DOOR_ACCESS_CHOICES': WorkOrderFormat.DOOR_ACCESS_CHOICES,
        'HOOTER_CHOICES': WorkOrderFormat.HOOTER_CHOICES,
        'PACKAGING_CHOICES': WorkOrderFormat.PACKAGING_CHOICES,
        'FAT_CHOICES': WorkOrderFormat.FAT_CHOICES,
    }
    
    return render(request, 'marketing/work_order_format_form.html', context)


@login_required
def work_order_format_delete(request, pk):
    """Delete Work Order Format entry"""
    workorderformat = get_object_or_404(WorkOrderFormat, pk=pk)
    if request.method == 'POST':
        work_order.delete()
        messages.success(request, 'Work Order deleted successfully!')
        return redirect('marketing:work_order_format_list')
    
    context = {
        'work_order': work_order,
    }
    
    return render(request, 'marketing/work_order_format_confirm_delete.html', context)


# Weekly Status Report System Views
@login_required
def wsr_dashboard(request):
    """WSR Dashboard - Main overview of all Weekly Status Reports"""
    user_info = get_user_info_dict(request)
    username = user_info['username']
    
    # Get statistics
    weekly_summary_count = WeeklySummary.objects.filter(created_by_username=username).count()
    hot_orders_count = HotOrders.objects.filter(created_by_username=username).count()
    calling_details_count = CallingDetails.objects.filter(created_by_username=username).count()
    dsr_count = DSR.objects.filter(created_by_username=username).count()
    
    # Recent activities
    recent_weekly_summaries = WeeklySummary.objects.filter(created_by_username=username).order_by('-created_at')[:5]
    
    context = {
        'weekly_summary_count': weekly_summary_count,
        'hot_orders_count': hot_orders_count,
        'calling_details_count': calling_details_count,
        'dsr_count': dsr_count,
        'recent_weekly_summaries': recent_weekly_summaries,
    }
    
    return render(request, 'marketing/wsr_dashboard.html', context)


@login_required
def wsr_sheets(request):
    """WSR Sheets - Tabbed interface for all 8 sheets"""
    context = {
        'REGION_CHOICES': WeeklySummary.REGION_CHOICES,
        'PRODUCT_CHOICES': WeeklySummary.PRODUCT_CHOICES,
        'STATUS_CHOICES': HotOrders.STATUS_CHOICES,
        'TEAM_CHOICES': DSR.TEAM_CHOICES,
    }
    
    return render(request, 'marketing/wsr_sheets.html', context)



# Dashboard Detail Views
@login_required
def sales_pipeline_details(request):
    """Sales Pipeline Details View"""
    return render(request, 'marketing/sales_pipeline_details.html')

@login_required
def regional_performance_details(request):
    """Regional Performance Details View"""
    from datetime import timedelta
    
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    
    # Get all regions
    regions = Region.objects.all().order_by('name')
    
    # Calculate overall metrics
    total_revenue = PurchaseOrder.objects.filter(
        created_at__date__gte=first_day_of_month
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    total_customers = Customer.objects.count()
    new_customers_this_month = Customer.objects.filter(
        created_at__date__gte=first_day_of_month
    ).count()
    
    # Calculate regional performance
    regional_data = []
    total_target = Decimal('0')
    total_achieved = Decimal('0')
    
    for region in regions:
        region_target = region.monthly_target or Decimal('0')
        total_target += region_target
        
        # Get sales for this region this month
        region_sales = PurchaseOrder.objects.filter(
            customer__region=region,
            created_at__date__gte=first_day_of_month
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        total_achieved += region_sales
        
        # Calculate achievement percentage
        achievement_pct = (region_sales / region_target * 100) if region_target > 0 else 0
        
        # Get customer count for this region
        region_customers = Customer.objects.filter(region=region).count()
        
        # Determine status
        if achievement_pct >= 90:
            status = 'Outstanding'
            status_class = 'bg-green-100 text-green-800'
        elif achievement_pct >= 80:
            status = 'On Track'
            status_class = 'bg-green-100 text-green-800'
        elif achievement_pct >= 60:
            status = 'Good'
            status_class = 'bg-yellow-100 text-yellow-800'
        else:
            status = 'Needs Attention'
            status_class = 'bg-red-100 text-red-800'
        
        regional_data.append({
            'region': region,
            'target': region_target,
            'achieved': region_sales,
            'achievement_pct': round(achievement_pct, 1),
            'customers': region_customers,
            'team_size': 0,  # Would need Employee/Team model
            'status': status,
            'status_class': status_class,
        })
    
    # Find best performing region
    best_region = max(regional_data, key=lambda x: x['achievement_pct']) if regional_data else None
    
    # Calculate average achievement
    avg_achievement = sum(r['achievement_pct'] for r in regional_data) / len(regional_data) if regional_data else 0
    
    # Calculate overall achievement percentage
    overall_achievement = (total_achieved / total_target * 100) if total_target > 0 else 0
    
    context = {
        'total_revenue': total_revenue,
        'best_region': best_region,
        'avg_achievement': round(avg_achievement, 1),
        'total_customers': total_customers,
        'new_customers_this_month': new_customers_this_month,
        'regional_data': regional_data,
        'overall_achievement': round(overall_achievement, 1),
    }
    
    return render(request, 'marketing/regional_performance_details.html', context)

@login_required
def recent_activities_details(request):
    """Recent Activities Details View"""
    return render(request, 'marketing/recent_activities_details.html')

@login_required
def alerts_notifications_details(request):
    """Alerts & Notifications Details View"""
    return render(request, 'marketing/alerts_notifications_details.html')